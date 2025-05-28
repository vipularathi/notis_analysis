import re
import pandas as pd
from datetime import datetime, timedelta, timezone
import os
import progressbar
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pyodbc
from sqlalchemy import create_engine, text, insert
import psycopg2
import time
import csv
import paramiko
import logging
from logging.handlers import TimedRotatingFileHandler
import sys
from db_config import (engine_str,
                       n_tbl_notis_trade_book, s_tbl_notis_trade_book,
                       n_tbl_notis_raw_data, s_tbl_notis_raw_data,
                       n_tbl_notis_nnf_data, s_tbl_notis_nnf_data)

holidays_25 = ['2025-02-26', '2025-03-14', '2025-03-31', '2025-04-10', '2025-04-14', '2025-04-18', '2025-05-01', '2025-08-15', '2025-08-27', '2025-10-02', '2025-10-21', '2025-10-22', '2025-11-05', '2025-12-25']
# holidays_25.append('2024-03-20') #add unusual holidays
today = datetime.now().date()
# today = datetime(year=2025, month=2, day=18).date()
# yesterday = today-timedelta(days=1)
b_days = pd.bdate_range(start=today-timedelta(days=7), end=today, freq='C', weekmask='1111100', holidays=holidays_25).date.tolist()
# b_days = b_days.append(pd.DatetimeIndex([pd.Timestamp(year=2024, month=1, day=20)])) #add unusual trading days
today, yesterday = sorted(b_days)[-1], sorted(b_days)[-2]

root_dir = os.path.dirname(os.path.abspath(__file__))
bhav_dir = os.path.join(root_dir, 'bhavcopy')
modified_dir = os.path.join(root_dir, 'modified_data')
table_dir = os.path.join(root_dir, 'table_data')
bse_dir = os.path.join(root_dir, 'bse_data')
logs_dir = os.path.join(root_dir, 'logs')
volt_dir = os.path.join(root_dir, 'nse_fo_voltality_file')
nnf_file_path = os.path.join(root_dir, "Final_NNF_ID.xlsx")
eod_test_dir = os.path.join(root_dir, 'eod_testing')
eod_input_dir = os.path.join(root_dir, 'eod_original')
eod_output_dir = os.path.join(root_dir, 'eod_data')
test_dir = os.path.join(root_dir, 'testing')
eod_net_pos_input_dir = os.path.join(root_dir, 'test_net_position_original')
eod_net_pos_output_dir = os.path.join(root_dir, 'test_net_position_code')
zipped_dir = os.path.join(root_dir, 'zipped_files')

dir_list = [bhav_dir, modified_dir, table_dir, eod_input_dir, bse_dir, logs_dir, volt_dir, zipped_dir, test_dir, logs_dir]
status = [os.makedirs(_dir, exist_ok=True) for _dir in dir_list if not os.path.exists(_dir)]

engine = create_engine(engine_str, pool_size = 20, max_overflow = 10, pool_pre_ping=True, pool_recycle=900)

def define_logger():
    # Logging Definitions
    log_lvl = logging.DEBUG
    console_log_lvl = logging.INFO
    _logger = logging.getLogger('arathi')
    # logger.setLevel(log_lvl)
    _logger.setLevel(console_log_lvl)
    log_file = os.path.join(logs_dir, f'logs_arathi_{datetime.now().strftime("%Y%m%d")}.log')
    handler = TimedRotatingFileHandler(log_file, when='D', delay=True)
    handler.setLevel(log_lvl)
    console = logging.StreamHandler(stream=sys.stdout)
    console.setLevel(console_log_lvl)
    formatter = logging.Formatter('%(asctime)s %(levelname)s <%(funcName)s> %(message)s')
    handler.setFormatter(formatter)
    console.setFormatter(formatter)
    _logger.addHandler(handler)  # Comment to disable file logs
    _logger.addHandler(console)
    # logger.propagate = False  # Removes AWS Level Logging as it tracks root propagation as well
    return _logger
def read_data_db(nnf=False, for_table='ENetMIS', from_time:str='', to_time:str='', from_source=False):
    global engine
    if not nnf and for_table == 'ENetMIS':
        # Sql connection parameters
        sql_server = "rms.ar.db"
        sql_database = "ENetMIS"
        sql_username = "notice_user"
        sql_password = "Notice@2024"
        if not from_time:
            sql_query = "SELECT * FROM [ENetMIS].[dbo].[NSE_FO_AA100_view]"
        else:
            sql_query = f"SELECT * FROM [ENetMIS].[dbo].[NSE_FO_AA100_view] WHERE CreateDate BETWEEN '{from_time}' AND '{to_time}';"
        try:
            sql_connection_string = (
                f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                f"SERVER={sql_server};"
                f"DATABASE={sql_database};"
                f"UID={sql_username};"
                f"PWD={sql_password}"
            )
            with pyodbc.connect(sql_connection_string) as sql_conn:
                df = pd.read_sql_query(sql_query, sql_conn)
            logger.info(f"Data fetched from SQL Server. Shape:{df.shape}")
            return df
        except (pyodbc.Error, psycopg2.Error) as e:
            logger.info("Error occurred:", e)
    elif nnf and for_table != 'ENetMIS':
        # engine = create_engine(engine_str)
        with engine.begin() as conn:
            df = pd.read_sql_table(n_tbl_notis_nnf_data, con=conn)
        logger.info(f"Data fetched from {for_table} table. Shape:{df.shape}")
        return df
    elif not nnf and for_table == 'TradeHist':
        sql_server = '172.30.100.41'
        sql_port = '1450'
        sql_db = 'OMNE_ARD_PRD'
        sql_userid = 'Pos_User'
        sql_paswd = 'Pass@Word1'

        if not from_time:
            logger.info(f'Fetching today\'s BSE trade data till now.')
            sql_query = (
                f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
                f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker, mnmExchangeTime "
                f"from [OMNE_ARD_PRD].[dbo].[TradeHist] "
                f"where mnmExchSeg = 'bse_fo' "
                f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
            sql_query2 = (
                f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
                f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker, mnmExchangeTime "
                f"from [OMNE_ARD_PRD_HNI].[dbo].[TradeHist] "
                f"where mnmExchSeg = 'bse_fo' "
                f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
        else:
            logger.info(f'Fetching BSE trade data from {from_time} to {to_time}')
            sql_query = (
                f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
                f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker, mnmExchangeTime "
                f"from [OMNE_ARD_PRD].[dbo].[TradeHist] "
                f"where mnmExchSeg = 'bse_fo' "
                f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
                f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
            sql_query2 = (
                f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
                f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker, mnmExchangeTime "
                f"from [OMNE_ARD_PRD_HNI].[dbo].[TradeHist] "
                f"where mnmExchSeg = 'bse_fo' "
                f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
                f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
        try:
            sql_engine_str = (
                f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                f"SERVER={sql_server},{sql_port};"
                f"DATABASE={sql_db};"
                f"UID={sql_userid};"
                f"PWD={sql_paswd};"
            )
            with pyodbc.connect(sql_engine_str) as sql_conn:
                df_bse = pd.read_sql_query(sql_query, sql_conn)
                df_bse_hni = pd.read_sql_query(sql_query2,sql_conn)
            logger.info(f'data fetched for bse: {df_bse.shape, df_bse_hni.shape}')
            final_bse_df = pd.concat([df_bse,df_bse_hni], ignore_index=True)
            return final_bse_df
        except (pyodbc.Error, psycopg2.Error) as e:
            logger.info(f'Error in fetching data: {e}')
    elif not nnf and for_table!='ENetMIS':
        with engine.begin() as conn:
            df = pd.read_sql_table(for_table, con=conn)
        logger.info(f"Data fetched from {for_table} table. Shape:{df.shape}")
        return df
    elif not nnf and for_table == 'Source_2':
        sql_server = '172.30.100.40'
        sql_port = '1450'
        sql_db = 'OMNE_ARD_PRD_AA100_3.19'
        sql_userid = 'Pos_User'
        sql_paswd = 'Pass@Word1'
        if not from_time:
            logger.info(f'Fetching today\'s NSE&BSE trades from Source:2 till now.')
            sql_query = (
                f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
                f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
                f"from [OMNE_ARD_PRD_3.19].[dbo].[TradeHist] "
                f"where mnmSymbolName in ('NIFTY','BANKNIFTY','MIDCPNIFTY','FINNIFTY','SENSEX') "
                f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
            sql_query2 = (
                f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
                f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
                f"from [OMNE_ARD_PRD_AA100_3.19].[dbo].[TradeHist] "
                f"where mnmSymbolName in ('NIFTY','BANKNIFTY','MIDCPNIFTY','FINNIFTY','SENSEX') "
                f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
        else:
            logger.info(f'Fetching NSE&BSE trade data(Source:2) from {from_time} to {to_time}')
            sql_query = (
                f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
                f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
                f"from [OMNE_ARD_PRD_3.19].[dbo].[TradeHist] "
                f"where mnmSymbolName in ('NIFTY','BANKNIFTY','MIDCPNIFTY','FINNIFTY','SENSEX') "
                f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
                f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
            sql_query2 = (
                f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
                f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
                f"from [OMNE_ARD_PRD_AA100_3.19].[dbo].[TradeHist] "
                f"where mnmSymbolName in ('NIFTY','BANKNIFTY','MIDCPNIFTY','FINNIFTY','SENSEX') "
                f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
                f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
        try:
            sql_engine_str = (
                f"DRIVER={{ODBC Driver 17 for SQL Server}};"
                f"SERVER={sql_server},{sql_port};"
                f"DATABASE={sql_db};"
                f"UID={sql_userid};"
                f"PWD={sql_paswd};"
            )
            with pyodbc.connect(sql_engine_str) as sql_conn:
                df_bse = pd.read_sql_query(sql_query, sql_conn)
                df_bse_hni = pd.read_sql_query(sql_query2,sql_conn)
            logger.info(f'data fetched for bse: {df_bse.shape, df_bse_hni.shape}')
            final_bse_df = pd.concat([df_bse,df_bse_hni], ignore_index=True)
            return final_bse_df
        except (pyodbc.Error, psycopg2.Error) as e:
            logger.info(f'Error in fetching data: {e}')

def read_notis_file(filepath):
    wb = load_workbook(filepath, read_only=True)
    sheet = wb.active
    total_rows = sheet.max_row
    logger.info('Reading Notis file...')
    pbar = progressbar.ProgressBar(max_value=total_rows, widgets=[progressbar.Percentage(), ' ',
                                                           progressbar.Bar(marker='=', left='[', right=']'),
                                                           progressbar.ETA()])

    data = []
    pbar.update(0)
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        data.append(row)
        pbar.update(i)
    pbar.finish()

    df = pd.DataFrame(data[1:], columns=data[0])
    logger.info('Notis file read')
    return df

def read_file(filepath):
    file_extension = os.path.splitext(filepath)[-1].lower()
    data = []
    if file_extension == '.xlsx':
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active
        total_rows = sheet.max_row
        logger.info('Reading Excel file...')
        pbar = progressbar.ProgressBar(max_value=total_rows, widgets=[progressbar.Percentage(), ' ',
                                                               progressbar.Bar(marker='=', left='[', right=']'),
                                                               progressbar.ETA()])

        pbar.update(0)
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            data.append(row)
            pbar.update(i)
        pbar.finish()
        # df = pd.DataFrame(data[1:], columns=data[0])
        logger.info('Excel file read')

    elif file_extension == '.csv':
        total_rows = sum(1 for _ in open(filepath, 'r', encoding='utf-8')) - 1
        logger.info(f'Reading CSV file...')
        pbar = progressbar.ProgressBar(max_value=total_rows, widgets=[progressbar.Percentage(), ' ',
                                                               progressbar.Bar(marker='=', left='[', right=']'),
                                                               progressbar.ETA()])
        pbar.update(0)
        with open(filepath, 'r', encoding='utf-8') as csv_file:
            for i, row in enumerate(csv_file, start=0):
                data.append(row.strip().split(','))
                pbar.update(i)
        pbar.finish()
        logger.info('CSV file read')
    df = pd.DataFrame(data[1:], columns=data[0])
    return df

def write_notis_postgredb(df, table_name, raw=False, truncate_required=False):
    start_time = time.time()
    if truncate_required:
        truncate_tables(table_name)
    logger.info(f'Writing {"Raw" if raw else "Modified"} data to database...')
    total_rows = len(df)
    if not raw:
        for col in df.columns:
            if type(df[col][0]) == type(pd.to_datetime('2025-04-04').date()):
                print(f'common changing col- {col}')
                df[col] = pd.to_datetime(df[col], dayfirst=True, format='mixed').dt.strftime('%d/%m/%Y')
    pbar = progressbar.ProgressBar(max_value=total_rows, widgets=[
        progressbar.Percentage(), ' ',
        progressbar.Bar(marker='=', left='[', right=']'),
        progressbar.ETA()
    ])

    if raw:
        list_str_int64 = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 16, 17, 18, 19, 21, 23, 27, 28]
        list_str_none = [15, 20, 25, 30, 31, 32, 33, 34, 35, 36, 37]
        list_none_str = [38]
        for i in list_str_int64:
            # df_db.loc[:, f'Column{i}'] = df_db.loc[:, f'Column{i}'].astype('int64')
            column_name = f'Column{i}'
            df[f'Column{i}'] = df[f'Column{i}'].astype('int64')
        for i in list_str_none:
            df[f'Column{i}'] = None
        for i in list_none_str:
            df[f'Column{i}'] = df[f'Column{i}'].astype('str')
    chunk_size = 1000
    for i in range(0, total_rows, chunk_size):
        chunk = df.iloc[i:i + chunk_size]
        with engine.begin() as conn:
            chunk.to_sql(table_name, conn, index=False, if_exists='append', method='multi')
        pbar.update(min(i + chunk_size, total_rows))

    pbar.finish()
    logger.info(f'Data successfully inserted in table => {table_name}')
    end_time = time.time()
    logger.info(f'Total time taken: {end_time - start_time} seconds')

def write_notis_data(df, filepath):
    file_extention = os.path.splitext(filepath)[-1].lower()
    if file_extention == '.xlsx':
        logger.info('Writing Notis file to excel...')
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        rows = list(dataframe_to_rows(df, index=False, header=True))
        total_rows = len(rows)
        pbar = progressbar.ProgressBar(max_value=total_rows, widgets=[progressbar.Percentage(), ' ',
                                                                      progressbar.Bar(marker='=', left='[', right=']'),
                                                                      progressbar.ETA()])
        for i, row in enumerate(rows, start=1):
            ws.append(row)
            pbar.update(i)
        pbar.finish()
        logger.info('Saving the file...')
        wb.save(filepath)
    elif file_extention == '.csv':
        logger.info('Writing Notis file to CSV...')
        total_rows = len(df)
        pbar = progressbar.ProgressBar(max_value=total_rows,widgets=[progressbar.Percentage(),' ',progressbar.Bar(marker='=',left='[',right=']'),progressbar.ETA()])
        pbar.update(0)
        with open(filepath,mode='w',encoding='utf-8',newline='') as f:
            writer = csv.writer(f)
            writer.writerow(df.columns)
            for row_num, row in enumerate(df.itertuples(index=False, name=None), start=1):
                writer.writerow(row)
                pbar.update(row_num)
        pbar.finish()
        logger.info('Saving the file...')

def get_date_from_jiffy(dt_val):
    """
    Converts the Jiffy format date to a readable format.
    :param dt_val: long
    :return: long (epoch time in seconds)
    """
    # Jiffy is 1/65536 of a second since Jan 1, 1980
    base_date = datetime(1980, 1, 1, tzinfo=timezone.utc)
    date_time = int((base_date.timestamp() + (dt_val / 65536)))
    new_date = datetime.fromtimestamp(date_time, timezone.utc)
    formatted_date = new_date.astimezone(timezone(timedelta(hours=5, minutes=30))).strftime("%Y-%m-%d %I:%M:%S")
    return formatted_date

def get_date_from_jiffy_new(dt_val):
    """
    Converts the Jiffy format date to a readable format.
    :param dt_val: long
    :return: long (epoch time in seconds)
    """
    # Jiffy is 1/65536 of a second since Jan 1, 1980
    # base_date = datetime(1980, 1, 1, tzinfo=timezone.utc)
    # date_time = int((base_date.timestamp() + (dt_val / 65536)))
    # new_date = datetime.fromtimestamp(date_time, timezone.utc)
    # formatted_date = new_date.astimezone(timezone(timedelta(hours=5, minutes=30))).strftime("%Y-%m-%d %I:%M:%S")
    formatted_date = int(dt_val/65536) + 315513000
    return formatted_date

def get_date_from_non_jiffy(dt_val):
    """
    Converts the 1980 format date time to a readable format.
    :param dt_val: long
    :return: long (epoch time in seconds)
    :sample: 1742322599
    """
    # Assuming dt_val is seconds since Jan 1, 1980
    base_date = datetime(1980, 1, 1, tzinfo=timezone.utc)
    if (type(dt_val) == str):
        dt_val = int(dt_val)
    # date_time = int(base_date.timestamp() + dt_val)
    date_time = base_date.timestamp() + dt_val
    new_date = datetime.fromtimestamp(date_time, timezone.utc)
    formatted_date = new_date.astimezone(timezone(timedelta(hours=5, minutes=30))).strftime("%Y-%m-%d %I:%M:%S")
    return formatted_date

# Not working for some cases correctly
def get_date_from_non_jiffy_new(dt_val):
    """
    Converts the 1980 format date time to a readable format.
    :param dt_val: long
    :return: long (epoch time in seconds)
    :sample: 1742322599
    """
    # Assuming dt_val is seconds since Jan 1, 1980
    base_date_1980 = datetime(1980, 1, 1, tzinfo=timezone.utc).timestamp()
    base_date_1970 = datetime(1970, 1, 1, tzinfo=timezone.utc).timestamp()
    if (type(dt_val) == str):
        dt_val = int(dt_val)
    if dt_val > base_date_1970:
        calc_time = dt_val
    else:
        calc_time = base_date_1980 + dt_val
    # date_time = base_date.timestamp() + dt_val
    new_date = datetime.fromtimestamp(calc_time, timezone.utc)
    formatted_date = new_date.astimezone(timezone(timedelta(hours=5, minutes=30))).strftime("%Y-%m-%d %I:%M:%S")
    return formatted_date

def download_bhavcopy():
    host = '192.168.112.81'
    username = 'greek'
    password = 'greeksoft'
    filename = f"regularNSEBhavcopy_{today.strftime('%d%m%Y')}.csv"  # sample=regularBhavcopy_13022025
    remote_path = rf'/home/greek/NSE_BSE_Broadcast/NSE/Bhavcopy/Files/{filename}'
    local_path = os.path.join(bhav_dir, filename)
    try:
        transport = paramiko.Transport((host, 22))
        transport.connect(username=username, password=password)
        sftp = paramiko.SFTPClient.from_transport(transport)
        sftp.get(remote_path, local_path)
        sftp.close()
        transport.close()
    except Exception as e:
        logger.info(f'Error: {e}')

def truncate_tables(tablename):
    # engine = create_engine(engine_str)
    with engine.begin() as conn:
        res = conn.execute(text(f'select count(*) from "{tablename}"'))
        row_count = res.scalar()
        if row_count > 0:
            conn.execute(text(f'truncate table "{tablename}"'))
            logger.info(f'Existing data from table {tablename} deleted')
        else:
            logger.info(f'No data in table {tablename}, no need to delete')

def revise_eod_net_pos(for_dt:str = '', modify_sensex:bool=False):# modify_sensex=True if only sensex data is to be revised
    final_eod_df = pd.DataFrame()
    for_date = pd.to_datetime(for_dt).date()
    if not for_date:
        print(f'No date entered, hence exiting.')
        return
    else:
        sent_df = read_file(
            rf"D:\notis_analysis\eod_original\EOD Net positions {for_date.strftime('%d%m%Y')} BSE.xlsx")
        sent_df.columns = [re.sub(rf'\s|\.', '', each) for each in sent_df.columns]
        sent_df.ExpiryDate = pd.to_datetime(sent_df.ExpiryDate, dayfirst=True, format='mixed').dt.date
        sent_df['Broker'] = sent_df.apply(lambda row: 'CP' if row['PartyCode'].upper().endswith('CP') else 'non CP',
                                          axis=1)
        sent_df['OptionType'] = sent_df.apply(
            lambda row: 'XX' if row['OptionType'].upper().startswith('F') else row['OptionType'], axis=1)
        sent_df.drop(columns=['PartyCode'], inplace=True)
        sent_df.rename(columns={'Symbol': 'Underlying', 'ExpiryDate': 'Expiry', 'StrikePrice': 'Strike'},
                       inplace=True)
        sent_df = sent_df.add_prefix('Eod')
        sent_df.rename(columns={'EodNetQty': 'FinalNetQty'}, inplace=True)
        col_to_add = ['EodNetQuantity', 'EodClosingPrice', 'buyQty', 'buyAvgPrice', 'sellQty', 'sellAvgPrice',
                      'IntradayVolume', 'FinalSettlementPrice']
        for col in col_to_add:
            sent_df[col] = 0
        final_eod_df = sent_df.copy()
        # if only sensex data is to be revised, modify_sensex=True
        if modify_sensex:
            truncated_sent_df = sent_df.query('EodUnderlying == "SENSEX"')
            eod_df = read_data_db(for_table=f'NOTIS_EOD_NET_POS_CP_NONCP_{for_date}')
            eod_df.EodExpiry = pd.to_datetime(eod_df.EodExpiry, dayfirst=True, format='mixed').dt.date
            concat_eod_df = pd.concat([eod_df, truncated_sent_df], ignore_index=True)
            final_eod_df = concat_eod_df.copy()
        write_notis_postgredb(df=final_eod_df,table_name=f'NOTIS_EOD_NET_POS_CP_NONCP_{for_date}',truncate_required=True)
        p=0

logger = define_logger()