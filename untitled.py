import numpy as np
import pandas as pd
import os
import time
from datetime import datetime, timedelta, timezone, date
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import progressbar
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session, sessionmaker
from urllib.parse import quote
import pytz
import gzip
import xlsxwriter
import warnings
from fastapi import FastAPI, Query, status, Depends
from fastapi.encoders import jsonable_encoder
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
import uvicorn
import openpyxl
import io
import gzip
import progressbar
import csv
import sys
import re
import pyodbc
import psycopg2
from common import test_dir
# today = datetime(year=2025, month=1, day=24).date()
# yesterday = datetime(year=2025, month=1, day=23).date()
pd.set_option('display.max_columns', None)
warnings.filterwarnings('ignore')
# holidays_25 = ['2025-02-26', '2025-03-14', '2025-03-31', '2025-04-10', '2025-04-14', '2025-04-18', '2025-05-01', '2025-08-15', '2025-08-27', '2025-10-02', '2025-10-21', '2025-10-22', '2025-11-05', '2025-12-25']
# # holidays_25.append('2024-03-20') #add unusual holidays
# # today = datetime.now().date()
# today = datetime(year=2025, month=1, day=21).date()
# b_days = pd.bdate_range(start=today-timedelta(days=7), end=today, freq='C', weekmask='1111100', holidays=holidays_25).date.tolist()
# # b_days = b_days.append(pd.DatetimeIndex([pd.Timestamp(year=2024, month=1, day=20)])) #add unusual trading days
#
# # yesterday = today-timedelta(days=1)
# # yesterday = datetime(year=2025, month=1, day=13).date()
# today, yesterday = sorted(b_days)[-1], sorted(b_days)[-2]

root_dir = os.path.dirname(os.path.abspath(__file__))
nnf_file_path = os.path.join(root_dir, "Final_NNF_ID.xlsx")
eod_test_dir = os.path.join(root_dir, 'eod_testing')
eod_input_dir = os.path.join(root_dir, 'eod_original')
eod_output_dir = os.path.join(root_dir, 'eod_data')
table_dir = os.path.join(root_dir, 'table_data')
bhav_path = os.path.join(root_dir, 'bhavcopy')
test_dir = os.path.join(root_dir, 'testing')
bse_dir = os.path.join(root_dir, 'bse_data')
eod_net_pos_input_dir = os.path.join(root_dir, 'test_net_position_original')
eod_net_pos_output_dir = os.path.join(root_dir, 'test_net_position_code')
zipped_dir=os.path.join(root_dir, 'zipped_files')
dir_list = [zipped_dir]
status = [os.makedirs(_dir, exist_ok=True) for _dir in dir_list if not os.path.exists(_dir)]

# # eod_df = read_notis_file(os.path.join(eod_input_dir, f'EOD Position_{yesterday.strftime("%d_%b_%Y")}_1.xlsx'))
# eod_df = read_notis_file(os.path.join(eod_net_pos_input_dir, f'net_position_eod_{yesterday.strftime("%d_%m_%Y")}.xlsx')) # net_position_eod_23_01_2025.xlsx
# # # eod_df = read_notis_file(os.path.join(eod_dir, rf'NOTIS_DESK_WISE_FINAL_NET_POSITION_{yesterday.strftime("%Y-%m-%d")}_testing_1.xlsx'))
# eod_df.columns = eod_df.columns.str.replace(' ', '')
# eod_df.drop(columns=[col for col in eod_df.columns if col is None], inplace=True)
# eod_df = eod_df.add_prefix('Eod')
# # # eod_df = read_notis_file(rf"C:\Users\vipulanand\Downloads\Book1.xlsx")
# eod_df.EodExpiry = eod_df.EodExpiry.dt.date
# # eod_df['EodClosingQty'] = eod_df['EodClosingQty'].astype('int64')
# # eod_df['EodMTM'] = eod_df['EodClosingQty'] * eod_df['EodClosingPrice']
# # eod_df = eod_df.iloc[:,1:]
# grouped_eod = eod_df.groupby(by=['EodSymbol','EodExpiry','EodStrike','EodType']).agg({'EodEODQty':'sum'}).reset_index()
# grouped_eod = grouped_eod.drop_duplicates()
#
#
#
# tablenam = f'NOTIS_DESK_WISE_NET_POSITION_{today.strftime("%Y-%m-%d")}'
# modified_df = read_data_db(for_table=tablenam)
# # # modified_df1 = read_notis_file(os.path.join(table_dir, f'NOTIS_DESK_WISE_NET_POSITION_{today.strftime("%Y_%m_%d")}.xlsx'))
# modified_df.expiryDate = modified_df.expiryDate.astype('datetime64[ns]')
# modified_df.expiryDate = modified_df.expiryDate.dt.date
# modified_df.loc[modified_df['optionType'] == 'XX', 'strikePrice'] = 0
# modified_df.strikePrice = modified_df.strikePrice.apply(lambda x: x/100 if x>0 else x)
# modified_df.strikePrice = modified_df.strikePrice.astype('int64')
# # grouped_modified_df = modified_df.groupby(by=['mainGroup','symbol', 'expiryDate', 'strikePrice', 'optionType']).agg({'buyAvgQty':'sum','sellAvgQty':'sum','volume':'sum'}).reset_index()
# # grouped_modified_df = grouped_modified_df.drop_duplicates()
# grouped_modified_df = modified_df.groupby(by=['symbol', 'expiryDate', 'strikePrice', 'optionType']).agg({'buyAvgQty':'sum','sellAvgQty':'sum'}).reset_index()
# grouped_modified_df['IntradayNetQty'] = grouped_modified_df['buyAvgQty'] - grouped_modified_df['sellAvgQty']
# grouped_modified_df.rename(columns={'buyAvgQty':'buyQty','sellAvgQty':'sellQty'})
# # bhav_df = read_notis_file(os.path.join(bhav_path, rf'regularBhavcopy_{today.strftime("%d%m%Y")}.xlsx')) # regularBhavcopy_14012025.xlsx
# # bhav_df.columns = bhav_df.columns.str.replace(' ', '')
# # bhav_df.columns = bhav_df.columns.str.capitalize()
# # bhav_df = bhav_df.add_prefix('Bhav')
# # bhav_df.BhavExpiry = bhav_df.BhavExpiry.apply(lambda x: pd.to_datetime(get_date_from_non_jiffy(x))).dt.strftime('%Y-%m-%d')
# # bhav_df.BhavExpiry = bhav_df.BhavExpiry.apply(lambda x: pd.to_datetime(x).date())
# # bhav_df.loc[bhav_df['BhavOptiontype'] == 'XX', 'BhavStrikeprice'] = 0
# # bhav_df = bhav_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
# # bhav_df.BhavStrikeprice = bhav_df.BhavStrikeprice.apply(lambda x: x/100 if x>0 else x)
# # bhav_df.BhavStrikeprice = bhav_df.BhavStrikeprice.astype('int64')
# # # modified_df.expiryDate = modified_df.expiryDate.astype('datetime64[ns]')
# # # eod_df.EodExpiry = eod_df.EodExpiry.dt.date
# # # modified_df.expiryDate = modified_df.expiryDate.dt.date
# # # modified_df.strikePrice = modified_df.strikePrice.apply(lambda x: x/100 if x>0 else x)
# # # modified_df.strikePrice = modified_df.strikePrice.astype('int64')
# # # modified_df.loc[modified_df['optionType'] == 'XX', 'strikePrice'] = 0
# # # eod_df['EodClosingQty'] = eod_df['EodClosingQty'].astype('int64')
# # # eod_df['EodMTM'] = eod_df['EodClosingQty'] * eod_df['EodClosingPrice']
# # col_keep = ['BhavSymbol', 'BhavExpiry', 'BhavStrikeprice', 'BhavOptiontype','BhavClosingprice']
# # bhav_df = bhav_df[col_keep]
# # # col_drop = ['BhavTotalValue','BhavOpenInterest','BhavChangeInOpenInterest']
# # # bhav_df = bhav_df.drop(columns=[col for col in bhav_df.columns if col in col_drop])
# # bhav_df = bhav_df.drop_duplicates()
#
# # merged_df = grouped_modified_df.merge(eod_df, left_on=["mainGroup", "symbol", "expiryDate", "strikePrice", "optionType"], right_on=['EodMainGroup', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType'], how='outer')
# # merged_df = eod_df.merge(grouped_modified_df, right_on=["mainGroup", "symbol", "expiryDate", "strikePrice", "optionType"], left_on=['EodMainGroup', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType'], how='outer')
# merged_df = grouped_eod.merge(grouped_modified_df, right_on=["symbol", "expiryDate", "strikePrice", "optionType"], left_on=['EodSymbol', 'EodExpiry', 'EodStrike', 'EodType'], how='outer')
# merged_df.fillna(0, inplace=True)
# merged_df = merged_df.drop_duplicates()
#
# coltd1 = ['EodSymbol', 'EodStrike', 'EodType', 'EodExpiry']
# coltd2 = ['symbol', 'strikePrice', 'optionType', 'expiryDate']
# for i in range(len(coltd1)):
#     merged_df.loc[merged_df[coltd1[i]] == 0, coltd1[i]] = merged_df[coltd2[i]]
#     merged_df.loc[merged_df[coltd2[i]] == 0, coltd2[i]] = merged_df[coltd1[i]]
# merged_df['EodNetQty'] = merged_df['EodEODQty'] + merged_df['IntradayNetQty']
#
# a=0
# # merged_bhav_df = merged_df.merge(bhav_df, left_on=["symbol", "expiryDate", "strikePrice", "optionType"], right_on=['BhavSymbol', 'BhavExpiry', 'BhavStrikeprice', 'BhavOptiontype'], how='left')
# # # for index, row in merged_bhav_df.iterrows():
# # #     if abs(row['volume']) > 0:
# # #         if (row['MTM'] > 0 and abs(row['MTM'])>abs(row['EodMTM'])) or (row['EodMTM'] > 0 and abs(row['EodMTM'])>abs(row['MTM'])):
# # #             sign = 1
# # #         else:
# # #             sign = -1
# # #         # merged_bhav_df.loc[index, 'NetAvgPrice'] = (abs(row['MTM']) + abs(
# # #         #     (row['BhavClosingprice'] * row['EodClosingQty']))) / (abs(row['volume']) + abs(row['EodClosingQty']))
# # #         merged_bhav_df.loc[index, 'NetAvgPrice'] = abs(row['NetQty']) / abs(row['BhavClosingprice'])
# # merged_bhav_df = merged_bhav_df.drop_duplicates()
# #
# # def find_expired_mtm(row):
# #     if row['expired'] == True: # ClosingQty=NetQty
# #             # if (row['MTM'] > 0 and abs(row['MTM'])>abs(row['EodMTM'])) or (row['EodMTM'] > 0 and abs(row['EodMTM'])>abs(row['MTM'])):
# #             #     sign = 1
# #             #     return row['MTM']+row['EodMTM']+(row['NetQty']*row['BhavClosingprice'])
# #             # else:
# #             #     sign = -1
# #             #     return -1*(row['MTM']+row['EodMTM']+(row['NetQty']*row['BhavClosingprice']))
# #             if row['NetQty'] != 0:
# #                 if row.EodOptionType == 'PE':
# #                     row.rate = max((row.EodStrike - row.Spot),0)
# #                 else:
# #                     row.rate = max((row.Spot - row.EodStrike),0)
# #                 if row.NetQty < 0:
# #                     row['BuyQty'] = -1*row.NetQty
# #                     row.BuyRate = row.rate
# #                     row.SellQty = 0
# #                     row.SellRate = 0
# #                 else:
# #                     row['BuyQty'] = 0
# #                     row.BuyRate = 0
# #                     row.SellQty = row.NetQty
# #                     row.SellRate = row.rate
# #                 row.BuyValue = row.BuyRate * row.BuyQty
# #                 row.SellValue = row.SellRate * row.SellQty
# #
# #             return (-1*row['MTM'])+row['EodMTM']+(row['NetQty']*row['BhavClosingprice'])
# #
# # merged_bhav_df.loc[merged_bhav_df['expiryDate'] == today, 'expired'] = True
# # merged_bhav_df['Spot'] = merged_bhav_df.apply(lambda row: row['BhavClosingprice'] if row['expiryDate'] == today else '', axis=1)
# # # merged_bhav_df['spot'] = np.where(merged_bhav_df['expiryDate'] == today, merged_bhav_df['BhavClosingprice'], merged_bhav_df['spot'])
# # # merged_bhav_df['NetAvgPrice'] = merged_bhav_df.apply(lambda row: abs(row['NetQty'])/abs(row['BhavClosingprice']) if abs(row['volume'])>0 else None, axis=1)
# # merged_bhav_df['expiredMTM'] = merged_bhav_df.apply(find_expired_mtm, axis=1)
# # # col_to_keep = modified_df.columns.tolist()+['EodLong', 'EodShort','EodClosingQty','EodClosingPrice','EodSubGroup','EodMainGroup', 'EodMTM', 'expired', 'NetQty','BhavClosingprice', 'NetAvgPrice', 'expiredMTM']
# # # merged_bhav_df.drop(columns=[col for col in merged_bhav_df.columns if col not in col_to_keep], axis=1, inplace=True)
# # # col_drop = ['EodMTM','mainGroup', 'account', 'brokerID', 'tokenNumber','MTM', 'symbol', 'expiryDate', 'strikePrice', 'optionType','BhavSymbol', 'BhavExpiry', 'BhavStrikeprice', 'BhavOptiontype', 'NetAvgPrice']
# # # merged_bhav_df = merged_bhav_df.drop(columns=[col for col in merged_bhav_df.columns if col in col_drop])
# # merged_bhav_df['Long'] = merged_bhav_df['EodLong'] + merged_bhav_df['buyAvgQty']
# # merged_bhav_df['Short'] = merged_bhav_df['EodShort'] + merged_bhav_df['sellAvgQty']
# # # col_keep = ['EodUnderlying', 'EodStrike', 'EodOptionType', 'EodExpiry','EodSubGroup', 'EodMainGroup', 'Long','Short','NetQty','BhavClosingprice']
# # # merged_bhav_df.drop(columns=[col for col in merged_bhav_df.columns if col not in col_keep], axis=1, inplace=True)
# # # merged_bhav_df.columns = merged_bhav_df.columns.str.replace('Eod','')
# # merged_bhav_df.rename(columns={'NetQty':'ClosingQty','BhavClosingprice':'ClosingPrice'}, inplace=True )
# # # merged_bhav_df = merged_bhav_df[['Underlying', 'Strike', 'OptionType', 'Expiry', 'Long', 'Short', 'ClosingQty', 'ClosingPrice', 'SubGroup', 'MainGroup']]
# # merged_bhav_df = merged_bhav_df.drop_duplicates()
# # merged_bhav_df.rename(columns={'buyAvgQty':'BuyQty','sellAvgQty':'SellQty'}, inplace=True )
# # merged_bhav_df.drop(columns=['EodMTM','mainGroup','symbol', 'expiryDate', 'strikePrice', 'optionType','BhavSymbol', 'BhavExpiry', 'BhavStrikeprice', 'BhavOptiontype','expired', 'expiredMTM', 'Long', 'Short'], axis=1, inplace=True)
# # # # merged_df.drop(columns=eod_df.columns.tolist(), axis=1, inplace=True)
# # # merged_bhav_df['Long'] = merged_bhav_df['buyAvgQty'] + merged_bhav_df['EodLong']
# # # merged_bhav_df['Short'] = merged_bhav_df['sellAvgQty'] + merged_bhav_df['EodShort']
# # # col_keep = ['symbol', 'strikePrice', 'optionType', 'expiryDate', 'Long', 'Short', 'NetQty', 'BhavClosingprice', 'EodSubGroup', 'mainGroup']
# # # merged_bhav_df.drop(columns=[col for col in merged_bhav_df.columns if col not in col_keep], axis=1, inplace=True)
# # # merged_bhav_df.rename(columns={'symbol':'Underlying', 'strikePrice':'Strike', 'optionType':'OptionType', 'NetQty':'ClosingQty', 'BhavClosingprice':'ClosingPrice', 'EodSubGroup':'SubGroup', 'mainGroup':'mainGroup'})
# # # merged_bhav_df = merged_bhav_df[col_keep]
# # a=0
# # def update_qty(row):
# #     if row.Long > row.Short:
# #         row.Long = row.ClosingQty
# #         row.Short = 0
# #     elif row.Long < row.Short:
# #         row.Short = row.ClosingQty
# #         row.Long = 0
# #     return row
# # # merged_bhav_df = merged_bhav_df.apply(update_qty, axis=1)
# # write_notis_data(merged_bhav_df, os.path.join(eod_output_dir, f'Eod_{today.strftime("%Y_%m_%d")}_test_1.xlsx'))
# # write_notis_postgredb(merged_bhav_df, table_name=n_tbl_notis_net_position, raw=False)
# # print(f'file made for {today}')
# # # write_notis_data(modified_df, f'desk_{today.strftime("%Y-%m-%d")}.xlsx')
# # # write_notis_data(eod_df, f'eod_{today.strftime("%Y-%m-%d")}.xlsx')
# # # write_notis_data(bhav_df, f'bhav_{today.strftime("%Y-%m-%d")}.xlsx')
# # # print(eod_df.head(),'\n',modified_df.head())
# filtered_merged = merged_df.query("EodEODQty != 0 or IntradayNetQty != 0")
# write_notis_data(filtered_merged, os.path.join(eod_net_pos_output_dir, f'test_net_pos_{today.strftime("%d_%m_%y")}.xlsx'))
# b=0
#
# # pbar = progressbar.ProgressBar(max_value=10, widgets=[progressbar.Percentage(), ' ', progressbar.Bar(marker='=', left='[', right=']'), progressbar.ETA()])
# # pbar.update(1)
# # for i in range(10):
# #     time.sleep(1)
# #     pbar.update(i + 1)
# # pbar.finish()

# n_tbl_notis_trade_book = "NOTIS_TRADE_BOOK"
# # n_tbl_notis_trade_book = "NOTIS_DESK_WISE_EOD_POSITION_2025-01-21"
# df = read_notis_file(rf"D:\notis_analysis\modified_data\NOTIS_DATA_29JAN2025.xlsx")
# # df = read_notis_file(rf"D:\notis_analysis\table_data\NOTIS_DESK_WISE_NET_POSITION_2025_01_20_2.xlsx")
# # df = read_notis_file(rf"D:\notis_analysis\eod_data\Eod_2025_01_21_test_2.xlsx")
# a=0
# write_notis_postgredb(df, n_tbl_notis_trade_book)
#
# # grouped_a = a.groupby(by=['MainGroup','SubGroup','sym','expDt', 'strPrc', 'optType','bsFlg'], as_index=False).agg({'trdQty':'sum', 'trdPrc':'mean'})


# bhav_df = read_notis_file(rf"D:\notis_analysis\testing\regularBhavcopy_10022025.xlsx")
# bhav_df1 = read_file(rf"D:\notis_analysis\testing\regularBhavcopy_10022025.xlsx")
# bhav_df2 = read_file(rf"D:\notis_analysis\testing\regularBhavcopy_10022025.csv")
a=0
# ex_dt = '11-02-2025  06:03:30'
# ex = datetime(2025, 2, 12, 9, 36, 50, microsecond=297695)
# epoch_time = int(time.mktime(ex.timetuple()))
# print(epoch_time)
b=0
# class ServiceApp:
#     def __init__(self):
#         super().__init__()
#         self.app = FastAPI(title='NOTIS_Net_Position', description='Notis_net_position', docs_url='/docs', openapi_url='/openapi.json')
#         self.app.add_middleware(CORSMiddleware, allow_origins = ['*'], allow_credentials = True, allow_methods=['*'], allow_headers=['*'])
#         self.add_routes()
#
#     def add_routes(self):
#         self.app.add_api_route('/netPosition/intraday', methods=['GET'], endpoint=self.get_intraday_net_position)
#
#     def get_intraday_net_position(self):
#         # df_db = read_data_db()
#         # nnf_file_path = os.path.join(root_dir, "Final_NNF_ID.xlsx")
#         # if not os.path.exists(nnf_file_path):
#         #     raise FileNotFoundError("NNF File not found. Please add the NNF file and try again.")
#         # readable_mod_time = datetime.fromtimestamp(os.path.getmtime(nnf_file_path))
#         # if readable_mod_time.date() == today:  # Check if the NNF file is modified today or not
#         #     print(f'New NNF Data found, modifying the nnf data in db . . .')
#         #     df_nnf = pd.read_excel(nnf_file_path, index_col=False)
#         #     df_nnf = df_nnf.loc[:, ~df_nnf.columns.str.startswith('Un')]
#         #     df_nnf.columns = df_nnf.columns.str.replace(' ', '', regex=True)
#         #     df_nnf.dropna(how='all', inplace=True)
#         #     df_nnf = df_nnf.drop_duplicates()
#         #     write_notis_postgredb(df_nnf, n_tbl_notis_nnf_data, raw=False)
#         # else:
#         #     df_nnf = read_data_db(nnf=True, for_table=n_tbl_notis_nnf_data)
#         #     df_nnf = df_nnf.drop_duplicates()
#         # modified_df = modify_file(df_db, df_nnf)
#         # modified_df.expiryDate = modified_df.expiryDate.astype('datetime64[ns]')
#         # modified_df.expiryDate = modified_df.expiryDate.dt.date
#         # modified_df.loc[modified_df['optionType'] == 'XX', 'strikePrice'] = 0
#         # modified_df.strikePrice = modified_df.strikePrice.apply(lambda x: x / 100 if x > 0 else x)
#         # modified_df.strikePrice = modified_df.strikePrice.astype('int64')
#         # grouped_modified_df = modified_df.groupby(by=['symbol', 'expiryDate', 'strikePrice', 'optionType']).agg(
#         #     {'buyAvgQty': 'sum', 'buyAvgPrice': 'sum', 'sellAvgQty': 'sum', 'sellAvgPrice': 'sum'}).reset_index()
#         # grouped_modified_df['IntradayVolume'] = grouped_modified_df['buyAvgQty'] - grouped_modified_df['sellAvgQty']
#         # grouped_modified_df.rename(columns={'buyAvgQty': 'buyQty', 'sellAvgQty': 'sellQty'})
#         # return grouped_modified_df.to_dict(orient='records')
#
#         # tablenam = f'NOTIS_DESK_WISE_NET_POSITION_{today.strftime("%Y-%m-%d")}'
#         tablenam = f'NOTIS_DESK_WISE_NET_POSITION_2025-02-14'
#         desk_db_df = read_data_db(for_table=tablenam)
#         desk_db_df.expiryDate = desk_db_df.expiryDate.astype('datetime64[ns]')
#         desk_db_df.expiryDate = desk_db_df.expiryDate.dt.date
#         desk_db_df.loc[desk_db_df['optionType'] == 'XX', 'strikePrice'] = 0
#         desk_db_df.strikePrice = desk_db_df.strikePrice.apply(lambda x: x / 100 if x > 0 else x)
#         desk_db_df.strikePrice = desk_db_df.strikePrice.astype('int64')
#         grouped_desk_db_df = desk_db_df.groupby(by=['symbol', 'expiryDate', 'strikePrice', 'optionType']).agg(
#             {'buyAvgQty': 'sum', 'buyAvgPrice': 'sum', 'sellAvgQty': 'sum', 'sellAvgPrice': 'sum'}).reset_index()
#         grouped_desk_db_df['IntradayVolume'] = grouped_desk_db_df['buyAvgQty'] - grouped_desk_db_df['sellAvgQty']
#         grouped_desk_db_df.rename(columns={'buyAvgQty': 'buyQty', 'sellAvgQty': 'sellQty'})
#         return grouped_desk_db_df.to_dict(orient='records')
#
# service = ServiceApp()
# app = service.app
#
# if __name__ == '__main__':
#     uvicorn.run('untitled:app', host='0.0.0.0', port=8851, workers=5)
c=0
# df = read_file(rf"D:\notis_analysis\modified_data\NOTIS_TRADE_DATA_14FEB2025.xlsx")
# col_keep = ['trdQty', 'trdPrc', 'bsFlg','ctclid','sym', 'inst', 'expDt', 'strPrc', 'optType','CreateDate', 'TerminalID', 'TerminalName', 'UserID', 'SubGroup', 'MainGroup', 'NeatID']
# df = df[col_keep]
# df.expDt = pd.to_datetime(df.expDt).dt.date
# grouped_df = df.groupby(by=['ctclid','sym', 'inst', 'expDt', 'strPrc', 'optType','bsFlg'], as_index=False).agg({
#     'trdQty':'sum',
#     'trdPrc':'mean',
#     'CreateDate': lambda x: x.unique()[0],
#     'TerminalID': lambda x: x.unique()[0],
#     'TerminalName': lambda x: x.unique()[0],
#     'UserID': lambda x: x.unique()[0],
#     'SubGroup': lambda x: x.unique()[0],
#     'MainGroup': lambda x: x.unique()[0],
#     'NeatID': lambda x: x.unique()[0]
# })
b=0
# cur_time = datetime.now().time()
# tar_time = datetime.strptime('1530', '%H%M').time()
c=0
# def get_zerodha_data(from_time: datetime = Query(), to_time: datetime = Query()):
#     db_name = f'data_analytics'
#     pg_user = 'postgres'
#     pg_pass = 'Rathi@321'
#     pg_host = '192.168.100.173'
#     pg_port = '5432'
#     pg_pass_encoded = quote(pg_pass)
#
#     ist = pytz.timezone('Asia/Kolkata')
#     from_datetime = ist.localize(datetime.strptime(f'{today} {from_time}', '%Y-%m-%d %H:%M:%S'))
#     to_datetime = ist.localize(datetime.strptime(f'{today} {to_time}', '%Y-%m-%d %H:%M:%S'))
#     minute_list = []
#     start_time = from_datetime
#     while start_time <= to_datetime:
#         minute_list.append(start_time)
#         start_time += timedelta(minutes=1)
#
#     engine_str = f"postgresql+psycopg2://{pg_user}:{pg_pass_encoded}@{pg_host}:{pg_port}/{db_name}"
#     engine = create_engine(engine_str)
#     conn = engine.connect()
#     query = f"""
#         SELECT * FROM snap
#         WHERE timestamp >= '{from_datetime}'
#         AND timestamp <= '{to_datetime}'
#         ORDER BY id ASC
#     """
#     df = pd.read_sql(query, con=conn)
#     return df
d=0
# to download bhavcopy file from the server
# import paramiko
# host = '192.168.112.81'
# username = 'greek'
# password = 'greeksoft'
# filename = f"regularBhavcopy_{yesterday.strftime('%d%m%Y')}.csv" #sample=regularBhavcopy_13022025
# remote_path = rf'/home/greek/NSE_BSE_Broadcast/NSE/Bhavcopy/Files/{filename}'
# local_path = os.path.join(test_dir, filename)
# try:
#     transport = paramiko.Transport((host, 22))
#     transport.connect(username=username, password=password)
#     sftp = paramiko.SFTPClient.from_transport(transport)
#     sftp.get(remote_path, local_path)
#     sftp.close()
#     transport.close()
# except Exception as e:
#     print(f'Error: {e}')
e=0
# tablename = f'notis_raw_data_2025-02-17'
# df = read_data_db(for_table=tablename)
# list_str_int64 = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 16, 17, 18, 19, 21, 23, 27, 28]
# list_str_none = [15, 20, 25, 30, 31, 32, 33, 34, 35, 36, 37]
# list_none_str = [38]
# for i in list_str_int64:
#     column_name = f'Column{i}'
#     df[f'Column{i}'] = df[f'Column{i}'].astype('int64')
# for i in list_str_none:
#     df[f'Column{i}'] = None
# for i in list_none_str:
#     df[f'Column{i}'] = df[f'Column{i}'].astype('str')
# print('Starting file modification...')
# df.rename(columns={
#     'Column1': 'seqNo', 'Column2': 'mkt', 'Column3': 'trdNo',
#     'Column4': 'trdTm', 'Column5': 'Tkn', 'Column6': 'trdQty',
#     'Column7': 'trdPrc', 'Column8': 'bsFlg', 'Column9': 'ordNo',
#     'Column10': 'brnCd', 'Column11': 'usrId', 'Column12': 'proCli',
#     'Column13': 'cliActNo', 'Column14': 'cpCD', 'Column15': 'remarks',
#     'Column16': 'actTyp', 'Column17': 'TCd', 'Column18': 'ordTm',
#     'Column19': 'Booktype', 'Column20': 'oppTmCd', 'Column21': 'ctclid',
#     'Column22': 'status', 'Column23': 'TmCd', 'Column24': 'sym',
#     'Column25': 'ser', 'Column26': 'inst', 'Column27': 'expDt',
#     'Column28': 'strPrc', 'Column29': 'optType', 'Column30': 'sessionID',
#     'Column31': 'echoback', 'Column32': 'Fill1', 'Column33': 'Fill2',
#     'Column34': 'Fill3', 'Column35': 'Fill4', 'Column36': 'Fill5', 'Column37': 'Fill6'
# }, inplace=True)
#
# df['ordTm'] = df['ordTm'].apply(lambda x: get_date_from_non_jiffy(int(x)))
#
# df['expDt'] = df['expDt'].apply(lambda x: get_date_from_non_jiffy(int(x)))
# df.expDt = df.expDt.astype('datetime64[ns]').dt.date
#
# df['trdTm'] = df['trdTm'].apply(lambda x: get_date_from_jiffy(int(x)))
# # --------------------------------------------------------------------------------------------------------------------------------
# df['bsFlg'] = np.where(df['bsFlg'] == 1, 'B', 'S')
# e=0
# # def collected_ctcl(series):
# #     return list(series.unique())
# # user_choice = int(input('Enter how would you like to make net-pos,\n 1-ctclidwise\n 2-symbolwise\n 3-both'))
# # if user_choice == 1 or user_choice == 3:
# #     print('Calculating ctclid wise Net-position...')
# #     grouped_df = df.groupby(['ctclid','sym','expDt','strPrc','optType','bsFlg'], as_index=False).agg({'trdQty':'sum','trdPrc':'mean'})
# # else:
# #     print('calculating sym wise netposition')
# #     grouped_df = df.groupby(['sym', 'expDt', 'strPrc', 'optType', 'bsFlg'], as_index=False).agg({'trdQty': 'sum', 'trdPrc': 'mean','ctclid':collected_ctcl})
# f=0
# # pivot_df = grouped_df.pivot_table(
# #     index=['ctclid', 'sym', 'expDt', 'strPrc', 'optType'],
# #     columns='bsFlg',
# #     values=['trdQty', 'trdPrc'],
# #     aggfunc='sum',
# #     fill_value=0
# # )
#
# pivot_df = df.pivot_table(
#     index=['ctclid', 'sym', 'expDt', 'strPrc', 'optType'],
#     columns='bsFlg',
#     values=['trdQty', 'trdPrc'],
#     aggfunc={'trdQty': 'sum', 'trdPrc': 'mean'},
#     fill_value=0
# )
#
# # pivot_df.columns = ['BuyPrc', 'SellPrc','BuyVol', 'SellVol']
# # pivot_df = pivot_df.reset_index()
# # pivot_df = pivot_df[['ctclid', 'sym', 'expDt', 'strPrc', 'optType', 'BuyVol', 'BuyPrc', 'SellVol', 'SellPrc']]
# pivot_df.columns = ['BuyPrc', 'SellPrc','BuyVol', 'SellVol']
# pivot_df = pivot_df.reset_index()
# pivot_df = pivot_df[['ctclid', 'sym', 'expDt', 'strPrc', 'optType', 'BuyVol', 'BuyPrc', 'SellVol', 'SellPrc']]
f=0
# engine = create_engine(engine_str)
# sessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
# today=datetime.now().date()
# def get_db():
#     db = sessionLocal()
#     try:
#         yield db
#     finally:
#         db.close()
# def test_page_download_trade_data(for_date: date,db:Session=Depends(get_db)):
#     for_dt = pd.to_datetime(for_date).date()
#     tablename = f'NOTIS_TRADE_BOOK' if for_dt == today else f'NOTIS_TRADE_BOOK_{for_dt}'
#     zip_path = os.path.join(zipped_dir,f'zippped_{tablename}.xlsx.gz')
#     if not os.path.exists(zip_path):
#         query=text(rf'select * from "{tablename}"')
#         result = db.execute(query).fetchall()
#         with gzip.open(zip_path, 'wb') as f:
#             workbook=xlsxwriter.Workbook(f,{'in_memory':True})
#             worksheet=workbook.add_worksheet()
#             headers = result.keys()
#             for col, header in enumerate(headers):
#                 worksheet.write(0, col, header)
#             for row_num, row in enumerate(result, start=1):
#                 for col_num, cell in enumerate(row):
#                     worksheet.write(row_num, col_num, cell)
#             workbook.close()
#         # return FileResponse(zip_path, media_type='application/gzip')
#         return pd.read_excel(zip_path)
#     else:
#         return pd.read_excel(zip_path)
q=0
# def test_xlsx_gz(for_date:date):
#     for_dt = pd.to_datetime(for_date).date()
#     tablename = f'NOTIS_TRADE_BOOK' if for_dt == today else f'NOTIS_TRADE_BOOK_{for_dt}'
#     zip_path = os.path.join(zipped_dir, f'zippped_{tablename}.xlsx.gz')
#     if not os.path.exists(zip_path):
#         query = text(rf'select * from "{tablename}"')
#         with engine.connect() as conn:
#             result = conn.execute(query)
#         total_rows = len(result.fetchall())
#         # # wb=openpyxl.Workbook()
#         # # ws=wb.active
#         # # headers = result.keys()
#         # # ws.append(headers)
#         # # for row in result:
#         # #     ws.append(row)
#         # # buffer = io.BytesIO()
#         # # wb.save(buffer)
#         # # buffer.seek(0)
#         # # with gzip.open(zip_path, 'wb', encoding='utf-8') as f:
#         # #     f.write(buffer.getvalue())
#         # with gzip.open(zip_path, 'wb') as f:
#         #     workbook=xlsxwriter.Workbook(f,{'in_memory':True})
#         #     worksheet=workbook.add_worksheet()
#         #     headers = result.keys()
#         #     for col, header in enumerate(headers):
#         #         worksheet.write(0, col, header)
#         #     for row_num, row in enumerate(result, start=1):
#         #         for col_num, cell in enumerate(row):
#         #             worksheet.write(row_num, col_num, cell)
#         #     workbook.close()
#         pbar=progressbar.ProgressBar(max_value=total_rows,widgets=[progressbar.Percentage(),' ',progressbar.Bar(marker='=',left='[',right=']'),progressbar.ETA()])
#         buffer = io.BytesIO()
#         wb = xlsxwriter.Workbook(buffer,{'in_memory':True})
#         ws = wb.add_worksheet()
#         # pbar.update(0)
#         header = result.keys()
#         for col, header_value in enumerate(header):
#             ws.write(0, col, header_value)
#         pbar.update(0)
#         for row_num, row in enumerate(result, start=1):
#             for col_num, cell in enumerate(row):
#                 ws.write(row_num, col_num, cell)
#             pbar.update(row_num+1)
#         wb.close()
#         pbar.finish()
#         buffer.seek(0)
#         with gzip.open(zip_path, 'wb') as f:
#             f.write(buffer.getvalue())
#         return 'Done'
#     else:
#         return True
q=0
# def test_xlsx_gz_2(for_date:date):
#     for_dt = pd.to_datetime(for_date).date()
#     tablename = f'NOTIS_TRADE_BOOK' if for_dt == today else f'NOTIS_TRADE_BOOK_{for_dt}'
#     zip_path = os.path.join(zipped_dir, f'zippped_{tablename}.xlsx.gz')
#     xl_path = os.path.join(zipped_dir, f'xl_{tablename}.xlsx')
#     if not os.path.exists(zip_path):
#         query = text(rf'select * from "{tablename}"')
#         buffer=io.BytesIO()
#         wb=xlsxwriter.Workbook(buffer,{'in_memory':True})
#         ws=wb.add_worksheet()
#         with engine.connect() as conn:
#             result = conn.execute(query)
#             total_rows = result.rowcount
#             pbar = progressbar.ProgressBar(max_value=total_rows+1, widgets=[progressbar.Percentage(),' ',progressbar.Bar(marker='=',left='[',right=']'),progressbar.ETA()])
#             header = result.keys()
#             for col, row_value in enumerate(header):
#                 ws.write(0,col, row_value)
#             pbar.update(1)
#             for row_num, row in enumerate(result, start=1):
#                 for col_num, cell in enumerate(row):
#                     ws.write(row_num, col_num, cell)
#                 pbar.update(row_num+1)
#             pbar.finish()
#         wb.close()
#         buffer.seek(0)
#         print('zipping file . . .')
#         with gzip.open(zip_path, 'wb') as f:
#             f.write(buffer.getvalue())
#         with open(xl_path, 'wb') as f:
#             f.write(buffer.getvalue())
#         df = pd.read_sql_table(tablename, con=engine)
#         print(f"Data fetched from {tablename} table. Shape:{df.shape}")
#         print('done')
#     else:
#         return True
y=0
# def test_xlsx_gz_3(for_date:date):
#     for_dt = pd.to_datetime(for_date).date()
#     tablename = f'NOTIS_TRADE_BOOK' if for_dt == today else f'NOTIS_TRADE_BOOK_{for_dt}'
#     zip_path = os.path.join(zipped_dir, f'zippped_{tablename}.csv.gz')
#     csv_path = os.path.join(zipped_dir, f'zippped_{tablename}.csv')
#     # if not os.path.exists(zip_path):
#     query = text(rf'select * from "{tablename}"')
#     with engine.connect() as conn:
#         result = conn.execute(query)
#         total_rows = result.rowcount
#         pbar = progressbar.ProgressBar(max_value=total_rows+1,widgets=[progressbar.Percentage(),' ',progressbar.Bar(marker='=', left='[', right=']'), progressbar.ETA()])
#         buffer=io.BytesIO()
#         with gzip.GzipFile(fileobj=buffer, mode='wb') as gz:
#             writer = csv.writer(io.TextIOWrapper(gz, encoding='utf-8', newline=''))
#             header = result.keys()
#             writer.writerow(header)
#             pbar.update(1)
#             for row_num, row in enumerate(result, start=1):
#                 writer.writerow(row)
#                 pbar.update(row_num+1)
#             pbar.finish()
#     with open(zip_path, 'wb') as f:
#         f.write(buffer.getvalue())
#     print('done')

# def test_csv(for_date:date):
#     for_dt = pd.to_datetime(for_date).date()
#     tablename = f'NOTIS_TRADE_BOOK' if for_dt == today else f'NOTIS_TRADE_BOOK_{for_dt}'
#     csv_path = os.path.join(zipped_dir, f'zippped_{tablename}.csv')
#     query = text(rf'select * from "{tablename}"')
#     with engine.connect() as conn:
#         result = conn.execute(query)
#         total_rows = result.rowcount
#         pbar = progressbar.ProgressBar(max_value=total_rows+1,widgets=[progressbar.Percentage(),' ',progressbar.Bar(marker='=', left='[', right=']'), progressbar.ETA()])
#         with open(csv_path, mode='w', encoding='utf-8', newline='') as csvfile:
#             writer = csv.writer(csvfile)
#             header = result.keys()
#             writer.writerow(header)
#             pbar.update(1)
#             for row_num, row in enumerate(result, start=1):
#                 writer.writerow(row)
#                 pbar.update(row_num + 1)
#             pbar.finish()
#     print('done')
t=0
# engine=create_engine(engine_str)
# def test_modif(for_date:date):
#     for_dt = pd.to_datetime(for_date).date()
#     tablename = f'NOTIS_TRADE_BOOK' if for_dt == today else f'NOTIS_TRADE_BOOK_{for_dt}'
#     zip_path = os.path.join(zipped_dir, f'zippped_{tablename}.csv.gz')
#     csv_path = os.path.join(zipped_dir, f'csv_{tablename}.csv')
#     query = text(rf'select * from "{tablename}"')
#     with engine.connect() as conn:
#         result = conn.execute(query)
#         total_rows = result.rowcount
#         pbar = progressbar.ProgressBar(max_value=total_rows + 1, widgets=[progressbar.Percentage(), ' ',
#                                                                           progressbar.Bar(marker='=', left='[',
#                                                                                           right=']'),
#                                                                           progressbar.ETA()])
#         buffer = io.StringIO()
#         writer = csv.writer(buffer)
#         header = result.keys()
#         writer.writerow(header)
#         pbar.update(1)
#         for row_num, row in enumerate(result, start=1):
#             writer.writerow(row)
#             pbar.update(row_num + 1)
#         pbar.finish()
#
#     with open(csv_path, mode='w', encoding='utf-8', newline='') as csvfile:
#         csvfile.write(buffer.getvalue())
#
#     with gzip.open(zip_path, mode='wt', encoding='utf-8', newline='') as zipfile:
#         zipfile.write(buffer.getvalue())
o=0
# oi=pd.to_datetime('20250224').date()
# a = test_modif(oi)
r=0
# for_tab = ['NOTIS_TRADE_BOOK','NOTIS_DESK_WISE_NET_POSITION', 'NOTIS_NNF_WISE_NET_POSITION', 'NOTIS_USERID_WISE_NET_POSITION']
# tod = datetime(year=2025,month=2,day=28).date()
# for each in for_tab:
#     df = read_data_db(for_table=f"{each}_{tod.strftime('%Y-%m-%d')}")
#     print(f"data read from {each}_{tod.strftime('%Y-%m-%d')}")
#     write_notis_data(df=df, filepath=os.path.join(test_dir,f"{each}_{tod.strftime('%Y-%m-%d')}.csv"))
#     print(f"data written to {each}_{tod.strftime('%Y-%m-%d')}.csv")
e=0
# def truncate_tables():
#     table_name = ["NOTIS_TRADE_BOOK_2024-12-31","NOTIS_DESK_WISE_NET_POSITION","NOTIS_NNF_WISE_NET_POSITION","NOTIS_USERID_WISE_NET_POSITION"]
#     engine = create_engine(engine_str)
#     with engine.connect() as conn:
#         for each in table_name:
#             print(each)
#             res = conn.execute(text(f'select count(*) from "{each}"'))
#             row_count = res.scalar()
#             if row_count > 0:
#                 conn.execute(text(f'delete from "{each}"'))
#                 print(f'Existing data from table {each} deleted')
#             else:
#                 print(f'No data in table {each}, no need to delete')
# truncate_tables()
u=0
# import os
# import io
# import gzip
# import progressbar
# import pandas as pd
# from sqlalchemy.sql import text
# from openpyxl import Workbook
# from fastapi.responses import FileResponse
#
#
# def export_db_to_xlsx(tablename, zip_path):
#     if not os.path.exists(zip_path):
#         stt = datetime.now()
#         query = text(f'select * from "{tablename}"')
#         with engine.connect() as conn:
#             result = conn.execute(query)
#         total_rows = result.rowcount
#         batch_size = 1_000_000  # 1 million rows per sheet
#         num_sheets = (total_rows // batch_size) + (1 if total_rows % batch_size else 0)
#
#         print(f'Total rows in DB: {total_rows}, Splitting into {num_sheets} sheets')
#
#         buffer = io.BytesIO()
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Sheet1"
#
#         header = list(result.keys())
#         ws.append(header)
#
#         pbar = progressbar.ProgressBar(max_value=total_rows + 1, widgets=[
#             progressbar.Percentage(), ' ',
#             progressbar.Bar(marker='=', left='[', right=']'),
#             progressbar.ETA()
#         ])
#         pbar.update(1)
#
#         sheet_index = 1
#         row_num = 0
#         for row in result:
#             if row_num > 0 and row_num % batch_size == 0:
#                 sheet_index += 1
#                 ws = wb.create_sheet(title=f"Sheet{sheet_index}")
#                 ws.append(header)
#             ws.append(tuple(row))
#             row_num += 1
#             pbar.update(row_num + 1)
#
#         pbar.finish()
#
#         wb.save(buffer)
#         buffer.seek(0)
#
#         with gzip.open(zip_path, 'wb') as f_out:
#             f_out.write(buffer.getvalue())
#         ett=datetime.now()
#         print(f"Data exported successfully with {num_sheets} sheets.")
#         print(f'total time taken for zip_path1:{(ett-stt).total_seconds()}')
#         # return FileResponse(path=zip_path, media_type='application/gzip')
#     else:
#         return FileResponse(path=zip_path, media_type='application/gzip')
#
# def export_db_to_xlsx_2(tablename, zip_path):
#     if not os.path.exists(zip_path):
#         stt=datetime.now()
#         query = text(f'select * from "{tablename}"')
#         with engine.connect() as conn:
#             result = conn.execute(query)
#         total_rows = result.rowcount
#         batch_size = 1_000_000  # 1 million rows per sheet
#         num_sheets = (total_rows // batch_size) + (1 if total_rows % batch_size else 0)
#         print(f'Total rows in DB: {total_rows}, Splitting into {num_sheets} sheets')
#         buffer = io.BytesIO()
#         wb = xlsxwriter.Workbook(buffer, {'in_memory':True})
#         # ws = wb.add_worksheet()
#         header = result.keys()
#
#         pbar = progressbar.ProgressBar(max_value=total_rows + 1, widgets=[progressbar.Percentage(), ' ',progressbar.Bar(marker='=', left='[',right=']'),progressbar.ETA()])
#         # # for col, header_value in enumerate(header):
#         # #     ws.write(0,col,header_value)
#         # #
#         # # for row_num, row in enumerate(result, start=1):
#         # #     for col_num, cell in enumerate(row):
#         # #         ws.write(row_num,col_num,cell)
#         # #     pbar.update(row_num)
#         # # pbar.finish()
#         # # wb.close()
#         # #
#         # # buffer.seek(0)
#         # # with gzip.open(zip_path,'wb') as f:
#         # #     f.write(buffer.getvalue())
#         # for sheet_num in range(num_sheets):
#         #     ws = wb.add_worksheet(f'Sheet{sheet_num+1}')
#         #     for col, header_value in enumerate(header):
#         #         ws.write(0,col,header_value)
#         #     start_index = sheet_num*batch_size
#         #     last_index = min(start_index,total_rows)
#         #     for row_num, row in enumerate(result[start_index:last_index], start=1):
#         #         for col_num, cell in enumerate(row):
#         #             ws.write(row_num,col_num,cell)
#         #         pbar.update(start_index+row_num)
#         # pbar.finish()
#         # wb.close()
#
#         sheet_idx = 0
#         ws = wb.add_worksheet(f'Sheet{sheet_idx+1}')
#         for col, header_value in enumerate(header):
#             ws.write(0, col, header_value)
#         row_cnt = 1
#         with engine.connect() as conn:
#             result = conn.execution_options(stream_results=True).execute(query)
#             for row_num, row in enumerate(result, start=1):
#                 if row_cnt>batch_size:
#                     sheet_idx+=1
#                     row_cnt = 1
#                     ws = wb.add_worksheet(f'Sheet{sheet_idx+1}')
#                     for col, header_value in enumerate(header):
#                         ws.write(0, col, header_value)
#
#                 for col_num, cell in enumerate(row):
#                     ws.write(row_cnt, col_num, cell)
#                 row_cnt+=1
#                 pbar.update(row_num)
#         pbar.finish()
#         wb.close()
#
#         buffer.seek(0)
#         with gzip.open(zip_path, 'wb') as f:
#             f.write(buffer.getvalue())
#         ett=datetime.now()
#         print(f'total time taken for zip_path2:{(ett-stt).total_seconds()}')
#         print('done')
#
# def export_db_to_xlsx_3(tablename, zip_path):
#     if not os.path.exists(zip_path):
#         stt=datetime.now()
#         query = text(f'select * from "{tablename}"')
#         count_query = text(f'select count(*) from "{tablename}"')
#         with engine.connect() as conn:
#             total_rows = conn.execute(count_query).scalar()
#         # total_rows = result.rowcount
#         batch_size = 1_000_000  # 1 million rows per sheet
#         num_sheets = (total_rows // batch_size) + (1 if total_rows % batch_size else 0)
#         print(f'Total rows in DB: {total_rows}, Splitting into {num_sheets} sheets')
#         buffer = io.BytesIO()
#         # wb = xlsxwriter.Workbook(buffer, {'in_memory':True})
#         # writer = pd.ExcelWriter(buffer, engine='xlsxwriter')
#         writer = pd.ExcelWriter(buffer, engine='openpyxl')
#         # ws = wb.add_worksheet()
#         # header = result.keys()
#
#         pbar = progressbar.ProgressBar(max_value=total_rows + 1, widgets=[progressbar.Percentage(), ' ',progressbar.Bar(marker='=', left='[',right=']'),progressbar.ETA()])
#
#         # # sheet_idx = 0
#         # # ws = wb.add_worksheet(f'Sheet{sheet_idx+1}')
#         # # for col, header_value in enumerate(header):
#         # #     ws.write(0, col, header_value)
#         # # row_cnt = 1
#         # # with engine.connect() as conn:
#         # #     result = conn.execution_options(stream_results=True).execute(query)
#         # #     for row_num, row in enumerate(result, start=1):
#         # #         if row_cnt>batch_size:
#         # #             sheet_idx+=1
#         # #             row_cnt = 1
#         # #             ws = wb.add_worksheet(f'Sheet{sheet_idx+1}')
#         # #             for col, header_value in enumerate(header):
#         # #                 ws.write(0, col, header_value)
#         # #
#         # #         for col_num, cell in enumerate(row):
#         # #             ws.write(row_cnt, col_num, cell)
#         # #         row_cnt+=1
#         # #         pbar.update(row_num)
#         # # pbar.finish()
#         # # wb.close()
#         # to_update = 0
#         # with engine.connect() as conn:
#         #     for i, df in enumerate(pd.read_sql(query, conn, chunksize=batch_size)):
#         #         df.to_excel(writer, sheet_name=f'Sheet{i+1}', index=False)
#         #         to_update+=len(df)
#         #         pbar.update(to_update)
#         #         print(f'Written sheet{i+1} with {len(df)}rows')
#         # writer.close()
#         # pbar.finish()
#
#         to_update = 0
#         with engine.connect() as conn:
#             for i, df in enumerate(pd.read_sql(query, conn, chunksize=batch_size)):
#                 temp_csv = io.StringIO()
#                 df.to_csv(temp_csv, index=False)
#                 temp_csv.seek(0)
#                 df = pd.read_csv(temp_csv, index_col=False)
#                 df.to_excel(writer, sheet_name=f'Sheet{i+1}', index=False)
#                 to_update+=len(df)
#                 pbar.update(to_update)
#                 sys.stdout.flush()
#                 print(f'Written sheet{i+1} with {len(df)} rows')
#         writer.close()
#         pbar.finish()
#
#         buffer.seek(0)
#         with gzip.open(zip_path, 'wb') as f:
#             f.write(buffer.getvalue())
#         ett=datetime.now()
#         print(f'total time taken for zip_path3:{(ett-stt).total_seconds()}')
#         print('done')
#
# tablename = 'NOTIS_TRADE_BOOK_2025-02-24'
# zip_path = os.path.join(zipped_dir, f'zipped_{tablename}_older.xlsx.gz')
# zip_path1 = os.path.join(zipped_dir, f'zipped_{tablename}_test1.xlsx.gz')
# zip_path2 = os.path.join(zipped_dir, f'zipped_{tablename}_test2.xlsx.gz')
# zip_path3 = os.path.join(zipped_dir, f'zipped_{tablename}_test3.xlsx.gz')
# # export_db_to_xlsx(tablename, zip_path1)
# # export_db_to_xlsx_2(tablename, zip_path2)
# # export_db_to_xlsx_3(tablename, zip_path3)
u=0
# tablename = 'NOTIS_TRADE_BOOK_2025-03-04'
# zip_path4 = os.path.join(zipped_dir, f'zipped_{tablename}_test_2.xlsx.gz')
# def get_db():
#     engine=create_engine(engine_str)
#     sessionLocal = sessionmaker(autoflush=False, autocommit=False, bind=engine)
#     db = sessionLocal()
#     try:
#         yield db
#     finally:
#         db.close()
# def export_test_1(tablename, zip_path, db):
#     if not os.path.exists(zip_path):
#         stt=datetime.now()
#         total_rows = db.execute(text(rf'select count(*) from "{tablename}"')).scalar()
#         page_size = 5_00_000
#         num_pages = total_rows//page_size + (1 if (total_rows%page_size) else 0)
#         buffer = io.BytesIO()
#         wb = xlsxwriter.Workbook(buffer,{'in_memory':True})
#         # pbar = progressbar.ProgressBar(max_value=total_rows + 1, widgets=[progressbar.Percentage(),'', progressbar.Bar(marker='=', left='[',right=']'), progressbar.ETA()])
#         for page in range(num_pages):
#             query = f'select * from "{tablename}" limit {page_size} offset {(page)*page_size}'
#             print(query)
#             pbar = progressbar.ProgressBar(max_value=total_rows + 1, widgets=[progressbar.Percentage(), '',
#                                                                               progressbar.Bar(marker='=', left='[',
#                                                                                               right=']'),
#                                                                               progressbar.ETA()])
#             result = db.execute(query)
#             ws = wb.add_worksheet(f'Sheet{page+1}')
#             for col, header in enumerate(result.keys()):
#                 ws.write(0,col,header)
#             for rn,row in enumerate(result, start=1):
#                 for col, cell in enumerate(row):
#                     ws.write(rn, col, cell)
#                 pbar.update(rn)
#             pbar.finish()
#             p = 0
#         wb.close()
#         print('fetching data from buffer')
#         buffer.seek(0)
#         print('Writing to xlsx file and zipping . . ')
#         with gzip.open(zip_path, 'wb') as f:
#             f.write(buffer.getvalue())
#         ett=datetime.now()
#         print(f'total time taken for zip_path4:{(ett-stt).total_seconds()}')
#
#
# db=next(get_db())
# export_test_1(tablename, zip_path4, db)
y=0
# today=datetime(year=2025,month=3,day=11)
# file_pattern = rf'EOD Position[ _]{today.day}[-_]{today.strftime("%b").capitalize()}[_-]{today.year}' #sample=EOD Position 28-Jan-2025 or EOD Position_11_Mar_2025
# matched_files = [f for f in os.listdir(test_dir) if re.match(file_pattern,f)]
# # matched_files = [f for f in os.listdir(rf"C:\Users\vipulanand\Downloads") if re.match(file_pattern,f)]
i=0
# df = pd.read_excel(rf"D:\notis_analysis\Final_NNF_ID.xlsx", index_col=False)
# df.columns = df.columns.str.replace(r'\s+','',regex=True)
# df.drop(columns=df.columns[df.columns.str.startswith('Un')], inplace=True)
# df.dropna(how='all', inplace=True)
# terminal_list = df.TerminalID.unique().tolist()
# o=0
# sql_server = '172.30.100.41'
# sql_port = '1450'
# sql_db = 'OMNE_ARD_PRD'
# sql_userid = 'Pos_User'
# sql_paswd = 'Pass@Word1'
# sql_paswd_encoded = quote(sql_paswd)
# # sql_query = "select * from [OMNE_ARD_PRD].[dbo].[TradeHist]"
# # sql_query = ("select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser,mnmFillSize, mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice"
# #              "from TradeHist "
# #              "where mnmAccountId='AA100' and mnmExchange='BSE'")
# sql_query = ("select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice from TradeHist where mnmAccountId='AA100' and mnmExchange='BSE'")
# try:
#     sql_engine_str = (
#         f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#         f"SERVER={sql_server},{sql_port};"
#         f"DATABASE={sql_db};"
#         f"UID={sql_userid};"
#         f"PWD={sql_paswd};"
#     )
#     with pyodbc.connect(sql_engine_str) as sql_conn:
#         df_bse=pd.read_sql_query(sql_query,sql_conn)
#     print(f'data fetched for bse: {df.shape}')
# except (pyodbc.Error, psycopg2.Error) as e:
#     print(f'Error in fetching data: {e}')
# # df_bse = read_file(os.path.join(bse_dir,'testbse_18Mar2025.xlsx'))
# df_bse.columns = [re.sub(r'mnm|\s','',each) for each in df_bse.columns]
# df_bse.ExpiryDate = df_bse.ExpiryDate.apply(lambda x:pd.to_datetime(get_date_from_non_jiffy(int(x))).date())
# to_int_list = ['FillPrice', 'FillSize','StrikePrice']
# for each in to_int_list:
#     df_bse[each] = df_bse[each].astype(np.int64)
# df_bse['AvgPrice'] = df_bse['AvgPrice'].astype(float).astype(np.int64)
# df_bse['StrikePrice'] = (df_bse['StrikePrice']/100).astype(np.int64)
# df_bse['Symbol'] = df_bse['TradingSymbol'].apply(lambda x:'SENSEX' if x.upper().startswith('SEN') else x)
# df_bse.rename(columns={'User':'TerminalID'}, inplace=True)
# p=0
# pivot_df = df_bse.pivot_table(
#     index=['TerminalID','Symbol','ExpiryDate','OptionType','StrikePrice'],
#     columns=['TransactionType'],
#     values=['FillSize','AvgPrice'],
#     aggfunc={'FillSize':'sum','AvgPrice':'mean'},
#     fill_value=0
# )
# pivot_df.columns = ['BuyPrc','SellPrc','BuyVol','SellVol']
# pivot_df.reset_index(inplace=True)
# pivot_df['BSEIntradayVol'] = pivot_df.BuyVol - pivot_df.SellVol
o=0
# mod_df = read_file(rf"D:\notis_analysis\modified_data\NOTIS_TRADE_DATA_12MAR2025.csv")
# # to make deskwise, useridwise and nnfwise tables
# p=0
# # to_use = trdqty,trdprc,bsflg,broker,sym,expdt,strprc,opttype,terminalid,subgroup,maingroup
# list_to_int = ['trdQty','strPrc']
# for each in list_to_int:
#     mod_df[each] = mod_df[each].astype(np.int64)
# mod_df['trdPrc'] = mod_df['trdPrc'].astype(np.float64)
# mod_df['expDt'] = pd.to_datetime(mod_df['expDt']).dt.date
# o=0
# pivot_df = mod_df.pivot_table(
#     index=['MainGroup','SubGroup','broker','sym','expDt','strPrc','optType'],
#     columns=['bsFlg'],
#     values=['trdQty','trdPrc'],
#     aggfunc={'trdQty':'sum','trdPrc':'mean'},
#     fill_value=0
# )
# pivot_df.columns = ['BuyAvgPrc','SellAvgPrc','BuyQty','SellQty']
# pivot_df.reset_index(inplace=True)
# pivot_df.rename(columns={'MainGroup':'mainGroup','SubGroup':'subGroup','sym':'symbol','expDt':'expiryDate','strPrc':'strikePrice','optType':'optionType','BuyAvgPrc':'buyAvgPrice',	'SellAvgPrc':'sellAvgPrice','BuyQty':'buyAvgQty','SellQty':'sellAvgQty'})
# pivot_df.volume = pivot_df.buyAvgQty - pivot_df.sellAvgQty
p=0
# # stt = datetime.combine(datetime.today,datetime.datetime.time(9,15))
# stt=datetime.now().replace(hour=9,minute=15)
# # ett=datetime.combine(datetime.today,datetime.datetime.time(15,30))
# ett=datetime.now().replace(hour=15,minute=30)
# while datetime.now()<stt:
#     time.sleep(1)
# while datetime.now()<ett:
#     now=datetime.now()
#     print('\nnow strf 0000 is ',now.strftime('%Y-%m-%d %H:%M')+':00.000')
#     print('complete strf ',now.strftime('%Y-%m-%d %H:%M:%S.%f'),'\n','truncated strf',now.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3])
#     next_min = (now+timedelta(minutes=1)).replace(second=0,microsecond=0)
#     print('next min ',next_min,'\n')
#     time.sleep((next_min-now).total_seconds()+30)
q=0
# import re
# import pandas as pd
# import numpy as np
# from datetime import datetime, timedelta, timezone
# import os
# from dateutil.relativedelta import relativedelta
# import progressbar
# from openpyxl import load_workbook, Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# import pyodbc
# from sqlalchemy import create_engine, text, insert
# import psycopg2
# import time
# import warnings
# from db_config import engine_str, n_tbl_notis_nnf_data
# from common import get_date_from_non_jiffy, get_date_from_jiffy, today, yesterday, root_dir
#
# warnings.filterwarnings("ignore")
#
# def read_data_db(nnf=False, for_table='ENetMIS', from_time:str='', to_time:str=''):
#     if not nnf and for_table == 'ENetMIS':
#         print(f'fetching raw data from {from_time} to {to_time}')
#         # Sql connection parameters
#         sql_server = "rms.ar.db"
#         sql_database = "ENetMIS"
#         sql_username = "notice_user"
#         sql_password = "Notice@2024"
#         # sql_query = "SELECT * FROM [ENetMIS].[dbo].[NSE_FO_AA100_view]"
#         sql_query = f"SELECT * FROM [ENetMIS].[dbo].[NSE_FO_AA100_view] WHERE CreateDate BETWEEN '{from_time}' AND '{to_time}';"
#
#         try:
#             sql_connection_string = (
#                 f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#                 f"SERVER={sql_server};"
#                 f"DATABASE={sql_database};"
#                 f"UID={sql_username};"
#                 f"PWD={sql_password}"
#             )
#             with pyodbc.connect(sql_connection_string) as sql_conn:
#                 df = pd.read_sql_query(sql_query, sql_conn)
#             print(f"Data fetched from SQL Server. Shape:{df.shape}")
#             return df
#
#         except (pyodbc.Error, psycopg2.Error) as e:
#             print("Error occurred:", e)
#
#     elif nnf and for_table != 'ENetMIS':
#         engine = create_engine(engine_str)
#         df = pd.read_sql_table(n_tbl_notis_nnf_data, con=engine)
#         print(f"Data fetched from {for_table} table. Shape:{df.shape}")
#         return df
#
#     elif not nnf and for_table!='ENetMIS':
#         engine = create_engine(engine_str)
#         df = pd.read_sql_table(for_table, con=engine)
#         print(f"Data fetched from {for_table} table. Shape:{df.shape}")
#         return df
#
# def main(from_time:str='', to_time:str=''):
#     print(f'\n\ntrade data fetched for {datetime.now().time()}')
#     # # today = datetime(year=2025, month=3, day=7).date()
#     # df_db = read_data_db(from_time=from_time,to_time=to_time)
#     # print(f'length fetched at {from_time} ',len(df_db))
#     # write_notis_postgredb1(df_db, table_name=n_tbl_test_raw, raw=True)
#     # # df_db = read_data_db(for_table=f'notis_raw_data_{today.strftime("%Y-%m-%d")}')
#     # # write_notis_postgredb(df_db, table_name=n_tbl_notis_raw_data, raw=True)
#     # # modify_filepath = os.path.join(modified_dir, f'NOTIS_TRADE_DATA_{today.strftime("%d%b%Y").upper()}.csv')
#     # nnf_file_path = os.path.join(root_dir, "Final_NNF_ID.xlsx")
#     # # if not os.path.exists(nnf_file_path):
#     # #     raise FileNotFoundError("NNF File not found. Please add the NNF file and try again.")
#     #
#     # # readable_mod_time = datetime.fromtimestamp(os.path.getmtime(nnf_file_path))
#     # # if readable_mod_time.date() == today: # Check if the NNF file is modified today or not
#     # #     print(f'New NNF Data found, modifying the nnf data in db . . .')
#     # #     df_nnf = pd.read_excel(nnf_file_path, index_col=False)
#     # #     df_nnf = df_nnf.loc[:, ~df_nnf.columns.str.startswith('Un')]
#     # #     df_nnf.columns = df_nnf.columns.str.replace(' ', '', regex=True)
#     # #     df_nnf.dropna(how='all', inplace=True)
#     # #     df_nnf = df_nnf.drop_duplicates()
#     # #     write_notis_postgredb(df_nnf, n_tbl_notis_nnf_data, raw=False)
#     # # else:
#     # df_nnf = read_data_db(nnf=True, for_table = n_tbl_notis_nnf_data)
#     # df_nnf = df_nnf.drop_duplicates()
#     #
#     # modified_df = modify_file1(df_db, df_nnf)
#     # write_notis_postgredb1(modified_df, table_name=n_tbl_test_mod, raw=False)
#     modified_df = read_data_db(for_table='test_mod_2025-03-21')
#     desk_db_df = read_data_db(for_table='NOTIS_DESK_WISE_NET_POSITION_2025-03-20')
#     modified_df['expDt'] = pd.to_datetime(modified_df['expDt']).dt.date
#     pivot_df = modified_df.pivot_table(
#         index=['MainGroup','SubGroup','broker','sym','expDt','strPrc','optType'],
#         columns=['bsFlg'],
#         values=['trdQty','trdPrc'],
#         aggfunc={'trdQty':'sum','trdPrc':'mean'},
#         fill_value=0
#     )
#     if modified_df.bsFlg.unique().tolist()[0] == 'B':
#         pivot_df['SellAvgPrc']=0;pivot_df['SellQty']=0
#     elif modified_df.bsFlg.unique().tolist()[0] == 'S':
#         pivot_df['BuyAvgPrc']=0;pivot_df['BuyQty']=0
#     else:
#         pivot_df.columns = ['BuyAvgPrc','SellAvgPrc','BuyQty','SellQty']
#     pivot_df.reset_index(inplace=True)
#     pivot_df.rename(columns={'MainGroup':'mainGroup','SubGroup':'subGroup','sym':'symbol','expDt':'expiryDate','strPrc':'strikePrice','optType':'optionType','BuyAvgPrc':'buyAvgPrice',	'SellAvgPrc':'sellAvgPrice','BuyQty':'buyAvgQty','SellQty':'sellAvgQty'}, inplace=True)
#     pivot_df.volume = pivot_df.buyAvgQty - pivot_df.sellAvgQty
#     w=0
#     # write_notis_data(modified_df, modify_filepath)
#     # write_notis_data(modified_df, rf'C:\Users\vipulanand\Documents\Anand Rathi Financial Services Ltd (Synced)\OneDrive - Anand Rathi Financial Services Ltd\notis_files\NOTIS_TRADE_DATA_{today.strftime("%d%b%Y").upper()}.csv')
#     # print('file saved in modified_data folder')
#     # download_bhavcopy()
#     # print(f'Today\'s bhavcopy downloaded and stored at {bhav_dir}')
#     # print(f'Updating cp-noncp eod table...')
#     # stt = datetime.now()
#     # calc_eod_cp_noncp1()
#     # ett = datetime.now()
#     # print(f'Eod(cp-noncp) updation completed. Total time taken: {(ett - stt).seconds} seconds')
#
# if __name__ == '__main__':
#     main()
i=0
# start_time = datetime.strptime("2025-03-24 09:15", "%Y-%m-%d %H:%M")
# end_time = datetime.strptime("2025-03-24 15:30", "%Y-%m-%d %H:%M")
# def read_data_db(nnf=False, for_table='ENetMIS', from_time:str='', to_time:str=''):
#     if not nnf and for_table == 'ENetMIS':
#         print(f'fetching raw data from {from_time} to {to_time}')
#         # Sql connection parameters
#         sql_server = "rms.ar.db"
#         sql_database = "ENetMIS"
#         sql_username = "notice_user"
#         sql_password = "Notice@2024"
#         sql_query = f"SELECT * FROM [ENetMIS].[dbo].[NSE_FO_AA100_view] WHERE CreateDate BETWEEN '{from_time}' AND '{to_time}';"
#
#         try:
#             sql_connection_string = (
#                 f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#                 f"SERVER={sql_server};"
#                 f"DATABASE={sql_database};"
#                 f"UID={sql_username};"
#                 f"PWD={sql_password}"
#             )
#             with pyodbc.connect(sql_connection_string) as sql_conn:
#                 df = pd.read_sql_query(sql_query, sql_conn)
#             print(f"Data fetched from SQL Server. Shape:{df.shape}")
#             return len(df)
#
#         except (pyodbc.Error, psycopg2.Error) as e:
#             print("Error occurred:", e)
#
#     elif nnf and for_table != 'ENetMIS':
#         engine = create_engine(engine_str)
#         df = pd.read_sql_table(n_tbl_notis_nnf_data, con=engine)
#         print(f"Data fetched from {for_table} table. Shape:{df.shape}")
#         return df
#
#     elif not nnf and for_table!='ENetMIS':
#         engine = create_engine(engine_str)
#         # df = pd.read_sql_table(for_table, con=engine)
#         # print(f"Data fetched from {for_table} table. Shape:{df.shape}")
#         # return df
#         with engine.begin() as conn:
#             res = conn.execute(text(f'select count(*) from "{for_table}" where "CreateDate" BETWEEN \'{from_time}\' AND \'{to_time}\''))
#             row_count = res.scalar()
#         return row_count
# # Generate every minute time from start_time to end_time
# current_time = start_time
# mismatch_list = []
# while current_time <= end_time:
#     print('\n',current_time.strftime("%H:%M"))
#     from_time = current_time.replace(second=0,microsecond=0).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
#     to_time = (current_time + timedelta(minutes=1)).replace(second=0, microsecond=0).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
#     pg_db = 'test_mod_2025-03-24'
#     sql_len = read_data_db(from_time=from_time, to_time=to_time)
#     pg_len = read_data_db(for_table=pg_db, from_time=from_time, to_time=to_time)
#     if sql_len != pg_len:
#         mismatch_list.append((current_time.strftime("%Y-%m-%d %H:%M:%S"), sql_len, pg_len))
#     print(f'SQL Server: {sql_len}, PostgreSQL: {pg_len}')
#     current_time += timedelta(minutes=1)
# print(f'mismatch list is \n {mismatch_list}')
e=0
# def test_func():
#     stt = datetime.now().replace(hour=9, minute=15)
#     ett = datetime.now().replace(hour=15, minute=35)
#     print(f'test started at {datetime.now()}')
#
#     while datetime.now() < stt:
#         time.sleep(1)
#
#     while datetime.now() < ett:
#         now = datetime.now()
#         next_call = (now + timedelta(minutes=1)).replace(second=30, microsecond=0)
#         if now.second >= 30:
#             next_call += timedelta(minutes=1)
#
#         print('\n')
#         print(now.strftime('%Y-%m-%d %H:%M:%S'))
#         time.sleep((next_call - now).total_seconds())
p=0
# from notis_main_per_minute import write_notis_postgredb1
# import tkinter as tk
# from tkinter import messagebox
# import pyttsx3
# def speak_message(message):
#     engine = pyttsx3.init()
#     engine.say(message)
#     engine.runAndWait()
# def show_alert(error_msg):
#     root=tk.Tk()
#     root.withdraw()
#     messagebox.showerror("Error", error_msg)
# def get_bse_trade(from_time, to_time):
#     print(f'fetching bse trade from {from_time} to {to_time}')
#     sql_server = '172.30.100.41'
#     sql_port = '1450'
#     sql_db = 'OMNE_ARD_PRD'
#     sql_userid = 'Pos_User'
#     sql_paswd = 'Pass@Word1'
#     sql_paswd_encoded = quote(sql_paswd)
#     # sql_query = "select * from [OMNE_ARD_PRD].[dbo].[TradeHist]"
#     # sql_query = ("select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser,mnmFillSize, mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice"
#     #              "from TradeHist "
#     #              "where mnmAccountId='AA100' and mnmExchange='BSE'")
#     # sql_query = ("select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker from TradeHist where mnmAccountId='AA100' and mnmExchange='BSE'")
#     sql_query = (f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker from TradeHist where (mnmSymbolName = 'BSXOPT' or mnmSymbolName = 'BSE') and mnmExchangeTime between \'{from_time}\' and \'{to_time}\'")
#     # sql_query = (f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker from TradeHist where (mnmSymbolName = 'BSXOPT' or mnmSymbolName = 'BSE') and mnmExchangeTime between '25-Mar-2025 14:04:00' and '25-Mar-2025 14:05:00'")
#     try:
#         sql_engine_str = (
#             f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#             f"SERVER={sql_server},{sql_port};"
#             f"DATABASE={sql_db};"
#             f"UID={sql_userid};"
#             f"PWD={sql_paswd};"
#         )
#         with pyodbc.connect(sql_engine_str) as sql_conn:
#             df_bse=pd.read_sql_query(sql_query,sql_conn)
#         print(f'data fetched for bse: {df_bse.shape}')
#     # except (pyodbc.Error, psycopg2.Error) as e:
#     #     print(f'Error in fetching data: {e}')
#         df_bse = df_bse.query("mnmTransactionType != 'L'")
#         if df_bse.empty:
#             print(f'No data for {from_time} hence skipping')
#             return
#         df_bse.replace('', 0, inplace=True)
#         # df_bse = read_file(os.path.join(bse_dir,'test_bse172025_1.xlsx'))
#         df_bse.columns = [re.sub(r'mnm|\s','',each) for each in df_bse.columns]
#         # df_bse.ExpiryDate = df_bse.ExpiryDate.apply(lambda x:pd.to_datetime(int(x), unit='s').date() if x !='' else x)
#         df_bse.ExpiryDate = df_bse.ExpiryDate.apply(lambda x:pd.to_datetime(int(x), unit='s').date())
#         # df_bse.replace('', 0, inplace=True)
#         to_int_list = ['FillPrice', 'FillSize','StrikePrice']
#         for each in to_int_list:
#             df_bse[each] = df_bse[each].astype(np.int64)
#         df_bse['AvgPrice'] = df_bse['AvgPrice'].astype(float).astype(np.int64)
#         df_bse['StrikePrice'] = (df_bse['StrikePrice']/100).astype(np.int64)
#         df_bse['Symbol'] = df_bse['TradingSymbol'].apply(lambda x:'SENSEX' if x.upper().startswith('SEN') else x)
#         df_bse.rename(columns={'User':'TerminalID'}, inplace=True)
#         pivot_df = df_bse.pivot_table(
#             index=['TerminalID','Symbol','TradingSymbol','ExpiryDate','OptionType','StrikePrice','ExecutingBroker'],
#             columns=['TransactionType'],
#             values=['FillSize','AvgPrice'],
#             aggfunc={'FillSize':'sum','AvgPrice':'mean'},
#             fill_value=0
#         )
#         if len(df_bse.TransactionType.unique()) == 1:
#             if df_bse.TransactionType.unique().tolist()[0] == 'B':
#                 pivot_df['SellAvgPrc']=0;pivot_df['SellQty']=0
#             elif df_bse.TransactionType.unique().tolist()[0] == 'S':
#                 pivot_df['BuyAvgPrc']=0;pivot_df['BuyQty']=0
#         elif len(df_bse) == 0 or len(pivot_df) == 0:
#             pivot_df.columns = ['_'.join(col).strip() for col in pivot_df.columns.values]
#         pivot_df.columns = ['BuyPrc','SellPrc','BuyVol','SellVol']
#         pivot_df.reset_index(inplace=True)
#         pivot_df['BSEIntradayVol'] = pivot_df.BuyVol - pivot_df.SellVol
#         pivot_df.ExpiryDate = pivot_df.ExpiryDate.astype(str)
#         pivot_df['ExpiryDate'] = [re.sub(r'1970.*','',each) for each in pivot_df['ExpiryDate']]
#         to_int_list = ['BuyPrc','SellPrc','BuyVol','SellVol']
#         for col in to_int_list:
#             pivot_df[col] = pivot_df[col].astype(np.int64)
#         print(f'pivot shape: {pivot_df.shape}')
#         # pivot_df.replace(0,'', inplace=True)
#         # write_notis_data(pivot_df,os.path.join(bse_dir,f'BSE_TRADE_DATA_{today.strftime("%Y-%m-%d")}.xlsx'))
#         # # write_notis_data(pivot_df,os.path.join(bse_dir,f'BSE_TRADE_DATA_{datetime(year=2025,month=3,day=17).strftime("%Y-%m-%d")}.xlsx'))
#         write_notis_postgredb1(pivot_df,table_name=f'test_bse_{today.strftime("%Y-%m-%d")}')
#     except Exception as e:
#         print(f"Exception occured in BSE: {e}")
#         show_alert(error_msg=e)
#         speak_message(message=e)
#
# stt = datetime.now().replace(hour=9, minute=15)
# ett = datetime.now().replace(hour=15, minute=35)
# print(f'test started at {datetime.now()}')
#
# while datetime.now() < stt:
#     time.sleep(1)
#
# while datetime.now() < ett:
#     now = datetime.now()
#     if now.second == 1:
#         print('\n')
#         print('now time => ',now.strftime('%Y-%m-%d %H:%M:%S'))
#         get_bse_trade((now - timedelta(minutes=1)).replace(second=0).strftime('%d-%b-%Y %H:%M:%S'),now.replace(second=0).strftime('%d-%b-%Y %H:%M:%S'))
#         time.sleep(1)  # Sleep for 1 second to avoid multiple prints within the same second
#     else:
#         # Calculate the time to sleep until the next 30th second
#         next_30th_second = (now + timedelta(minutes=1)).replace(second=1,microsecond=0)
#         time.sleep((next_30th_second - now).total_seconds())
# # test_func()
# # now = datetime.now()
# # get_bse_trade(now.replace(second=0).strftime('%d-%b-%Y %H:%M:%S'), (now + timedelta(minutes=1)).replace(second=0).strftime('%d-%b-%Y %H:%M:%S'))
o=0
# from common import read_file
# df = read_file(rf"D:\notis_analysis\bse_data\BSE_TRADE_DATA_ALL_22APR2025.xlsx")
# col_keep = [
#     'ExchSeg','TransactionType','ExchangeTime','AvgPrice','FillSize',
#     'ExecutingBroker','TradingSymbol','ExpiryDate',
#     'OptionType','User','AccountId','StrikePrice'
# ]
# df = df[col_keep]
# df = df.applymap(lambda x: re.sub(r'\s+','',x) if isinstance(x,str) else x)
# chg_format = ['AvgPrice','ExpiryDate','FillSize','StrikePrice']
# for index,col in enumerate(chg_format, start=1):
#     if index == 1:
#         df[col] = df[col].astype(np.float64)
#     elif index == 2:
#         df[col] = df[col].astype('datetime64[ns]').dt.date
#     else:
#         df[col] = df[col].astype(np.int64)
# df = df.round(2)
# df['Broker'] = df['ExecutingBroker'].apply(lambda x: 'CP' if str(x).startswith('Y') else 'non CP')
# df['Symbol'] = df['TradingSymbol'].apply(lambda x: 'SENSEX' if x.startswith('S') else 'BANKEX')
# df.drop(columns=['TradingSymbol'], inplace=True)
# df['trdQtyPrc'] = df['FillSize'] * df['AvgPrice']
# pivot_df = df.pivot_table(
#     index=['Broker','Symbol','ExpiryDate','StrikePrice','OptionType'],
#     columns=['TransactionType'],
#     values=['FillSize','trdQtyPrc'],
#     aggfunc={'FillSize':'sum','trdQtyPrc':'sum'},
#     fill_value=0
# )
# if len(df.TransactionType.unique()) == 1:
#     if df.TransactionType.unique().tolist()[0] == 'B':
#         pivot_df['SellTrdQtyPrc'],pivot_df['SellQty'] = 0,0
#     else:
#         pivot_df['BuyTrdQtyPrc'],pivot_df['BuyQty'] = 0,0
# elif len(df) == 0 or len(pivot_df) == 0:
#     pivot_df.columns = ['_'.join(col).strip() for col in pivot_df.columns.values]
# pivot_df.columns = ['BuyQty','SellQty','BuyTrdQtyPrc','SellTrdQtyPrc']
# pivot_df['BuyAvgPrc'] = pivot_df.apply(lambda row: row['BuyTrdQtyPrc']/row['BuyQty'] if row['BuyQty']>0 else 0, axis=1)
# pivot_df['SellAvgPrc'] = pivot_df.apply(lambda row: row['SellTrdQtyPrc']/row['SellQty'] if row['SellQty']>0 else 0, axis=1)
# pivot_df.drop(columns=['SellTrdQtyPrc','BuyTrdQtyPrc'], inplace=True)
# pivot_df.reset_index(inplace=True)
u=0
# from common import get_date_from_jiffy_new, read_file
# df=read_file(rf"D:\notis_analysis\testing\rawtradebook_2025-04-17.xlsx")
p=0
# from common import read_data_db
# for_date = datetime.today().date().replace(day=28, month=4)
# nse_cp_noncp = read_data_db(for_table=f'NOTIS_EOD_NET_POS_CP_NONCP_{for_date}')
# bse_trade_data = read_data_db(for_table=f'BSE_TRADE_DATA_{for_date}')
# bse_trade_data['Broker'] = bse_trade_data['ExecutingBroker'].apply(lambda x: 'CP' if x.startswith("Y") else 'non CP')
o=0
# def read_data_db(nnf=False, for_table='ENetMIS', from_time:str='', to_time:str='', from_source=False):
#     global engine
#     if not nnf and for_table == 'ENetMIS':
#         # Sql connection parameters
#         sql_server = "rms.ar.db"
#         sql_database = "ENetMIS"
#         sql_username = "notice_user"
#         sql_password = "Notice@2024"
#         if not from_time:
#             sql_query = "SELECT * FROM [ENetMIS].[dbo].[NSE_FO_AA100_view]"
#         else:
#             sql_query = f"SELECT * FROM [ENetMIS].[dbo].[NSE_FO_AA100_view] WHERE CreateDate BETWEEN '{from_time}' AND '{to_time}';"
#         # if from_source:
#         #     sql_query = f"""
#         #                     WITH CTE AS (
#         #                         SELECT *,
#         #                                ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS RowNum
#         #                         FROM [ENetMIS].[dbo].[NSE_FO_AA100_view]
#         #                     )
#         #                     SELECT *
#         #                     FROM CTE
#         #                     WHERE RowNum > {offset} AND RowNum <= {offset + page_size};
#         #                     """
#         try:
#             sql_connection_string = (
#                 f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#                 f"SERVER={sql_server};"
#                 f"DATABASE={sql_database};"
#                 f"UID={sql_username};"
#                 f"PWD={sql_password}"
#             )
#             with pyodbc.connect(sql_connection_string) as sql_conn:
#                 df = pd.read_sql_query(sql_query, sql_conn)
#             print(f"Data fetched from SQL Server. Shape:{df.shape}")
#             return df
#         except (pyodbc.Error, psycopg2.Error) as e:
#             print("Error occurred:", e)
#     elif nnf and for_table != 'ENetMIS':
#         # engine = create_engine(engine_str)
#         with engine.begin() as conn:
#             df = pd.read_sql_table(n_tbl_notis_nnf_data, con=conn)
#         print(f"Data fetched from {for_table} table. Shape:{df.shape}")
#         return df
#     elif not nnf and for_table == 'TradeHist':
#         sql_server = '172.30.100.41'
#         sql_port = '1450'
#         sql_db = 'OMNE_ARD_PRD'
#         sql_userid = 'Pos_User'
#         sql_paswd = 'Pass@Word1'
#         if not from_time:
#             print(f'Fetching today\'s BSE trade data till now.')
#             # sql_query = (
#             #     f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker from TradeHist where (mnmSymbolName = 'BSXOPT' or mnmSymbolName = 'BSE')")
#             sql_query = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#             sql_query2 = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD_HNI].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#         else:
#             print(f'Fetching BSE trade data from {from_time} to {to_time}')
#             # sql_query = (
#             #     f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker from TradeHist where (mnmSymbolName = 'BSXOPT' or mnmSymbolName = 'BSE') and mnmExchangeTime between \'{from_time}\' and \'{to_time}\'")
#             sql_query = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#             sql_query2 = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD_HNI].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#         try:
#             sql_engine_str = (
#                 f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#                 f"SERVER={sql_server},{sql_port};"
#                 f"DATABASE={sql_db};"
#                 f"UID={sql_userid};"
#                 f"PWD={sql_paswd};"
#             )
#             with pyodbc.connect(sql_engine_str) as sql_conn:
#                 df_bse = pd.read_sql_query(sql_query, sql_conn)
#                 df_bse_hni = pd.read_sql_query(sql_query2,sql_conn)
#             print(f'data fetched for bse: {df_bse.shape, df_bse_hni.shape}')
#             final_bse_df = pd.concat([df_bse,df_bse_hni], ignore_index=True)
#             return final_bse_df
#         except (pyodbc.Error, psycopg2.Error) as e:
#             print(f'Error in fetching data: {e}')
#     elif not nnf and for_table!='ENetMIS':
#         # engine = create_engine(engine_str)
#         with engine.begin() as conn:
#             df = pd.read_sql_table(for_table, con=conn)
#         print(f"Data fetched from {for_table} table. Shape:{df.shape}")
#         return df
# server = '172.30.100.40,1450'
# database = OMNE_ARD_PRD_3.19
# database_1 = OMNE_ARD_PRD_AA100_3.19
t=0
# # to get trade for 2 nnf ids
# from common import read_data_db
# holidays_25 = ['2025-02-26', '2025-03-14', '2025-03-31', '2025-04-10', '2025-04-14', '2025-04-18', '2025-05-01', '2025-08-15', '2025-08-27', '2025-10-02', '2025-10-21', '2025-10-22', '2025-11-05', '2025-12-25']
# start_date = datetime.today().date().replace(day=1,month=4)
# end_date = datetime.now().date().replace(day=30,month=4)
# b_days = pd.bdate_range(start=start_date, end=end_date, freq='C', weekmask='1111100', holidays=holidays_25).date.tolist()
# main_df = pd.DataFrame()
# for each_date in b_days:
#     table_name = f"NOTIS_NNF_WISE_NET_POSITION_{each_date}"
#     df = read_data_db(for_table=table_name)
#     df1 = df.query("nnfID == 111111111111122 or nnfID == 400013041212000")
#     df1['for_date'] = each_date
#     main_df = pd.concat([main_df,df1], ignore_index=True)
#
# main_df['NetQty'] = main_df['buyAvgQty'] - main_df['sellAvgQty']
# main_df['tradeValue'] = (main_df['sellAvgPrice']*main_df['sellAvgQty']) - (main_df['buyAvgPrice']*main_df['buyAvgQty'])
# main_df['nnfID'] = main_df['nnfID'].astype(str)
e=0
# from common import today, yesterday, read_data_db
# from db_config import n_tbl_notis_nnf_data
# from nse_utility import NSEUtility
# table_name = f'NOTIS_EOD_NET_POS_CP_NONCP_{yesterday.strftime("%Y-%m-%d")}'
# df = read_data_db(for_table=table_name)
# df_nnf = read_data_db(nnf=True, for_table = n_tbl_notis_nnf_data)
# df_nnf = df_nnf.drop_duplicates()
# nse_df = read_data_db()
# modified_nse = NSEUtility.modify_file(df=nse_df,df_nnf=df_nnf)
w=0
# from common import read_data_db, read_file, write_notis_postgredb
# import re
#
# main_bse_df = pd.DataFrame()
# today=datetime.today().date().replace(day=7)
# yesterday=today-timedelta(days=1)

# def get_oi(for_date = Query()):
#     for_date = datetime.today().date().strftime('%Y-%m-%d')
#     table_to_read = f'NOTIS_EOD_NET_POS_CP_NONCP_{for_date}'
#     eod_df = read_data_db(for_table=table_to_read)
#     eod_df.columns = [re.sub(rf'Eod|\s','',each) for each in eod_df.columns]
#     grouped_df = eod_df.groupby(by=['Broker','Underlying','Expiry'], as_index=False).agg({'FinalNetQty':lambda x:x.abs().sum()})
#     json_data = grouped_df.to_json(orient='records')
#     return grouped_df
# get_oi()
u=0
# from bse_utility import BSEUtility
# bse_df = BSEUtility.get_bse_trade_data()

# def bse_modify_file(bse_raw_df):
#     bse_raw_df = bse_raw_df.query("mnmTransactionType != 'L'")
#     bse_raw_df.replace('', 0, inplace=True)
#     bse_raw_df.columns = [re.sub(r'mnm|\s', '', each) for each in bse_raw_df.columns]
#     bse_raw_df.ExpiryDate = bse_raw_df.ExpiryDate.apply(lambda x: pd.to_datetime(int(x), unit='s').date())
#     to_int_list = ['FillPrice', 'FillSize', 'StrikePrice']
#     for each in to_int_list:
#         bse_raw_df[each] = bse_raw_df[each].astype(np.int64)
#     bse_raw_df['AvgPrice'] = bse_raw_df['AvgPrice'].astype(float).round(2)
#     bse_raw_df['StrikePrice'] = (bse_raw_df['StrikePrice'] / 100).astype(np.int64)
#     bse_raw_df['Symbol'] = bse_raw_df['TradingSymbol'].apply(lambda x: 'SENSEX' if x.upper().startswith('SEN') else x)
#     bse_raw_df['Broker'] = bse_raw_df['AccountId'].apply(lambda x: 'non CP' if x.upper().startswith('AA') else 'CP')
#     bse_raw_df.rename(columns={'User': 'TerminalID','Symbol':'Underlying','ExpiryDate':'Expiry', 'StrikePrice':'Strike'}, inplace=True)
#     return bse_raw_df
#
# def calc_bse_eod_net_pos(desk_bse_df):
#     # read prev day eod table and group it
#     # read today's data and group it
#     # merge both grouped data, yesterday>today
#     eod_df = read_data_db(for_table=f'NOTIS_EOD_NET_POS_CP_NONCP_{yesterday.strftime("%Y-%m-%d")}')
#     eod_df = eod_df.replace(' ','',regex=True)
#     eod_df.columns = [re.sub(rf'Eod|\s','',each) for each in eod_df.columns]
#     eod_df.Expiry = pd.to_datetime(eod_df.Expiry, dayfirst=True, format='mixed').dt.date
#     eod_df.drop(
#         columns=['NetQuantity', 'buyQty', 'buyAvgPrice', 'sellQty', 'sellAvgPrice', 'IntradayVolume', 'ClosingPrice'],
#         inplace=True
#     )
#     eod_df.rename(columns={'FinalNetQty': 'NetQuantity', 'FinalSettlementPrice': 'ClosingPrice'}, inplace=True)
#     eod_df = eod_df.add_prefix('Eod')
#     nse_underlying_list = ['NIFTY','BANKNIFTY','MIDCPNIFTY','FINNIFTY']
#     eod_df = eod_df.query("EodUnderlying == 'SENSEX' and EodExpiry > @today and EodNetQuantity != 0 and EodUnderlying in @nse_underlying_list")
#     grouped_eod_df = eod_df.groupby(by=['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType'], as_index=False).agg({'EodNetQuantity':'sum','EodClosingPrice':'mean'})
#     # ============================================================================================
#     grouped_desk_df = desk_bse_df.groupby(by=['Broker','Underlying','Expiry','Strike','OptionType'], as_index=False).agg({'BuyQty':'sum','SellQty':'sum','buyAvgPrice':'mean','sellAvgPrice':'mean','IntradayVolume':'sum'})
#     # grouped_desk_df['IntradayVolume'] = grouped_desk_df['BuyQty'] - grouped_desk_df['SellQty']
#     # ============================================================================================
#     merged_df = grouped_eod_df.merge(
#         grouped_desk_df,
#         left_on=['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType'],
#         right_on=['Broker', 'Underlying', 'Expiry', 'Strike', 'OptionType'],
#         how='outer'
#     )
#     merged_df.fillna(0,inplace=True)
#     merged_df.drop_duplicates(inplace=True)
#     # ============================================================================================
#     coltd1 = ['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType']
#     coltd2 = ['Broker','Underlying','Expiry','Strike','OptionType']
#     for i in range(len(coltd1)):
#         merged_df.loc[merged_df[coltd1[i]] == 0, coltd1[i]] = merged_df[coltd2[i]]
#         merged_df.loc[merged_df[coltd2[i]] == 0, coltd2[i]] = merged_df[coltd1[i]]
#     merged_df['FinalNetQty'] = merged_df['EodNetQuantity'] + merged_df['IntradayVolume']
#     merged_df['FinalSettlementPrice'] = 0
#     merged_df.drop(columns=['Broker','Underlying','Expiry','Strike','OptionType'], inplace=True)
#     # ============================================================================================
#     col_to_int = ['BuyQty', 'SellQty','FinalSettlementPrice']
#     for col in col_to_int:
#         merged_df[col] = merged_df[col].astype(np.int64)
#     print(f'length of cp noncp for {today} is {merged_df.shape}')
#     return merged_df
#
# def add_to_bse_eod_net_pos(for_date):
#     if not for_date:
#         print(f'for_date is empty')
#     else:
#         sent_df = read_file(rf"D:\notis_analysis\eod_original\EOD Net positions {for_date.strftime('%d%m%Y')} BSE.xlsx")
#         sent_df.columns = [re.sub(rf'\s|\.','',each) for each in sent_df.columns]
#         sent_df.ExpiryDate = pd.to_datetime(sent_df.ExpiryDate, dayfirst=True, format='mixed').dt.date
#         sent_df['Broker'] = sent_df.apply(lambda row: 'CP' if row['PartyCode'].upper().endswith('CP') else 'non CP', axis=1)
#         sent_df['OptionType'] = sent_df.apply(lambda row: 'XX' if row['OptionType'].upper().startswith('F') else row['OptionType'], axis=1)
#         sent_df.drop(columns=['PartyCode'], inplace=True)
#         sent_df.rename(columns={'Symbol':'Underlying','ExpiryDate':'Expiry','StrikePrice':'Strike'}, inplace=True)
#         sent_df = sent_df.add_prefix('Eod')
#         sent_df.rename(columns={'EodNetQty':'FinalNetQty'}, inplace=True)
#         col_to_add = ['EodNetQuantity','EodClosingPrice','buyQty','buyAvgPrice','sellQty','sellAvgPrice','IntradayVolume','FinalSettlementPrice']
#         for col in col_to_add:
#             sent_df[col]=0
#         truncated_sent_df = sent_df.query('EodUnderlying == "SENSEX"')
#
#         eod_df = read_data_db(for_table=f'NOTIS_EOD_NET_POS_CP_NONCP_{for_date}')
#         eod_df.EodExpiry = pd.to_datetime(eod_df.EodExpiry, dayfirst=True, format='mixed').dt.date
#
#         concat_eod_df = pd.concat([eod_df,truncated_sent_df], ignore_index=True)
#         write_notis_postgredb(df=concat_eod_df,table_name=f'NOTIS_EOD_NET_POS_CP_NONCP_{for_date}',truncate_required=True)
#     u=0

# raw_bse_df = read_data_db(for_table='TradeHist')
# # raw_bse_df1 = read_data_db(for_table=f'BSE_TRADE_DATA_{today.strftime("%Y-%m-%d")}')
# modified_bse_df1 = bse_modify_file(raw_bse_df)
# main_bse_df = pd.concat([main_bse_df,modified_bse_df1],ignore_index=True)
# # main_bse_df = pd.concat([main_bse_df,raw_bse_df1],ignore_index=True)
# main_bse_df['trdQtyPrc'] = main_bse_df['FillSize']*main_bse_df['AvgPrice']
# pivot_df = main_bse_df.pivot_table(
#     index=['Broker','Underlying', 'Expiry', 'Strike', 'OptionType'],
#     columns=['TransactionType'],
#     values=['FillSize', 'trdQtyPrc'],
#     aggfunc={'FillSize': 'sum', 'trdQtyPrc': 'sum'},
#     fill_value=0
# )
# if len(main_bse_df.TransactionType.unique()) == 1:
#     if main_bse_df.TransactionType.unique().tolist()[0] == 'B':
#         pivot_df['SellTrdQtyPrc'] = 0;
#         pivot_df['SellQty'] = 0
#     elif main_bse_df.TransactionType.unique().tolist()[0] == 'S':
#         pivot_df['BuyTrdQtyPrc'] = 0;
#         pivot_df['BuyQty'] = 0
# elif len(main_bse_df) == 0 or len(pivot_df) == 0:
#     pivot_df.columns = ['_'.join(col).strip() for col in pivot_df.columns.values]
# pivot_df.columns = ['BuyQty', 'SellQty','BuyTrdQtyPrc', 'SellTrdQtyPrc']
# pivot_df.reset_index(inplace=True)
# pivot_df['buyAvgPrice'] = pivot_df.apply(lambda row: row['BuyTrdQtyPrc']/row['BuyQty'] if row['BuyQty']>0 else 0, axis=1)
# pivot_df['sellAvgPrice'] = pivot_df.apply(lambda row: row['SellTrdQtyPrc']/row['SellQty'] if row['SellQty']>0 else 0, axis=1)
# pivot_df.drop(columns=['BuyTrdQtyPrc', 'SellTrdQtyPrc'], inplace=True)
# pivot_df['IntradayVolume'] = pivot_df.BuyQty - pivot_df.SellQty
# pivot_df=pivot_df.round(2)
# eod_bse_df = calc_bse_eod_net_pos(pivot_df)
# add_to_bse_eod_net_pos(today)
i=0
# from common import read_data_db, yesterday, today
# # nse_empty,bse_empty = False, False
# # eod_table_name = f'NOTIS_EOD_NET_POS_CP_NONCP_{yesterday.strftime("%Y-%m-%d")}'
# # prev_eod = read_data_db(for_table=eod_table_name)
# # prev_eod.EodExpiry = pd.to_datetime(prev_eod.EodExpiry, dayfirst=True, format='mixed').dt.date
# eod_bse_df = pd.DataFrame()
# def a():
#     return eod_bse_df
# ret = a()
p=0
# from bse_utility import BSEUtility
# def read_data_db(nnf=False, for_table='ENetMIS', from_time:str='', to_time:str='', from_source=False):
#     if not nnf and for_table == 'ENetMIS':
#         # Sql connection parameters
#         sql_server = "rms.ar.db"
#         sql_database = "ENetMIS"
#         sql_username = "notice_user"
#         sql_password = "Notice@2024"
#         if not from_time:
#             sql_query = "SELECT * FROM [ENetMIS].[dbo].[NSE_FO_AA100_view]"
#         else:
#             sql_query = f"SELECT * FROM [ENetMIS].[dbo].[NSE_FO_AA100_view] WHERE CreateDate BETWEEN '{from_time}' AND '{to_time}';"
#         # if from_source:
#         #     sql_query = f"""
#         #                     WITH CTE AS (
#         #                         SELECT *,
#         #                                ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS RowNum
#         #                         FROM [ENetMIS].[dbo].[NSE_FO_AA100_view]
#         #                     )
#         #                     SELECT *
#         #                     FROM CTE
#         #                     WHERE RowNum > {offset} AND RowNum <= {offset + page_size};
#         #                     """
#         try:
#             sql_connection_string = (
#                 f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#                 f"SERVER={sql_server};"
#                 f"DATABASE={sql_database};"
#                 f"UID={sql_username};"
#                 f"PWD={sql_password}"
#             )
#             with pyodbc.connect(sql_connection_string) as sql_conn:
#                 df = pd.read_sql_query(sql_query, sql_conn)
#             print(f"Data fetched from SQL Server. Shape:{df.shape}")
#             return df
#         except (pyodbc.Error, psycopg2.Error) as e:
#             print("Error occurred:", e)
#     elif nnf and for_table != 'ENetMIS':
#         # engine = create_engine(engine_str)
#         with engine.begin() as conn:
#             df = pd.read_sql_table(n_tbl_notis_nnf_data, con=conn)
#         print(f"Data fetched from {for_table} table. Shape:{df.shape}")
#         return df
#     elif not nnf and for_table == 'TradeHist':
#         sql_server = '172.30.100.41'
#         sql_port = '1450'
#         sql_db = 'OMNE_ARD_PRD'
#         sql_userid = 'Pos_User'
#         sql_paswd = 'Pass@Word1'
#
#         if not from_time:
#             print(f'Fetching today\'s BSE trade data till now.')
#             # sql_query = (
#             #     f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker from TradeHist where (mnmSymbolName = 'BSXOPT' or mnmSymbolName = 'BSE')")
#             sql_query = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#             sql_query2 = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD_HNI].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#         else:
#             print(f'Fetching BSE trade data from {from_time} to {to_time}')
#             sql_query = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#             sql_query2 = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD_HNI].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#         try:
#             sql_engine_str = (
#                 f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#                 f"SERVER={sql_server},{sql_port};"
#                 f"DATABASE={sql_db};"
#                 f"UID={sql_userid};"
#                 f"PWD={sql_paswd};"
#             )
#             with pyodbc.connect(sql_engine_str) as sql_conn:
#                 df_bse = pd.read_sql_query(sql_query, sql_conn)
#                 df_bse_hni = pd.read_sql_query(sql_query2,sql_conn)
#             print(f'data fetched for bse: {df_bse.shape, df_bse_hni.shape}')
#             final_bse_df = pd.concat([df_bse,df_bse_hni], ignore_index=True)
#             return final_bse_df
#         except (pyodbc.Error, psycopg2.Error) as e:
#             print(f'Error in fetching data: {e}')
#     elif not nnf and for_table == 'TradeHist1':
#         sql_server = '172.30.100.40'
#         sql_port = '1450'
#         sql_db = 'OMNE_ARD_PRD_3.19'
#         sql_userid = 'Pos_User'
#         sql_paswd = 'Pass@Word1'
#         if not from_time:
#             print(f'Fetching today\'s BSE trade data till now.')
#             # sql_query = (
#             #     f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker from TradeHist where (mnmSymbolName = 'BSXOPT' or mnmSymbolName = 'BSE')")
#             sql_query = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD_3.19].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#             sql_query2 = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD_AA100_3.19].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#         else:
#             print(f'Fetching BSE trade data from {from_time} to {to_time}')
#             sql_query = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#             sql_query2 = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker "
#                 f"from [OMNE_ARD_PRD_HNI].[dbo].[TradeHist] "
#                 f"where mnmExchSeg = 'bse_fo' "
#                 f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#         try:
#             sql_engine_str = (
#                 f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#                 f"SERVER={sql_server},{sql_port};"
#                 f"DATABASE={sql_db};"
#                 f"UID={sql_userid};"
#                 f"PWD={sql_paswd};"
#             )
#             with pyodbc.connect(sql_engine_str) as sql_conn:
#                 df_bse = pd.read_sql_query(sql_query, sql_conn)
#                 df_bse_hni = pd.read_sql_query(sql_query2,sql_conn)
#             print(f'data fetched for bse: {df_bse.shape, df_bse_hni.shape}')
#             final_bse_df = pd.concat([df_bse,df_bse_hni], ignore_index=True)
#             return final_bse_df
#         except (pyodbc.Error, psycopg2.Error) as e:
#             print(f'Error in fetching data: {e}')
#     elif not nnf and for_table!='ENetMIS':
#         # engine = create_engine(engine_str)
#         with engine.begin() as conn:
#             df = pd.read_sql_table(for_table, con=conn)
#         print(f"Data fetched from {for_table} table. Shape:{df.shape}")
#         return df
# ard_prd_df = read_data_db(for_table='TradeHist')
# ard_prd_319_df = read_data_db(for_table='TradeHist1')
# ard_prd_mod = BSEUtility.bse_modify_file(ard_prd_df)
# ard_prd_319_mod = BSEUtility.bse_modify_file(ard_prd_319_df)
o=0
# from nse_utility import NSEUtility
# from bse_utility import BSEUtility
# from db_config import n_tbl_notis_nnf_data, engine_str
# def read_data_db(nnf=False, for_table='ENetMIS', from_time:str='', to_time:str='', from_source=False):
#     if not nnf and for_table == 'ENetMIS':
#         # Sql connection parameters
#         sql_server = "rms.ar.db"
#         sql_database = "ENetMIS"
#         sql_username = "notice_user"
#         sql_password = "Notice@2024"
#         if not from_time:
#             sql_query = "SELECT * FROM [ENetMIS].[dbo].[NSE_FO_AA100_view]"
#         else:
#             sql_query = f"SELECT * FROM [ENetMIS].[dbo].[NSE_FO_AA100_view] WHERE CreateDate BETWEEN '{from_time}' AND '{to_time}';"
#         try:
#             sql_connection_string = (
#                 f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#                 f"SERVER={sql_server};"
#                 f"DATABASE={sql_database};"
#                 f"UID={sql_username};"
#                 f"PWD={sql_password}"
#             )
#             with pyodbc.connect(sql_connection_string) as sql_conn:
#                 df = pd.read_sql_query(sql_query, sql_conn)
#             print(f"Data fetched from SQL Server. Shape:{df.shape}")
#             return df
#         except (pyodbc.Error, psycopg2.Error) as e:
#             print("Error occurred:", e)
#     elif not nnf and for_table == 'Source_2':
#         sql_server = '172.30.100.40'
#         sql_port = '1450'
#         sql_db = 'OMNE_ARD_PRD_AA100_3.19'
#         sql_userid = 'Pos_User'
#         sql_paswd = 'Pass@Word1'
#         if not from_time:
#             print(f'Fetching today\'s NSE&BSE trades from Source:2 till now.')
#             sql_query = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker, mnmExchangeTime "
#                 f"from [OMNE_ARD_PRD_3.19].[dbo].[TradeHist] "
#                 f"where mnmSymbolName in ('NIFTY','BANKNIFTY','MIDCPNIFTY','FINNIFTY','SENSEX') "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#             sql_query2 = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker, mnmExchangeTime "
#                 f"from [OMNE_ARD_PRD_AA100_3.19].[dbo].[TradeHist] "
#                 f"where mnmSymbolName in ('NIFTY','BANKNIFTY','MIDCPNIFTY','FINNIFTY','SENSEX') "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#         else:
#             print(f'Fetching NSE&BSE trade data(Source:2) from {from_time} to {to_time}')
#             sql_query = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker,mnmExchangeTime "
#                 f"from [OMNE_ARD_PRD_3.19].[dbo].[TradeHist] "
#                 f"where mnmSymbolName in ('NIFTY','BANKNIFTY','MIDCPNIFTY','FINNIFTY','SENSEX') "
#                 f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#             sql_query2 = (
#                 f"select mnmFillPrice,mnmSegment, mnmTradingSymbol,mnmTransactionType,mnmAccountId,mnmUser , mnmFillSize, "
#                 f"mnmSymbolName, mnmExpiryDate, mnmOptionType, mnmStrikePrice, mnmAvgPrice, mnmExecutingBroker, mnmExchangeTime "
#                 f"from [OMNE_ARD_PRD_AA100_3.19].[dbo].[TradeHist] "
#                 f"where mnmSymbolName in ('NIFTY','BANKNIFTY','MIDCPNIFTY','FINNIFTY','SENSEX') "
#                 f"and mnmExchangeTime between \'{from_time}\' and \'{to_time}\' "
#                 f"and (mnmAccountId = 'AA100' or mnmAccountId = 'CPAA100')")
#         try:
#             sql_engine_str = (
#                 f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#                 f"SERVER={sql_server},{sql_port};"
#                 f"DATABASE={sql_db};"
#                 f"UID={sql_userid};"
#                 f"PWD={sql_paswd};"
#             )
#             with pyodbc.connect(sql_engine_str) as sql_conn:
#                 df_bse = pd.read_sql_query(sql_query, sql_conn)
#                 df_bse_hni = pd.read_sql_query(sql_query2,sql_conn)
#             print(f'data fetched for bse: {df_bse.shape, df_bse_hni.shape}')
#             final_bse_df = pd.concat([df_bse,df_bse_hni], ignore_index=True)
#             return final_bse_df
#         except (pyodbc.Error, psycopg2.Error) as e:
#             print(f'Error in fetching data: {e}')
#     elif nnf and for_table != 'ENetMIS':
#         engine = create_engine(engine_str, pool_size = 20, max_overflow = 10)
#         with engine.begin() as conn:
#             df = pd.read_sql_table(n_tbl_notis_nnf_data, con=conn)
#         print(f"Data fetched from {for_table} table. Shape:{df.shape}")
#         return df
# old_nse_db = read_data_db(for_table='ENetMIS')
# df_nnf = read_data_db(nnf=True, for_table='nnf_data')
# modified_enet_db = NSEUtility.modify_file(df=old_nse_db,df_nnf=df_nnf)
# both_db = read_data_db(for_table='Source_2')
# modified_new_nse_db = BSEUtility.bse_modify_file(both_db)
# def convert_to_modified_notis_format(nse_new_source_db):
#     nse_new_source_db.columns = [re.sub(rf'mnm|\s','',each) for each in nse_new_source_db.columns]
#     nse_new_source_db.rename({
#         'ExchangeTime':'CreateDate','SymbolName':'sym','FillSize':'trdQty','TransactionType':'bsFlg',
#         'AvgPrice':'trdPrc','ExecutingBroker':'cpCD','Strike':'strPrc','Expiry':'expDt','OptionType':'optType',
#         'AccountId':'cliActNo','Broker':'broker'
#     })
p=0
# from common import revise_eod_net_pos
# revise_eod_net_pos(for_dt='2025-05-16')
o=0
# from pandas.testing import assert_frame_equal
#
# df1 = pd.read_excel(os.path.join(table_dir,f'BSE_TRADE_DATA_2025-05-29.xlsx'), index_col=False)
# df2 = pd.read_excel(os.path.join(bse_dir,f'BSE_TRADE_DATA_29MAY2025.xlsx'), index_col=False)
# res=df1.equals(df2)
# print(f'res equal={res}')
# # res_df = df1.compare(df2)
# res_np=np.array_equal(df1.values,df2.values)
# print(f'res np={res_np}')
# sorted_df1 = df1.sort_index(axis=1).sort_values(by=df1.columns.tolist()).reset_index(inplace=True)
# sorted_df1 = df1.sort_index(axis=1).sort_values(by=sorted(df1.columns.tolist())).reset_index(drop=True)
# sorted_df2 = df2.sort_index(axis=1).sort_values(by=sorted(df2.columns.tolist())).reset_index(drop=True)
# res_sorted=sorted_df1.equals(sorted_df2)
# print(f'res sorted={res_sorted}')
# res_compare = sorted_df1.compare(sorted_df2, align_axis=True)
# print(f'res compare= {res_compare}')
# try:
#     assert_frame_equal(sorted_df1,sorted_df2)
#     print(f'Assert DF Equal')
# except AssertionError as e:
#     print(f'Assert Df not Eq: {e}')
# res_avgprc=df1['AvgPrice'].equals(df2['AvgPrice'])
# print(f'res avg price={res_avgprc}')
# for col in df1.columns:
#     print(df1[col].equals(df2[col]))
# sorted_df12 = df1.apply(sorted, axis=0)
# sorted_df22 = df2.apply(sorted, axis=0)
# res_sorted2 = sorted_df12.equals(sorted_df22)
# print(f'res sorted2={res_sorted2}')
y=0
# to convert source2 expiry wheremnmTradingSymbol.startswith('SEN')
# from dateutil.relativedelta import relativedelta
# for each in raw_combined_trade_source2.mnmExpiryDate.unique():
#     new_dt = pd.to_datetime(each, unit='s') + relativedelta(years=10)
#     while new_dt.weekday() != 3:
#         new_dt -= timedelta(days=1)
#     print(f'raw={each}\tepoch_converted={new_dt}')
p=0

# import yfinance as yf
#
# nifty = yf.Ticker("^NSEI") # NIFTY 50
# sensex = yf.Ticker("^BSESN") # SENSEX
#
# print(nifty.history(period="2d")['Close'])
# print(sensex.history(period="2d")['Close'])
u=0
# from db_config import n_tbl_notis_eod_net_pos_cp_noncp
# from common import read_data_db
# df = read_data_db(for_table=n_tbl_notis_eod_net_pos_cp_noncp)
p=0
# from common import bhav_dir, yesterday
# import paramiko
#
# for_date = datetime(year=2025,month=6,day=27).date()
# def download_bhavcopy():
#     host = '192.168.112.81'
#     username = 'greek'
#     password = 'greeksoft'
#     filename = f"regularNSEBhavcopy_{for_date.strftime('%d%m%Y')}.csv"  # sample=regularBhavcopy_13022025
#     remote_path = rf'/home/greek/NSE_BSE_Broadcast/NSE/Bhavcopy/Files/{filename}'
#     local_path = os.path.join(bhav_dir, filename)
#     try:
#         transport = paramiko.Transport((host, 22))
#         transport.connect(username=username, password=password)
#         sftp = paramiko.SFTPClient.from_transport(transport)
#         sftp.get(remote_path, local_path)
#         sftp.close()
#         transport.close()
#         print(f'Bhavcopy for {for_date} downloaded to local server.')
#     except Exception as e:
#         print(f'Error: {e}')
# download_bhavcopy()
r=0
# #SRC2
# import pandas as pd
# import requests, json, os
# from datetime import datetime
# from sqlalchemy import create_engine
# from common import read_data_db, test_dir
# from db_config import inhouse_engine_str
#
# root_dir = os.getcwd()
#
# def get_src2_trade():
#     server_ip = "192.168.50.68"
#     adminusername = "user3_rms"
#     adminpassword = "user3_rms"
#     userid = 201
#
#     def get_token():
#         url = f"http://{server_ip}:8010/v1/loginrms"
#         payload = json.dumps({
#             "username": adminusername,
#             "password": adminpassword
#         })
#         headers = {
#             'accept': 'application/json',
#             'Content-Type': 'application/json'
#         }
#
#         response = requests.request("POST", url, headers=headers, data=payload)
#         response_msg = response.json()
#         authtoken = response_msg.get("token")
#         return authtoken
#
#     def get_algo2_trades():
#         url = f"http://{server_ip}:8010/v1/dcNetposition?user_id={userid}"
#         payload = {}
#         headers = {
#             'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
#                           'AppleWebKit/537.36 (KHTML, like Gecko) '
#                           'Chrome/137.0.0.0 Safari/537.36',
#             'Accept': 'application/json',
#             'auth-token': get_token()
#         }
#         response = requests.request("GET", url, headers=headers, data=payload)
#
#         if response.status_code == 200:
#             response = response.json()
#             data = response['data']
#             df = pd.DataFrame(data)
#             return df
#         else:
#             return pd.DataFrame()
#
#     algo2_df = get_algo2_trades()
#     if not algo2_df.empty:
#         algo2_df.rename(columns={'strikePrice': 'StrikePrice'}, inplace=True)
#         # print(f'SRC2\n{algo2_df}')
#         print('data fetched from src2(68 server)')
#         return algo2_df
#     else:
#         print('No Algo2 trades.')
#         return pd.DataFrame()
#
# # SRC1
# def get_src1_trade():
#     inhouse_engine = create_engine(inhouse_engine_str, pool_size = 20, max_overflow = 10, pool_pre_ping=True, pool_recycle=900)
#     with inhouse_engine.begin() as conn:
#         df = pd.read_sql_table('netPositionBSE', con=conn)
#         print('data fetched from src1(inhouse db)')
#         df.createdAt = df.createdAt.astype(str)
#         df.updatedAt = df.updatedAt.astype(str)
#     return df
#
# def get_src3_trade():
#     df_src3 = read_data_db(for_table='BSE_ENetMIS')
#     print('data fetched from src3(RMS)')
#     return df_src3
#
# src1_df = get_src1_trade()
# src2_df = get_src2_trade()
# src3_df = get_src3_trade()
w=0
# from common import read_data_db, write_notis_postgredb, today, yesterday
# # n_tbl_bse_trade_data = f'BSE_TRADE_DATA_{today}'
#
# # bse_df = read_data_db(for_table=n_tbl_bse_trade_data)
# # write_notis_postgredb(table_name=n_tbl_bse_trade_data,truncate_required=True)
# # tradehist_db = read_data_db(for_table='TradeHist')
# from_time = '14-07-2025 09:15:00'
# # to_time = '14-07-2025 15:12:00'
# to_time = datetime.now().replace(second=0, microsecond=0).strftime('%d-%m-%Y %H:%M:%S')
# rms_db = read_data_db(for_table='BSE_ENetMIS', from_time=from_time, to_time=to_time)
# trd_df = read_data_db(for_table='TradeHist', from_time=from_time, to_time=to_time)
o=0
# def calc_bse_eod_net_pos(desk_bse_df,for_unde):
#     if for_unde == 'BSE':
#         underlying_list = ['SENSEX','BANKEX']
#     else:
#         underlying_list = ['NIFTY', 'BANKNIFTY', 'MIDCPNIFTY', 'FINNIFTY']
#     yest_eod_df = read_data_db(for_table=f'NOTIS_EOD_NET_POS_CP_NONCP_{yesterday.strftime("%Y-%m-%d")}')
#     yest_eod_df.EodExpiry = pd.to_datetime(yest_eod_df.EodExpiry, dayfirst=True, format='mixed').dt.date
#
#     yest_eod_df = yest_eod_df.query("EodUnderlying in @underlying_list and EodExpiry >= @today and EodBroker != "
#                                     "'SRSPL' and FinalNetQty != 0")
#     yest_eod_df['EodNetQuantity'] = yest_eod_df['FinalNetQty']
#     yest_eod_df['PreFinalNetQty'] = yest_eod_df['FinalNetQty']
#     exclude_columns = ['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType',
#                        'EodNetQuantity', 'PreFinalNetQty', 'FinalNetQty']
#     yest_eod_df.loc[:, ~yest_eod_df.columns.isin(exclude_columns)] = 0
#     yest_eod_df = yest_eod_df.query('FinalNetQty != 0')
#
#     today_eod_df = read_data_db(for_table=f'NOTIS_EOD_NET_POS_CP_NONCP_{today.strftime("%Y-%m-%d")}')
#     today_eod_df.EodExpiry = pd.to_datetime(today_eod_df.EodExpiry, dayfirst=True, format='mixed').dt.date
#     today_eod_df = today_eod_df.query("EodUnderlying in @underlying_list and EodExpiry >= @today and EodBroker != 'SRSPl'")
#
#     if len(desk_bse_df) == 0 or desk_bse_df.empty:
#         if today_eod_df.empty or len(today_eod_df) == 0:
#             return yest_eod_df
#         return today_eod_df
#     yest_eod_df.columns = [re.sub(rf'Eod|\s|Expired', '', each) for each in yest_eod_df.columns]
#     yest_eod_df.Expiry = pd.to_datetime(yest_eod_df.Expiry, dayfirst=True, format='mixed').dt.date
#     yest_eod_df.drop(
#         columns=['NetQuantity', 'buyQty', 'buyAvgPrice', 'buyValue', 'sellQty', 'sellAvgPrice', 'sellValue',
#                  'PreFinalNetQty', 'Spot_close', 'Rate', 'Assn_value', 'SellValue', 'BuyValue', 'Qty'],
#         inplace=True
#     )
#     yest_eod_df.rename(columns={'FinalNetQty': 'NetQuantity'}, inplace=True)
#     yest_eod_df = yest_eod_df.add_prefix('Eod')
#     yest_eod_df = yest_eod_df.query("EodUnderlying in @underlying_list and EodExpiry >= @today and EodNetQuantity != 0 "
#                           "and EodBroker != 'SRSPL'")
#     grouped_eod_df = yest_eod_df.groupby(by=['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType'],
#                                     as_index=False).agg({'EodNetQuantity': 'sum'})
#     grouped_eod_df = grouped_eod_df.query("EodNetQuantity != 0")
#     grouped_eod_df = grouped_eod_df.drop_duplicates()
#     # ============================================================================================
#     desk_bse_df.Expiry = pd.to_datetime(desk_bse_df.Expiry, dayfirst=True, format='mixed').dt.date
#     grouped_desk_df = desk_bse_df.groupby(by=['Broker', 'Underlying', 'Expiry', 'Strike', 'OptionType'],
#                                           as_index=False).agg({
#         'BuyQty': 'sum', 'SellQty': 'sum', 'buyAvgPrice': 'mean', 'sellAvgPrice': 'mean',
#         'buyValue': 'sum', 'sellValue': 'sum'
#     })
#     grouped_desk_df['IntradayVolume'] = grouped_desk_df['BuyQty'] - grouped_desk_df['SellQty']
#     if len(grouped_eod_df) > len(grouped_desk_df):
#         merged_df = grouped_eod_df.merge(
#             grouped_desk_df,
#             left_on=['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType'],
#             right_on=['Broker', 'Underlying', 'Expiry', 'Strike', 'OptionType'],
#             how='outer'
#         )
#     else:
#         merged_df = grouped_desk_df.merge(
#             grouped_eod_df,
#             right_on=['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType'],
#             left_on=['Broker', 'Underlying', 'Expiry', 'Strike', 'OptionType'],
#             how='outer'
#         )
#     merged_df.fillna(0, inplace=True)
#     merged_df.drop_duplicates(inplace=True)
#     # ============================================================================================
#     coltd1 = ['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType']
#     coltd2 = ['Broker', 'Underlying', 'Expiry', 'Strike', 'OptionType']
#     for i in range(len(coltd1)):
#         merged_df.loc[merged_df[coltd1[i]] == 0, coltd1[i]] = merged_df[coltd2[i]]
#         merged_df.loc[merged_df[coltd2[i]] == 0, coltd2[i]] = merged_df[coltd1[i]]
#     merged_df['FinalNetQty'] = merged_df['EodNetQuantity'] + merged_df['IntradayVolume']
#     merged_df.drop(columns=['Broker', 'Underlying', 'Expiry', 'Strike', 'OptionType'], inplace=True)
#     # ============================================================================================
#     col_to_int = ['BuyQty', 'SellQty']
#     for col in col_to_int:
#         merged_df[col] = merged_df[col].astype(np.int64)
#     merged_df.rename(columns={'BuyQty': 'buyQty', 'SellQty': 'sellQty'}, inplace=True)
#     print(f'length of cp noncp for {today} is {merged_df.shape}')
#     return merged_df
y=0
# sql_server = "rms.ar.db"
# sql_database = "ENetMIS"
# sql_username = "notice_user"
# sql_password = "Notice@2024"
# sql_query = (
#     f"SELECT DISTINCT LEFT(LTRIM(RTRIM([time])), 5) AS hh_mm "
#     f"from [ENETMIS].[dbo].[BSE_FO_AA100_view] "
#     f"order by hh_mm desc"
# )
# try:
#     sql_connection_string = (
#         f"DRIVER={{ODBC Driver 17 for SQL Server}};"
#         f"SERVER={sql_server};"
#         f"DATABASE={sql_database};"
#         f"UID={sql_username};"
#         f"PWD={sql_password}"
#     )
#     with pyodbc.connect(sql_connection_string) as sql_conn:
#         df = pd.read_sql_query(sql_query, sql_conn)
#     print(f"Data fetched from SQL Server. Shape:{df.shape}")
# except (pyodbc.Error, psycopg2.Error) as e:
#     print("Error occurred:", e)
t=0
# # Finding last tuesday(SENSEX)
# import calendar
# from common import holidays_25, read_data_db
# from Untitled_bse_utility import BSEUtility
#
# # today = datetime.today().date()
# # b_days = pd.bdate_range(start=today - timedelta(days=45), end=today + timedelta(days=45),freq='C',weekmask='1111100',
# #                         holidays=holidays_25).date.tolist()
# # def find_tuesday(year,month):
# #     last_day = date(year=year, month=month, day=calendar.monthrange(year=year, month=month)[1])
# #     offset = (last_day.weekday() -1) % 7
# #     last_tues = last_day.replace(day=last_day.day-offset)
# #     b_days = [each for each in b_days if each.month == last_tues.month]
# #     if last_tues in b_days:
# #         return last_tues
# #     else:
# #         b_days = [each for each in b_days if each < last_tues]
# #         return b_days[-1]
# # # ltues = find_tuesday(year=2025,month=7)
#
# def conv_exp(val):
#     if re.fullmatch(r'\d{6}', val):
#         return datetime.strptime(val,'%y%m%d').date()
#     elif re.fullmatch(r'\d{5}', val):
#         fulldate = val[:2]+'0'+val[2:]
#         return datetime.strptime(fulldate, '%y%m%d').date()
#     elif re.fullmatch(r'\d{2}[A-Z]{3}', val):
#         year = 2000+int(val[:2])
#         month = datetime.strptime(val[2:],'%b').month
#         start_date = datetime.today().date()
#         end_date = datetime(year=year,month=month,day=calendar.monthrange(year=year,month=month)[1]).date()
#         b_days = pd.bdate_range(start=start_date,end=end_date,freq='C',weekmask='1111100',
#                                 holidays=holidays_25).date.tolist()
#         # Finding last wednesday, weekday=2
#         offset = (end_date.weekday() - 2) % 7
#         last_wed = end_date.replace(day=end_date.day - offset)
#         if last_wed in b_days:
#             return last_wed
#         else:
#             b_days = [each for each in b_days if each < last_wed]
#             return b_days[-1]
# res = conv_exp('25AUG')
u=0
# # fetch annualized volatility
# from common import (today,yesterday, read_file, volt_dir, find_spot,
#                     read_data_db, holidays_25)
# from db_config import n_tbl_notis_eod_net_pos_cp_noncp
# import mibian
#
# def calc_dte(row):
#     # if row['EodExpiry'] == pd.to_datetime('2025-07-24').date():
#     #     return 1
#     # else:
#     #     bdays_left = pd.bdate_range(start=today, end=row['EodExpiry'],freq='C',weekmask='1111100', holidays=holidays_25)
#     #     actual_bdays_left = len(bdays_left) - 1
#     #     return actual_bdays_left
#     bdays_left = pd.bdate_range(start=today, end=row['EodExpiry'], freq='C', weekmask='1111100', holidays=holidays_25)
#     actual_bdays_left = len(bdays_left)
#     return actual_bdays_left
# def get_delta(row):
#     int_rate,annual_div = 5.5,0
#     # if row['EodExpiry'] == pd.to_datetime('2025-07-31').date() and row['EodOptionType'] == 'CE':
#     #     p=0
#     if row['EodOptionType'] == 'CE':
#         calc = mibian.Me(
#             [row['spot'],row['EodStrike'],int_rate,annual_div,row['dte']],
#             volatility=row['volatility']
#         )
#         return calc.callDelta
#     elif row['EodOptionType'] == 'PE':
#         calc = mibian.Me(
#             [row['spot'], row['EodStrike'],int_rate,annual_div,row['dte']],
#             volatility=row['volatility']
#         )
#         return calc.putDelta
#     else:
#         return 1.0
#
# sym_list = ['NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY', 'SENSEX', 'BANKEX']
# eod_df = read_data_db(for_table=n_tbl_notis_eod_net_pos_cp_noncp)
# col_keep = ['EodBroker','EodUnderlying','EodExpiry','EodStrike','EodOptionType','PreFinalNetQty']
# # eod_df.columns = [col for col in col_keep if col in eod_df.columns]
# eod_df.drop(columns=[col for col in eod_df.columns if col not in col_keep], inplace=True)
# eod_df['EodExpiry'] = pd.to_datetime(eod_df['EodExpiry'], dayfirst=True).dt.date
# volt_df = read_file(os.path.join(volt_dir, f'FOVOLT_{yesterday.strftime("%d%m%Y")}.csv'))
# volt_df.columns = [re.sub(r'\s','',each) for each in volt_df.columns]
# volt_df.rename(columns = {'ApplicableAnnualisedVolatility(N)=Max(ForL)':'AnnualizedReturn'}, inplace=True)
# volt_df = volt_df.iloc[:,[1,-1]].query("Symbol in @sym_list")
# spot_dict = find_spot()
# volt_dict = dict(zip(volt_df['Symbol'],volt_df['AnnualizedReturn']))
# eod_df['spot'] = eod_df['EodUnderlying'].map(spot_dict)
# eod_df['volatility'] = eod_df['EodUnderlying'].map(volt_dict)
# eod_df['volatility'] = eod_df['volatility'].astype(np.float64)
# eod_df['volatility'] = eod_df['volatility'] * 100
# eod_df['dte'] = eod_df.apply(calc_dte, axis=1)
# mask = eod_df['EodOptionType'] == 'XX'
# eod_df.loc[mask,'volatility'] = 1
# eod_df['deltaPerUnit'] = eod_df.apply(get_delta, axis=1).astype(np.float64)
# eod_df['deltaQty'] = eod_df['PreFinalNetQty'] * eod_df['deltaPerUnit']
# eod_df['deltaExposure(in Cr)'] = (eod_df['spot'] * eod_df['deltaQty']) / 10_000_000
# # mask = (eod_df['EodOptionType'] == 'CE' | eod_df['EodOptionType'] == 'PE')
# mask = eod_df['EodOptionType'].isin(['CE','PE'])
# eod_df.loc[mask,'EodOptionType'] = 'CE_PE'
# # for summerisation
# final_eod_df = pd.DataFrame()
# for each in ['XX','CE_PE']:
#     temp_eod_df = eod_df.query("EodOptionType == @each")
#     grouped_temp_eod_df = temp_eod_df.groupby(by=['EodOptionType','EodBroker','EodUnderlying'], as_index=False)['deltaExposure(in Cr)'].agg(
#         {'Long':lambda x:x[x>0].sum(),'Short':lambda x: x[x<0].sum(),'Net':'sum'}
#     )
#     total_dict = {
#         'EodOptionType':'',
#         'EodBroker':'Total',
#         'EodUnderlying':'',
#         'Long':grouped_temp_eod_df['Long'].sum(),
#         'Short':grouped_temp_eod_df['Short'].sum(),
#         'Net':grouped_temp_eod_df['Net'].sum()
#     }
#     grouped_temp_eod_df = pd.concat([grouped_temp_eod_df,pd.DataFrame([total_dict])], ignore_index=True)
#     final_eod_df = pd.concat([final_eod_df,grouped_temp_eod_df], ignore_index=True)
# # combined_dict = {
# #     'EodOptionType':'Combined',
# #     'EodBroker':'Total',
# #     'Long':grouped_temp_eod_df['Long'].sum(),
# #     'Short':grouped_temp_eod_df['Short'].sum(),
# #     'Net':grouped_temp_eod_df['Net'].sum()
# # }
# for each in ['deltaExposure(in Cr)','deltaQty']:
#     grouped_df = eod_df.groupby(by=['EodBroker','EodUnderlying'], as_index=False)[each].agg(
#             {'Long':lambda x: x[x>0].sum(),'Short':lambda x: x[x<0].sum(),'Net':'sum'}
#         )
#     if each == 'deltaExposure(in Cr)':
#         use = 'Combined'
#         grouped_df['EodOptionType'] = 'Combined'
#     else:
#         use = 'DeltaQty'
#         grouped_df['EodOptionType'] = 'DeltaQty'
#     total_dict = {
#         'EodOptionType':use,
#         'EodBroker':'Total',
#         'EodUnderlying':'',
#         'Long':grouped_df['Long'].sum(),
#         'Short':grouped_df['Short'].sum(),
#         'Net':grouped_df['Net'].sum()
#     }
#     grouped_df = pd.concat([grouped_df,pd.DataFrame([total_dict])], ignore_index=False)
#     # grouped_df.fillna('',inplace=True)
#     final_eod_df = pd.concat([final_eod_df,grouped_df], ignore_index=False)
u=0
# import random
# random_number = random.randint(1, 135)
# print("Random number:", random_number)
y=0
# import re, os, progressbar, pyodbc, warnings, psycopg2, time, mibian, math
# import pandas as pd
# import numpy as np
# from datetime import datetime, timedelta, timezone
# from py_vollib.black_scholes.greeks.analytical import delta
# from common import read_file, volt_dir, find_spot, holidays_25, holidays_26, read_data_db, write_notis_postgredb
# from db_config import n_tbl_spot_data
#
# today = pd.to_datetime('2025-08-22').date()
# # today = datetime.today().date()
# yesterday = today - timedelta(days=1)
# # yesterday = today - timedelta(days=1)
# # a=find_spot()
# # def calc_dte(row):
# #     total_holidays = holidays_25 + holidays_26
# #     bdays_left = pd.bdate_range(start=today, end=row['EodExpiry'], freq='C', weekmask='1111100', holidays=total_holidays)
# #     actual_bdays_left = len(bdays_left)
# #     return actual_bdays_left
# def get_delta(row):
#     int_rate = 5.5
#     annual_div = 0.0
#     spot = row['spot']
#     strike = row['EodStrike']
#     dte = row['dte']
#     # dte = (row['EodExpiry'] - today).days
#     vol = float(row['volatility'])
#     if row['EodOptionType'] == 'XX':
#         return 1.0
#     calc = mibian.BS(
#         [spot, strike, int_rate, dte],
#         volatility=vol
#     )
#     return calc.callDelta if row['EodOptionType'] == 'CE' else calc.putDelta
#     # if row['EodOptionType'] == 'CE':
#     #     # calc = mibian.BS(
#     #     #     [row['spot'],row['EodStrike'],int_rate,row['dte']],
#     #     #     volatility=row['volatility']
#     #     # )
#     #     return calc.callDelta
#     # else:
#     #     # calc = mibian.BS(
#     #     #     [row['spot'], row['EodStrike'],int_rate,row['dte']],
#     #     #     volatility=row['volatility']
#     #     # )
#     #     return calc.putDelta
#     # # else:
#     # #     return 1.0
#
# def get_delta_vollib(row):
#     int_rate = 0.055
#     spot = row['spot']
#     strike = row['EodStrike']
#     dte = row['dte']/365
#     vol = row['volatility'] / 100
#     if row['EodOptionType'] == 'XX':
#         return 1.0
#     elif row['EodOptionType'] == 'CE':
#         d = delta('c',spot,strike,dte,int_rate,vol)
#         return d
#     else:
#         d = delta('p',spot,strike,dte,int_rate,vol)
#         return d
#
# # def calc_delta(eod_df):
# #     eod_df = eod_df.copy()
# #     eod_df['EodExpiry'] = pd.to_datetime(eod_df['EodExpiry'], dayfirst=True).dt.date
# #     sym_list = ['NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY', 'SENSEX', 'BANKEX']
# #     col_keep = ['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType', 'PreFinalNetQty']
# #     eod_df.drop(columns=[col for col in eod_df.columns if col not in col_keep], inplace=True)
# #     eod_df = eod_df.query("EodExpiry != @today")
# #     volt_df = read_file(os.path.join(volt_dir, f'FOVOLT_{today.strftime("%d%m%Y")}.csv'))
# #     volt_df.columns = [re.sub(r'\s', '', each) for each in volt_df.columns]
# #     volt_df.rename(columns={'ApplicableAnnualisedVolatility(N)=Max(ForL)': 'AnnualizedReturn'}, inplace=True)
# #     volt_df = volt_df.iloc[:, [1, -1]].query("Symbol in @sym_list")
# #     volt_df = volt_df.applymap(lambda x:re.sub(r'\s+','',x) if isinstance(x,str) else x)
# #     volt_df['AnnualizedReturn'] = volt_df['AnnualizedReturn'].astype(np.float64)
# #     # volt_df = volt_df.reset_index()
# #     # spot_dict = find_spot()
# #     spot_dict = {
# #         'NIFTY':24980.65,
# #         'BANKNIFTY':55865.15,
# #         'SENSEX':81644.39
# #     }
# #     volt_dict = dict(zip(volt_df['Symbol'], volt_df['AnnualizedReturn']))
# #     eod_df['spot'] = eod_df['EodUnderlying'].map(spot_dict)
# #     eod_df['volatility'] = eod_df['EodUnderlying'].map(volt_dict)
# #     eod_df['volatility'] = eod_df['volatility'].astype(np.float64)
# #     eod_df['volatility'] = eod_df['volatility'] * 100
# #     # eod_df['volatility'] = eod_df['volatility'].round()
# #     eod_df['dte'] = eod_df['EodExpiry'].apply(lambda x: (x-today).days)
# #     mask = eod_df['EodOptionType'] == 'XX'
# #     eod_df.loc[mask, 'volatility'] = 1
# #     eod_df['deltaPerUnit'] = eod_df.apply(get_delta, axis=1).astype(np.float64)
# #     # eod_df['deltaPerUnit'] = eod_df.apply(get_delta_vollib, axis=1).astype(np.float64)
# #     eod_df['deltaQty'] = (eod_df['PreFinalNetQty'] * eod_df['deltaPerUnit'])
# #     eod_df['deltaExposure(in Cr)'] = ((eod_df['spot'] * eod_df['deltaQty']) / 10_000_000).round(2)
# #     eod_df.to_excel(os.path.join(test_dir, f'eod_delta_{today}_{datetime.today().strftime("%H%M")}.xlsx'),
# #                     index=False)
# #     mask = eod_df['EodOptionType'].isin(['CE', 'PE'])
# #     eod_df.loc[mask, 'EodOptionType'] = 'CE_PE'
# #     final_eod_df = pd.DataFrame()
# #     for each in ['XX', 'CE_PE']:
# #         temp_eod_df = eod_df.query("EodOptionType == @each")
# #         grouped_temp_eod_df = temp_eod_df.groupby(by=['EodOptionType', 'EodBroker', 'EodUnderlying'], as_index=False)[
# #             'deltaExposure(in Cr)'].agg(
# #             {'Long': lambda x: x[x > 0].sum(), 'Short': lambda x: x[x < 0].sum(), 'Net': 'sum'}
# #         )
# #         total_dict = {
# #             'EodOptionType': '',
# #             'EodBroker': 'Total',
# #             'EodUnderlying': '',
# #             'Long': grouped_temp_eod_df['Long'].sum(),
# #             'Short': grouped_temp_eod_df['Short'].sum(),
# #             'Net': grouped_temp_eod_df['Net'].sum()
# #         }
# #         grouped_temp_eod_df = pd.concat([grouped_temp_eod_df, pd.DataFrame([total_dict])], ignore_index=True)
# #         final_eod_df = pd.concat([final_eod_df, grouped_temp_eod_df], ignore_index=True)
# #     for each in ['deltaExposure(in Cr)', 'deltaQty']:
# #         grouped_df = eod_df.groupby(by=['EodBroker', 'EodUnderlying'], as_index=False)[each].agg(
# #             {'Long': lambda x: x[x > 0].sum(), 'Short': lambda x: x[x < 0].sum(), 'Net': 'sum'}
# #         )
# #         if each == 'deltaExposure(in Cr)':
# #             use = 'Combined'
# #             grouped_df['EodOptionType'] = 'Combined'
# #         else:
# #             use = 'DeltaQty'
# #             grouped_df['EodOptionType'] = 'DeltaQty'
# #             grouped_df['Long'] = grouped_df['Long'] / 100000
# #             grouped_df['Short'] = grouped_df['Short'] / 100000
# #             grouped_df['Net'] = grouped_df['Net'] / 100000
# #         total_dict = {
# #             'EodOptionType': use,
# #             'EodBroker': 'Total',
# #             'EodUnderlying': '',
# #             'Long': grouped_df['Long'].sum(),
# #             'Short': grouped_df['Short'].sum(),
# #             'Net': grouped_df['Net'].sum()
# #         }
# #         grouped_df = pd.concat([grouped_df, pd.DataFrame([total_dict])], ignore_index=False)
# #         final_eod_df = pd.concat([final_eod_df, grouped_df], ignore_index=False)
# #     return final_eod_df
#
# def calc_delta(eod_df):
#     delta_df = eod_df.copy()
#     delta_df['EodExpiry'] = pd.to_datetime(delta_df['EodExpiry'], dayfirst=True).dt.date
#     sym_list = ['NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY', 'SENSEX', 'BANKEX']
#     col_keep = ['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType', 'PreFinalNetQty']
#     delta_df.drop(columns=[col for col in delta_df.columns if col not in col_keep], inplace=True)
#     if datetime.now() > datetime.today().replace(hour=16,minute=50,second=0,microsecond=0):
#         volt_df = read_file(os.path.join(volt_dir, f'FOVOLT_{yesterday.strftime("%d%m%Y")}.csv'))
#         spot_dict = find_spot()
#     else:
#         delta_df = delta_df.query("EodExpiry != @today")
#         volt_df = read_file(os.path.join(volt_dir, f'FOVOLT_{today.strftime("%d%m%Y")}.csv'))
#         spot_df = read_data_db(for_table=n_tbl_spot_data)
#         spot_dict = spot_df.to_dict(orient='records')[0]
#     volt_df.columns = [re.sub(r'\s', '', each) for each in volt_df.columns]
#     volt_df.rename(columns={'ApplicableAnnualisedVolatility(N)=Max(ForL)': 'AnnualizedReturn'}, inplace=True)
#     volt_df = volt_df.iloc[:, [1, -1]].query("Symbol in @sym_list")
#     volt_df = volt_df.applymap(lambda x: re.sub(r'\s+', '', x) if isinstance(x, str) else x)
#     volt_df['AnnualizedReturn'] = volt_df['AnnualizedReturn'].astype(np.float64)
#     # spot_dict = find_spot()
#     volt_dict = dict(zip(volt_df['Symbol'], volt_df['AnnualizedReturn']))
#     delta_df['spot'] = delta_df['EodUnderlying'].map(spot_dict)
#     delta_df['volatility'] = delta_df['EodUnderlying'].map(volt_dict)
#     delta_df['volatility'] = delta_df['volatility'].astype(np.float64)
#     delta_df['volatility'] = delta_df['volatility'] * 100
#     delta_df['dte'] = delta_df['EodExpiry'].apply(lambda x: (x-today).days)
#     mask = delta_df['EodExpiry'] == today
#     delta_df.loc[mask,'dte'] = 1
#     mask = delta_df['EodOptionType'] == 'XX'
#     delta_df.loc[mask, 'volatility'] = 1
#     delta_df['deltaPerUnit'] = delta_df.apply(get_delta, axis=1).astype(np.float64)
#     delta_df['deltaQty'] = (delta_df['PreFinalNetQty'] * delta_df['deltaPerUnit'])
#     delta_df['deltaExposure(in Cr)'] = (delta_df['spot'] * delta_df['deltaQty']) / 10_000_000
#     delta_df1 = delta_df.copy()
#     # delta_df1.to_excel(os.path.join(test_dir, f'eod_delta_{today}_{datetime.today().strftime("%H%M")}.xlsx'),index=False)
#     final_delta_df = pd.DataFrame()
#     mask = delta_df1['EodOptionType'].isin(['CE', 'PE'])
#     delta_df1.loc[mask, 'EodOptionType'] = 'CE_PE'
#     for each in ['XX', 'CE_PE']:
#         temp_delta_df1 = delta_df1.query("EodOptionType == @each")
#         grouped_temp_delta_df1 = temp_delta_df1.groupby(by=['EodOptionType', 'EodBroker', 'EodUnderlying'], as_index=False)[
#             'deltaExposure(in Cr)'].agg(
#             {'Long': lambda x: x[x > 0].sum(), 'Short': lambda x: x[x < 0].sum(), 'Net': 'sum'}
#         )
#         total_dict = {
#             'EodOptionType': each,
#             'EodBroker': 'Total',
#
#             'Long': grouped_temp_delta_df1['Long'].sum(),
#             'Short': grouped_temp_delta_df1['Short'].sum(),
#             'Net': grouped_temp_delta_df1['Net'].sum()
#         }
#         grouped_temp_delta_df1 = pd.concat([grouped_temp_delta_df1, pd.DataFrame([total_dict])], ignore_index=True)
#         final_delta_df = pd.concat([final_delta_df, grouped_temp_delta_df1], ignore_index=True)
#     for each in ['deltaExposure(in Cr)', 'deltaQty']:
#         grouped_df = delta_df1.groupby(by=['EodBroker', 'EodUnderlying'], as_index=False)[each].agg(
#             {'Long': lambda x: x[x > 0].sum(), 'Short': lambda x: x[x < 0].sum(), 'Net': 'sum'}
#         )
#         if each == 'deltaExposure(in Cr)':
#             use = 'Combined'
#             grouped_df['EodOptionType'] = 'Combined'
#         else:
#             use = 'DeltaQty'
#             grouped_df['EodOptionType'] = 'DeltaQty'
#             grouped_df['Long'] = grouped_df['Long'] / 100000
#             grouped_df['Short'] = grouped_df['Short'] / 100000
#             grouped_df['Net'] = grouped_df['Net'] / 100000
#         total_dict = {
#             'EodOptionType': use,
#             'EodBroker': 'Total',
#
#             'Long': grouped_df['Long'].sum(),
#             'Short': grouped_df['Short'].sum(),
#             'Net': grouped_df['Net'].sum()
#         }
#         grouped_df = pd.concat([grouped_df, pd.DataFrame([total_dict])], ignore_index=False)
#         final_delta_df = pd.concat([final_delta_df, grouped_df], ignore_index=False)
#     delta_df2 = delta_df.copy()
#     for each in ['deltaExposure(in Cr)', 'deltaQty']:
#         grouped_df = delta_df2.groupby(by=['EodUnderlying'], as_index=False)[each].agg(
#             {'Long': lambda x:x[x>0].sum(), 'Short': lambda x:x[x<0].sum(), 'Net':'sum'}
#         )
#         if each == 'deltaExposure(in Cr)':
#             use = 'Underlying Combined'
#         else:
#             use = 'Underlying DeltaQty'
#             grouped_df['Long'] = grouped_df['Long'] / 100000
#             grouped_df['Short'] = grouped_df['Short'] / 100000
#             grouped_df['Net'] = grouped_df['Net'] / 100000
#         grouped_df['EodOptionType'] = use
#         total_dict = {
#             'EodOptionType': use,
#             'EodBroker':'Total',
#             'Long': grouped_df['Long'].sum(),
#             'Short': grouped_df['Short'].sum(),
#             'Net': grouped_df['Net'].sum()
#         }
#         grouped_df = pd.concat([grouped_df, pd.DataFrame([total_dict])], ignore_index=True)
#         final_delta_df = pd.concat([final_delta_df,grouped_df], ignore_index=True)
#     return final_delta_df
#
# eod_df = read_data_db(for_table=f"NOTIS_EOD_NET_POS_CP_NONCP_{today}")
# # eod_df = pd.read_excel(rf"D:\notis_analysis\input_data\eod_2025-08-19.xlsx", index_col=False)
# delta_df = calc_delta(eod_df)
# delta_df.to_excel(os.path.join(test_dir, f'final_delta_{today}_{datetime.today().strftime("%H%M")}_1.xlsx'),
#                   index=False)
# write_notis_postgredb(df=delta_df,table_name=f"NOTIS_DELTA_{today}",truncate_required=True)
p=0
# import requests
# def find_spot_v2(for_dt, index_list=[]):
#     if not index_list:
#         spot_dict = {}
#         index_list = ["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "SENSEX"]
#         url = 'http://192.168.112.219:8080/livedataname'
#         headers = {
#             'esegment': '["1"]',
#             'oi': '["1"]'
#         }
#         proxies = {"http": None, "https": None}
#         for each in index_list:
#             headers[f"inst_name"] = f'["{each}"]'
#             try:
#                 response = requests.get(url=url, headers=headers, proxies=proxies)
#                 if response.status_code == 200:
#                     for index, index_value in response.json().items():
#                         # spot_list.append({index:f'{index_value[2]}'})
#                         spot_dict[index] = index_value[2]
#             except Exception as e:
#                 print(f"Error in fetching spot data = {e}")
#     return spot_dict

# find_spot_v2(for_dt=datetime.today().date())
o=0
# from common import read_data_db, today, write_notis_postgredb
# from nse_utility import NSEUtility
# from bse_utility import BSEUtility
# def get_nse_data():
#     # # logger.info(f'fetching NSE trades...')
#     # df_db = read_data_db()
#     # if df_db is None or df_db.empty:
#     #     # logger.info(f'No NSE trade done today hence skipping')
#     #     df = pd.DataFrame()
#     #     return df
#     # # logger.info(f'Notis trade data fetched, shape={df_db.shape}')
#     # # write_notis_postgredb(df=df_db, table_name=n_tbl_notis_raw_data, raw=True, truncate_required=True)
#     # # modify_filepath = os.path.join(modified_dir, f'NOTIS_TRADE_DATA_{today.strftime("%d%b%Y").upper()}.csv')
#     # nnf_file_path = os.path.join(root_dir, "Final_NNF_ID.xlsx")
#     # if not os.path.exists(nnf_file_path):
#     #     raise FileNotFoundError("NNF File not found. Please add the NNF file and try again.")
#     # readable_mod_time = datetime.fromtimestamp(os.path.getmtime(nnf_file_path))
#     # if readable_mod_time.date() == today:  # Check if the NNF file is modified today or not
#     #     # logger.info(f'New NNF Data found, modifying the nnf data in db . . .')
#     #     df_nnf = pd.read_excel(nnf_file_path, index_col=False)
#     #     df_nnf = df_nnf.loc[:, ~df_nnf.columns.str.startswith('Un')]
#     #     df_nnf.columns = df_nnf.columns.str.replace(' ', '', regex=True)
#     #     df_nnf.dropna(how='all', inplace=True)
#     #     df_nnf = df_nnf.drop_duplicates()
#     #     # write_notis_postgredb(df=df_nnf, table_name=n_tbl_notis_nnf_data, truncate_required=True)
#     # else:
#     #     df_nnf = read_data_db(nnf=True, for_table=n_tbl_notis_nnf_data)
#     #     df_nnf = df_nnf.drop_duplicates()
#     modified_df = read_data_db(for_table='NOTIS_TRADE_BOOK_2025-08-22')
#     # write_notis_postgredb(modified_df, table_name=n_tbl_notis_trade_book, truncate_required=True)
#     # write_notis_data(modified_df, modify_filepath)
#     # write_notis_data(modified_df, rf'C:\Users\vipulanand\Documents\Anand Rathi Financial Services Ltd (Synced)\OneDrive - Anand Rathi Financial Services Ltd\notis_files\NOTIS_TRADE_DATA_{today.strftime("%d%b%Y").upper()}.csv')
#     # logger.info('file saved in modified_data folder')
#     modified_df['trdQtyPrc'] = modified_df['trdQty'] * (modified_df['trdPrc'] / 100)
#     pivot_df = modified_df.pivot_table(
#         index=['MainGroup', 'SubGroup', 'broker', 'ctclid', 'sym', 'expDt', 'strPrc', 'optType'],
#         columns=['bsFlg'],
#         values=['trdQty', 'trdQtyPrc', 'trdPrc'],
#         aggfunc={'trdQty': 'sum', 'trdQtyPrc': 'sum', 'trdPrc': ['min', 'max']},
#         fill_value=0
#     )
#     if len(modified_df.bsFlg.unique()) == 1:
#         if modified_df.bsFlg.unique().tolist()[0] == 'B':
#             pivot_df['SellTrdQtyPrc'] = pivot_df['SellQty'] = pivot_df['SellMax'] = pivot_df['SellMin'] = 0
#             pivot_df.columns = ['BuyMax', 'SellMax', 'BuyMin', 'SellMin', 'BuyQty', 'SellQty', 'BuyTrdQtyPrc',
#                                 'SellTrdQtyPrc']
#         elif modified_df.bsFlg.unique().tolist()[0] == 'S':
#             pivot_df['BuyTrdQtyPrc'] = pivot_df['BuyQty'] = pivot_df['BuyMax'] = pivot_df['BuyMin'] = 0
#             pivot_df.columns = ['BuyMax', 'SellMax', 'BuyMin', 'SellMin', 'BuyQty', 'SellQty', 'BuyTrdQtyPrc',
#                                 'SellTrdQtyPrc']
#     elif len(modified_df) == 0 or len(pivot_df) == 0:
#         pivot_df.columns = ['_'.join(col).strip() for col in pivot_df.columns.values]
#     else:
#         pivot_df.columns = ['BuyMax', 'SellMax', 'BuyMin', 'SellMin', 'BuyQty', 'SellQty', 'BuyTrdQtyPrc',
#                             'SellTrdQtyPrc']
#     pivot_df.reset_index(inplace=True)
#     pivot_df['BuyAvgPrc'] = pivot_df.apply(
#         lambda row: row['BuyTrdQtyPrc'] / row['BuyQty'] if row['BuyQty'] > 0 else 0.0,
#         axis=1)
#     pivot_df['SellAvgPrc'] = pivot_df.apply(
#         lambda row: row['SellTrdQtyPrc'] / row['SellQty'] if row['SellQty'] > 0 else 0.0, axis=1)
#     # pivot_df.drop(columns=['BuyTrdQtyPrc', 'SellTrdQtyPrc'], inplace=True)
#     pivot_df.rename(columns={'MainGroup': 'mainGroup', 'SubGroup': 'subGroup', 'sym': 'symbol', 'expDt': 'expiryDate',
#                              'strPrc': 'strikePrice', 'optType': 'optionType', 'BuyAvgPrc': 'buyAvgPrice',
#                              'BuyTrdQtyPrc': 'buyValue', 'BuyQty': 'buyAvgQty', 'SellQty': 'sellAvgQty',
#                              'SellAvgPrc': 'sellAvgPrice', 'SellTrdQtyPrc': 'sellValue'},
#                     inplace=True)
#     pivot_df.volume = pivot_df.buyAvgQty - pivot_df.sellAvgQty
#     return pivot_df
# def get_bse_data():
#     # # stt = datetime.now()
#     # # # logger.info(f'fetching BSE trades...')
#     # # df_bse1 = read_data_db(for_table='BSE_ENetMIS')
#     # # if df_bse1 is None or df_bse1.empty:
#     # #     # logger.info(f'No BSE trade done today hence skipping')
#     # #     df = pd.DataFrame()
#     # #     return df
#     # # # logger.info(f'BSE trade data fetched, shape={df_bse1.shape}')
#     # modified_bse_df1 = read_data_db(for_table='BSE_TRADE_DATA_2025-08-08')
#     # modified_bse_df1.TraderID = modified_bse_df1.TraderID.astype(np.int64)
#     #
#     # df_bse2 = read_data_db(for_table='TradeHist')
#     # modified_bse_df2 = BSEUtility.bse_modify_file(df_bse2)
#     # modified_bse_df2 = modified_bse_df2[
#     #     ['TerminalID', 'TradingSymbol', 'FillSize', 'TransactionType', 'ExchUser', 'Underlying', 'Strike', 'OptionType',
#     #      'Expiry']]
#     # modified_bse_df2.ExchUser = modified_bse_df2.ExchUser.astype(np.int64)
#     # modified_bse_df2.ExchUser = modified_bse_df2.ExchUser % 10000
#     # grouped_modified_bse_df2 = (
#     #     modified_bse_df2
#     #     .groupby(['ExchUser', 'TradingSymbol', 'FillSize', 'TransactionType',
#     #               'Underlying', 'Strike', 'OptionType', 'Expiry'], as_index=False)
#     #     .agg({'TerminalID': 'first'})
#     # )
#     #
#     # modified_bse_df = pd.merge(modified_bse_df1, grouped_modified_bse_df2,
#     #                            left_on=['TraderID', 'TradingSymbol', 'FillSize', 'TransactionType', 'Underlying',
#     #                                     'Strike', 'OptionType', 'Expiry'],
#     #                            right_on=['ExchUser', 'TradingSymbol', 'FillSize', 'TransactionType', 'Underlying',
#     #                                      'Strike', 'OptionType', 'Expiry'],
#     #                            how='left'
#     #                            )
#     # modified_bse_df['TerminalID'] = np.where(modified_bse_df['TraderID'] == 1011, '945440A',
#     #                                          modified_bse_df['TerminalID'])
#     # modified_bse_df.drop(columns=['ExchUser'], axis=1, inplace=True)
#     # modified_bse_df.fillna(0, inplace=True)
#     # # write_notis_postgredb(df=modified_bse_df, table_name=n_tbl_bse_trade_data, truncate_required=True)
#     # # write_notis_data(modified_bse_df, os.path.join(bse_dir, f'BSE_TRADE_DATA_{today.strftime("%d%b%Y").upper()}.xlsx'))
#     # # write_notis_data(modified_bse_df,rf'C:\Users\vipulanand\Documents\Anand Rathi Financial Services Ltd (Synced)\OneDrive - Anand Rathi Financial Services Ltd\notis_files\BSE_TRADE_DATA_{today.strftime("%d%b%Y").upper()}.xlsx')
#     modified_bse_df = read_data_db(for_table='BSE_TRADE_DATA_2025-08-22')
#     modified_bse_df['trdQtyPrc'] = modified_bse_df['FillSize'] * (modified_bse_df['AvgPrice'] / 100)
#     pivot_df = modified_bse_df.pivot_table(
#         index=['Broker', 'Underlying', 'Expiry', 'Strike', 'OptionType', 'TerminalID', 'TraderID'],
#         columns=['TransactionType'],
#         values=['FillSize', 'trdQtyPrc', 'AvgPrice'],
#         aggfunc={'FillSize': 'sum', 'trdQtyPrc': 'sum', 'AvgPrice': ['min', 'max']},
#         fill_value=0
#     )
#     if len(modified_bse_df.TransactionType.unique()) == 1:
#         if modified_bse_df.TransactionType.unique().tolist()[0] == 'B':
#             pivot_df['SellTrdQtyPrc'] = pivot_df['SellQty'] = pivot_df['sellMax'] = pivot_df['sellMin'] = 0
#             pivot_df.columns = ['buyMax', 'sellMax', 'buyMin', 'sellMin', 'BuyQty', 'SellQty', 'BuyTrdQtyPrc',
#                                 'SellTrdQtyPrc']
#         elif modified_bse_df.TransactionType.unique().tolist()[0] == 'S':
#             pivot_df['BuyTrdQtyPrc'] = pivot_df['BuyQty'] = pivot_df['buyMax'] = pivot_df['buyMin'] = 0
#             pivot_df.columns = ['buyMax', 'sellMax', 'buyMin', 'sellMin', 'BuyQty', 'SellQty', 'BuyTrdQtyPrc',
#                                 'SellTrdQtyPrc']
#     elif len(modified_bse_df) == 0 or len(pivot_df) == 0:
#         pivot_df.columns = ['_'.join(col).strip() for col in pivot_df.columns.values]
#     else:
#         pivot_df.columns = ['buyMax', 'sellMax', 'buyMin', 'sellMin', 'BuyQty', 'SellQty', 'BuyTrdQtyPrc',
#                             'SellTrdQtyPrc']
#     pivot_df.reset_index(inplace=True)
#     pivot_df['buyAvgPrice'] = pivot_df.apply(
#         lambda row: row['BuyTrdQtyPrc'] / row['BuyQty'] if row['BuyQty'] > 0 else 0, axis=1)
#     pivot_df['sellAvgPrice'] = pivot_df.apply(
#         lambda row: row['SellTrdQtyPrc'] / row['SellQty'] if row['SellQty'] > 0 else 0, axis=1)
#     # pivot_df.drop(columns=['BuyTrdQtyPrc', 'SellTrdQtyPrc'], inplace=True)
#     pivot_df['IntradayVolume'] = pivot_df['BuyQty'] - pivot_df['SellQty']
#     pivot_df.rename(columns={'BuyTrdQtyPrc': 'buyValue', 'SellTrdQtyPrc': 'sellValue'}, inplace=True)
#     # pivot_df = pivot_df.round(2)
#     ett = datetime.now()
#     # logger.info(f'BSE trade fetched. Total time taken: {(ett - stt).seconds} seconds')
#     return pivot_df
#
# nse_pivot_df = get_nse_data()
# bse_pivot_df = get_bse_data()
#
# nse_deal_df = NSEUtility.calc_nse_deal_sheet(nse_pivot_df)
# bse_deal_df = BSEUtility.calc_bse_deal_sheet(bse_pivot_df)
# final_deal_df = pd.concat([nse_deal_df, bse_deal_df], ignore_index=True)
# grouped_main_deal_df = final_deal_df.groupby(
#     by=['Broker', 'Underlying', 'Expiry', 'Strike', 'OptionType'],
#     as_index=False
# ).agg(
#     {'BuyMax': 'max', 'SellMax': 'max',
#      'BuyMin': 'min', 'SellMin': 'min',
#      'BuyQty': 'sum', 'SellQty': 'sum',
#      'BuyValue':'sum', 'SellValue':'sum'}
# )
# # write_notis_postgredb(df=grouped_main_deal_df, table_name='NOTIS_DEAL_SHEET_2025-08-22', truncate_required=True)
y=0
# bse_raw_df = pd.DataFrame(columns=['scid'])
# bse_raw_df['scid'] = ['SENSEX25AUGFUT','SENSEX2581980800CE']
# pattern = r'^([A-Z]+)(\d{5}|\d{6}|\d{2}[A-Z]{3})(\d{5})?([A-Z]{2}|[A-Z]{3})$'
# bse_raw_df[['Underlying','temp_expiry','Strike','OptionType']] = bse_raw_df['scid'].str.extract(pattern)
p=0
# from common import read_data_db, find_spot, write_notis_postgredb
# from db_config import n_tbl_spot_data, n_tbl_notis_eod_net_pos_cp_noncp
# df = read_data_db(for_table=n_tbl_spot_data)
# spot_dict = find_spot()
# cspot_df = pd.DataFrame([spot_dict])
# write_notis_postgredb(df=spot_df,table_name=n_tbl_spot_data,truncate_required=True)
i=0
from common import write_notis_postgredb
mod_eod_df = pd.read_excel(rf"D:\notis_analysis\table_data\NOTIS_EOD_NET_POS_CP_NONCP_2025-10-06_mod.xlsx",
                           index_col=False)
# mod_eod_df.EodExpiry = pd.to_datetime(mod_eod_df)
write_notis_postgredb(df=mod_eod_df,table_name='NOTIS_EOD_NET_POS_CP_NONCP_2025-10-06',truncate_required=True)
i=0