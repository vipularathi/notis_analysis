import io, re, csv, os, json, warnings, xlsxwriter, progressbar, gzip, uvicorn, time
import numpy as np
import pandas as pd
from datetime import datetime, timedelta, timezone, date
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import create_engine
from sqlalchemy.sql import text
from sqlalchemy.orm import sessionmaker, Session
from fastapi import FastAPI, Query, status, Response, Depends, UploadFile, File, HTTPException
from fastapi.encoders import jsonable_encoder
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse, PlainTextResponse
from fastapi.middleware.cors import CORSMiddleware

from db_config import (n_tbl_notis_trade_book, n_tbl_notis_raw_data,
                       n_tbl_notis_nnf_data,n_tbl_notis_delta_table,
                        n_tbl_srspl_trade_data, n_tbl_notis_eod_net_pos_cp_noncp,
                       engine_str, notis_engine_str, bse_engine_str)
from common import (get_date_from_non_jiffy,get_date_from_jiffy,
                    read_data_db, read_notis_file, read_file,
                    write_notis_data, write_notis_postgredb,
                    analyze_expired_instruments_v2, calc_delta_v2,
                    today, yesterday, bhav_dir,
                    logger, volt_dir, zipped_dir)

pd.set_option('display.max_columns', None)
warnings.filterwarnings('ignore')

engine = create_engine(engine_str, pool_pre_ping=True, pool_recycle=300)
sessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

notis_engine = create_engine(notis_engine_str, pool_pre_ping=True, pool_recycle=300)
sessionLocalNotis = sessionmaker(autocommit=False, autoflush=False, bind=notis_engine)

bse_engine = create_engine(notis_engine_str, pool_pre_ping=True, pool_recycle=300)
sessionLocalBSE = sessionmaker(autocommit=False, autoflush=False, bind=bse_engine)

def conv_str(obj):
    if isinstance(obj, datetime):
        # return obj.strftime('%Y-%m-%d %H:%M:%S')
        return obj.isoformat()
    return obj
def get_db(for_table:str=Query()):
    if for_table == f'sourcenotisraw':
        db = sessionLocalNotis()
    elif for_table == f'sourcebseraw':
        db = sessionLocalBSE()
    else:
        db = sessionLocal()
    try:
        yield db
    finally:
        db.close()
def get_notis_db():
    db = sessionLocalNotis()
    try:
        yield db
    finally:
        db.close()

def get_bse_db():
    db = sessionLocalBSE()
    try:
        yield db
    finally:
        db.close()

class ServiceApp:
    def __init__(self):
        self.app = FastAPI(title='NOTIS_Net_Position', description='Notis_net_position', docs_url='/docs', openapi_url='/openapi.json')
        self.app.add_middleware(CORSMiddleware, allow_origins = ['*'], allow_credentials = True, allow_methods=['*'], allow_headers=['*'])
        self.add_routes()
        self.main_mod_df = pd.DataFrame()

    def add_routes(self):
        self.app.add_api_route('/netPosition/intraday', methods=['GET'], endpoint=self.get_intraday_net_position)
        self.app.add_api_route('/netPosition/eod', methods=['GET'], endpoint=self.get_intraday_net_position)
        self.app.add_api_route('/data', methods=['GET'], endpoint=self.get_data)
        self.app.add_api_route('/netPosition/raw', methods=['GET'], endpoint=self.get_raw_net_position)
        self.app.add_api_route('/download', methods=['GET'], endpoint=self.download_data)
        self.app.add_api_route('/exposure', methods=['GET'], endpoint=self.get_exposure)
        self.app.add_api_route('/sourceData', methods=['GET'], endpoint=self.get_source_data)
        self.app.add_api_route('/downloadSourceData', methods=['GET'], endpoint=self.download_source_data)
        self.app.add_api_route('/get_oi', methods=['GET'], endpoint=self.get_oi)
        self.app.add_api_route('/upload', methods=['POST'], endpoint=self.upload_data)
        self.app.add_api_route('/nifty/future/oi', methods=['GET'], endpoint=self.calc_nifty_future_oi)
        self.app.add_api_route('/data/nnfTable', methods=['GET'], endpoint=self.get_nnf_table)
        
    def get_data(self, for_date:date=Query(), for_table:str=Query(), page:int=Query(1), page_size:int=Query(1000),db:Session=Depends(get_db)):
        for_dt = pd.to_datetime(for_date).date()
        if for_table == 'modifiedtradebook':
            tablename = f'NOTIS_TRADE_BOOK_{for_dt}'
        elif for_table == 'nnfwise':
            tablename = f'NOTIS_NNF_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'useridwise': #to_remove
            tablename = f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'deskwise':
            tablename = f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'rawtradebook':
            tablename = f'notis_raw_data_{for_dt}'
        elif for_table == 'eodnetposcp':
            tablename = f'NOTIS_EOD_NET_POS_CP_NONCP_{for_dt.strftime("%Y-%m-%d")}'
        elif for_table == f'bsetradebook':
            tablename = f'BSE_TRADE_DATA_{for_dt.strftime("%Y-%m-%d")}'
        elif for_table == f'delta':
            tablename = f'NOTIS_DELTA_{for_dt.strftime("%Y-%m-%d")}'
        elif for_table == f'deal':
            tablename = f'NOTIS_DEAL_SHEET_{for_dt.strftime("%Y-%m-%d")}'
        query=text(rf'Select * from "{tablename}" limit {page_size} offset {(page -1)*page_size}')
        result = db.execute(query).fetchall()
        total_rows = db.execute(text(rf'Select count(*) from "{tablename}"')).scalar()
        # total_rows = 18
        json_data = {
            'data':[{k: conv_str(v) for k, v in row._mapping.items()} for row in result],
            'total_rows':total_rows,
            'page':page,
            'page_size':page_size
        }
        if not len(result):
            json_data = pd.DataFrame(columns=['data','total_rows','page','page_size']).to_json(orient='records')
            return Response(content=json_data, media_type='application/json')
        else:
            compressed_data = gzip.compress(json.dumps(json_data).encode('utf-8'))
            logger.info(f'\ntotal_rows={json_data["total_rows"]}\tpage={json_data["page"]}\tpage_size={json_data["page_size"]}\n')
            return Response(content=compressed_data, media_type='application/gzip')

    def download_data(self,for_date:date=Query(),for_table:str=Query(), db:Session=Depends(get_db)):
        for_dt = pd.to_datetime(for_date).date()
        if for_table == 'modifiedtradebook':
            tablename = f'NOTIS_TRADE_BOOK_{today}' if for_dt == today else f'NOTIS_TRADE_BOOK_{for_dt}'
        elif for_table == 'nnfwise':
            tablename = f'NOTIS_NNF_WISE_NET_POSITION_{today}' if for_dt == today else f'NOTIS_NNF_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'useridwise':
            tablename = f'NOTIS_NNF_WISE_NET_POSITION_{today}' if for_dt == today else f'NOTIS_NNF_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'deskwise':
            tablename = f'NOTIS_DESK_WISE_NET_POSITION_{today}' if for_dt == today else f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'rawtradebook':
            tablename = f'notis_raw_data_{today}' if for_dt == today else f'notis_raw_data_{for_dt}'
        # netPosition eodNetPosition rawtradebooknetposi
        elif for_table == 'modifiedtradebooknetposi':
            tablename = f'NOTIS_DESK_WISE_NET_POSITION_{today}' if for_dt == today else f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'eodNetPosition':
            tablename = f'NOTIS_DESK_WISE_NET_POSITION_{today}' if for_dt == today else f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
        elif for_table == f'rawtradebooknetposi':
            tablename = f'notis_raw_data_{today}' if for_dt == today else f'notis_raw_data_{for_dt}'
        elif for_table == f'eodnetposcp':
            tablename = f'NOTIS_EOD_NET_POS_CP_NONCP_{today}' if for_dt == today else f'NOTIS_EOD_NET_POS_CP_NONCP_{for_dt.strftime("%Y-%m-%d")}'
        elif for_table == f'bsetradebook':
            tablename = f'BSE_TRADE_DATA_{today.strftime("%Y-%m-%d")}' if for_dt == today else f'BSE_TRADE_DATA_{for_dt.strftime("%Y-%m-%d")}'
        elif for_table == f'delta':
            tablename = f'NOTIS_DELTA_{for_dt.strftime("%Y-%m-%d")}'
        elif for_table == f'deal':
            tablename = f'NOTIS_DEAL_SHEET_{for_dt.strftime("%Y-%m-%d")}'
        # logger.info(f'tablename is {tablename}')
        if for_dt == today:
            zip_path = os.path.join(zipped_dir, f'zipped_{tablename}_{for_dt}.xlsx.gz')
        else:
            zip_path = os.path.join(zipped_dir, f'zipped_{tablename}.xlsx.gz')

        if for_table in ['modifiedtradebooknetposi','eodNetPosition','rawtradebooknetposi']:
            if for_table == 'modifiedtradebooknetposi' or for_table=='eodNetPosition':
                desk_db_df = read_data_db(for_table=tablename)
                desk_db_df.expiryDate = desk_db_df.expiryDate.astype('datetime64[ns]')
                desk_db_df.expiryDate = desk_db_df.expiryDate.dt.date
                desk_db_df.loc[desk_db_df['optionType'] == 'XX', 'strikePrice'] = 0
                desk_db_df.strikePrice = desk_db_df.strikePrice.apply(lambda x: x / 100 if x > 0 else x)
                desk_db_df.strikePrice = desk_db_df.strikePrice.astype('int64')
                grouped_desk_db_df = desk_db_df.groupby(by=['symbol', 'expiryDate', 'strikePrice', 'optionType']).agg(
                    {'buyAvgQty': 'sum', 'buyAvgPrice': 'mean', 'sellAvgQty': 'sum',
                     'sellAvgPrice': 'mean'}).reset_index()
                grouped_desk_db_df['IntradayVolume'] = grouped_desk_db_df['buyAvgQty'] - grouped_desk_db_df[
                    'sellAvgQty']
                grouped_desk_db_df.rename(columns={'buyAvgQty': 'buyQty', 'sellAvgQty': 'sellQty'})
                grouped_desk_db_df.expiryDate = grouped_desk_db_df.expiryDate.astype(str)
                if grouped_desk_db_df.empty:
                    return JSONResponse(content={"message": "No data available"}, status_code=204)
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    grouped_desk_db_df.to_excel(writer, index=False)
                buffer.seek(0)
                with gzip.open(zip_path, 'wb') as f:
                    f.write(buffer.getvalue())
                return FileResponse(zip_path,media_type='application/gzip')
            else:
                df = read_data_db(for_table=tablename)
                if not len(df):
                    return JSONResponse(content={"message": "No data available"}, status_code=204)
                list_str_int64 = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 16, 17, 18, 19, 21, 23, 27, 28]
                list_str_none = [15, 20, 25, 30, 31, 32, 33, 34, 35, 36, 37]
                list_none_str = [38]
                for i in list_str_int64:
                    column_name = f'Column{i}'
                    df[f'Column{i}'] = df[f'Column{i}'].astype('int64')
                for i in list_str_none:
                    df[f'Column{i}'] = None
                for i in list_none_str:
                    df[f'Column{i}'] = df[f'Column{i}'].astype('str')
                logger.info('Starting file modification...')
                df.rename(columns={
                    'Column1': 'seqNo', 'Column2': 'mkt', 'Column3': 'trdNo',
                    'Column4': 'trdTm', 'Column5': 'Tkn', 'Column6': 'trdQty',
                    'Column7': 'trdPrc', 'Column8': 'bsFlg', 'Column9': 'ordNo',
                    'Column10': 'brnCd', 'Column11': 'usrId', 'Column12': 'proCli',
                    'Column13': 'cliActNo', 'Column14': 'cpCD', 'Column15': 'remarks',
                    'Column16': 'actTyp', 'Column17': 'TCd', 'Column18': 'ordTm',
                    'Column19': 'Booktype', 'Column20': 'oppTmCd', 'Column21': 'ctclid',
                    'Column22': 'status', 'Column23': 'TmCd', 'Column24': 'sym',
                    'Column25': 'ser', 'Column26': 'inst', 'Column27': 'expDt',
                    'Column28': 'strPrc', 'Column29': 'optType', 'Column30': 'sessionID',
                    'Column31': 'echoback', 'Column32': 'Fill1', 'Column33': 'Fill2',
                    'Column34': 'Fill3', 'Column35': 'Fill4', 'Column36': 'Fill5', 'Column37': 'Fill6'
                }, inplace=True)
                df['ordTm'] = df['ordTm'].apply(lambda x: get_date_from_non_jiffy(int(x)))
                df['expDt'] = df['expDt'].apply(lambda x: get_date_from_non_jiffy(int(x)))
                df.expDt = df.expDt.astype('datetime64[ns]').dt.date
                df['trdTm'] = df['trdTm'].apply(lambda x: get_date_from_jiffy(int(x)))
                df['bsFlg'] = np.where(df['bsFlg'] == 1, 'B', 'S')
                pivot_df = df.pivot_table(
                    index=['ctclid', 'sym', 'expDt', 'strPrc', 'optType'],
                    columns='bsFlg',
                    values=['trdQty', 'trdPrc'],
                    aggfunc={'trdQty': 'sum', 'trdPrc': 'mean'},
                    fill_value=0
                )
                pivot_df.columns = ['BuyPrc', 'SellPrc', 'BuyVol', 'SellVol']
                pivot_df = pivot_df.reset_index()
                pivot_df = pivot_df[
                    ['ctclid', 'sym', 'expDt', 'strPrc', 'optType', 'BuyVol', 'BuyPrc', 'SellVol', 'SellPrc']]
                pivot_df['expDt'] = pivot_df['expDt'].astype(str)
                pivot_df['IntradayVolume'] = pivot_df['BuyVol'] - pivot_df['SellVol']

                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    pivot_df.to_excel(writer, index=False)
                buffer.seek(0)
                with gzip.open(zip_path, 'wb') as f:
                    f.write(buffer.getvalue())
                return FileResponse(zip_path, media_type='application/gzip')
        elif for_table == 'eodnetposcp':
            grouped_desk_db_df = read_data_db(for_table=tablename)
            # grouped_desk_db_df.drop(columns=['buyAvgPrice','sellAvgPrice'], inplace=True)
            if grouped_desk_db_df.empty:
                return JSONResponse(content={"message": "No data available"}, status_code=204)
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                grouped_desk_db_df.to_excel(writer, index=False)
            buffer.seek(0)
            with gzip.open(zip_path, 'wb') as f:
                f.write(buffer.getvalue())
            return FileResponse(zip_path, media_type='application/gzip')
        else:
            stt = datetime.now()
            total_rows = db.execute(text(rf'select count(*) from "{tablename}"')).scalar()
            page_size = 5_00_000
            num_pages = total_rows // page_size + (1 if (total_rows % page_size) else 0)
            logger.info(f'Total rows in DB: {total_rows}, Splitting into {num_pages} sheets')
            buffer = io.BytesIO()
            wb = xlsxwriter.Workbook(buffer, {'in_memory': True})
            for page in range(num_pages):
                query = f'select * from "{tablename}" limit {page_size} offset {(page) * page_size}'
                logger.info(query)
                pbar = progressbar.ProgressBar(
                    max_value=total_rows + 1,
                    widgets=[
                        progressbar.Percentage(), '',
                        progressbar.Bar(marker='=', left='[', right=']'),
                        progressbar.ETA()
                    ]
                )
                result = db.execute(text(query))
                ws = wb.add_worksheet(f'Sheet{page + 1}')
                for col, header in enumerate(result.keys()):
                    ws.write(0, col, header)
                for rn, row in enumerate(result, start=1):
                    for col, cell in enumerate(row):
                        ws.write(rn, col, cell)
                    pbar.update(rn)
                pbar.finish()
            wb.close()
            logger.info('fetching data from buffer')
            buffer.seek(0)
            logger.info('Writing to xlsx file and zipping . . ')
            with gzip.open(zip_path, 'wb') as f:
                f.write(buffer.getvalue())
            ett = datetime.now()
            logger.info(f'total time taken for zip_path:{(ett - stt).total_seconds()}')
            return FileResponse(path=zip_path, media_type='application/gzip')

    def get_intraday_net_position(self, for_date:date=Query()):
        for_dt = pd.to_datetime(for_date).date()
        if for_dt == today:
            tablename = f'NOTIS_DESK_WISE_NET_POSITION_{today}'
        else:
            tablename = f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
        desk_db_df = read_data_db(for_table=tablename)
        desk_db_df.expiryDate = desk_db_df.expiryDate.astype('datetime64[ns]')
        desk_db_df.expiryDate = desk_db_df.expiryDate.dt.date
        desk_db_df.loc[desk_db_df['optionType'] == 'XX', 'strikePrice'] = 0
        desk_db_df.strikePrice = desk_db_df.strikePrice.apply(lambda x: x / 100 if x > 0 else x)
        desk_db_df.strikePrice = desk_db_df.strikePrice.astype('int64')
        grouped_desk_db_df = desk_db_df.groupby(by=['symbol', 'expiryDate', 'strikePrice', 'optionType']).agg(
            {'buyAvgQty': 'sum', 'buyAvgPrice': 'mean', 'sellAvgQty': 'sum', 'sellAvgPrice': 'mean'}).reset_index()
        grouped_desk_db_df['IntradayVolume'] = grouped_desk_db_df['buyAvgQty'] - grouped_desk_db_df['sellAvgQty']
        grouped_desk_db_df.rename(columns={'buyAvgQty': 'buyQty', 'sellAvgQty': 'sellQty'})
        grouped_desk_db_df.expiryDate = grouped_desk_db_df.expiryDate.astype(str)
        json_data = grouped_desk_db_df.to_json(orient='records')
        if not len(grouped_desk_db_df):
            return Response(content=json_data, media_type='application/json')
        else:
            compressed_data = gzip.compress(json_data.encode('utf-8'))
            return Response(content=compressed_data, media_type='application/gzip')

    def get_raw_net_position(self, for_date:date=Query()):
        for_dt = pd.to_datetime(for_date).date()
        if for_dt == today:
            tablename = f'notis_raw_data_{today}'
        else:
            tablename = f'notis_raw_data_{for_dt}'
        df = read_data_db(for_table=tablename)
        if not len(df):
            json_data = df.to_json(orient='records')
            return Response(content=json_data, media_type='application/json')
        else:
            list_str_int64 = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 16, 17, 18, 19, 21, 23, 27, 28]
            list_str_none = [15, 20, 25, 30, 31, 32, 33, 34, 35, 36, 37]
            list_none_str = [38]
            for i in list_str_int64:
                column_name = f'Column{i}'
                df[f'Column{i}'] = df[f'Column{i}'].astype('int64')
            for i in list_str_none:
                df[f'Column{i}'] = None
            for i in list_none_str:
                df[f'Column{i}'] = df[f'Column{i}'].astype('str')
            logger.info('Starting file modification...')
            df.rename(columns={
                'Column1': 'seqNo', 'Column2': 'mkt', 'Column3': 'trdNo',
                'Column4': 'trdTm', 'Column5': 'Tkn', 'Column6': 'trdQty',
                'Column7': 'trdPrc', 'Column8': 'bsFlg', 'Column9': 'ordNo',
                'Column10': 'brnCd', 'Column11': 'usrId', 'Column12': 'proCli',
                'Column13': 'cliActNo', 'Column14': 'cpCD', 'Column15': 'remarks',
                'Column16': 'actTyp', 'Column17': 'TCd', 'Column18': 'ordTm',
                'Column19': 'Booktype', 'Column20': 'oppTmCd', 'Column21': 'ctclid',
                'Column22': 'status', 'Column23': 'TmCd', 'Column24': 'sym',
                'Column25': 'ser', 'Column26': 'inst', 'Column27': 'expDt',
                'Column28': 'strPrc', 'Column29': 'optType', 'Column30': 'sessionID',
                'Column31': 'echoback', 'Column32': 'Fill1', 'Column33': 'Fill2',
                'Column34': 'Fill3', 'Column35': 'Fill4', 'Column36': 'Fill5', 'Column37': 'Fill6'
            }, inplace=True)
            df['ordTm'] = df['ordTm'].apply(lambda x: get_date_from_non_jiffy(int(x)))
            df['expDt'] = df['expDt'].apply(lambda x: get_date_from_non_jiffy(int(x)))
            df.expDt = df.expDt.astype('datetime64[ns]').dt.date
            df['trdTm'] = df['trdTm'].apply(lambda x: get_date_from_jiffy(int(x)))
            df['bsFlg'] = np.where(df['bsFlg'] == 1, 'B', 'S')
            pivot_df = df.pivot_table(
                index=['ctclid', 'sym', 'expDt', 'strPrc', 'optType'],
                columns='bsFlg',
                values=['trdQty', 'trdPrc'],
                aggfunc={'trdQty': 'sum', 'trdPrc': 'mean'},
                fill_value=0
            )
            pivot_df.columns = ['BuyPrc', 'SellPrc', 'BuyVol', 'SellVol']
            pivot_df = pivot_df.reset_index()
            pivot_df = pivot_df[['ctclid', 'sym', 'expDt', 'strPrc', 'optType', 'BuyVol', 'BuyPrc', 'SellVol', 'SellPrc']]
            pivot_df['expDt'] = pivot_df['expDt'].astype(str)
            pivot_df['IntradayVolume'] = pivot_df['BuyVol'] - pivot_df['SellVol']

            json_data = pivot_df.to_json(orient='records')
            compressed_data = gzip.compress(json_data.encode('utf-8'))
            return Response(content=compressed_data, media_type='application/gzip')

    def get_source_data(self, for_date='', for_table:str=Query(), page:int=Query(1), page_size:int=Query(1000), db:Session=Depends(get_db)):
        offset = (page - 1) * page_size
        if for_table == f'sourcenotisraw':
            tablename = "[ENetMIS].[dbo].[NSE_FO_AA100_view]"
            query = text(f"""
                WITH CTE AS (
                    SELECT *,
                           ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS RowNum
                    FROM [ENetMIS].[dbo].[NSE_FO_AA100_view]
                )
                SELECT *
                FROM CTE
                WHERE RowNum > {offset} AND RowNum <= {offset + page_size};
                """)
            result = db.execute(query).fetchall()
            total_rows = db.execute(text(rf'Select count(*) from [ENetMIS].[dbo].[NSE_FO_AA100_view]')).scalar()
        elif for_table == f'sourcebseraw':
            query = text(f"""
                WITH CTE AS (
                    SELECT *,
                           ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS RowNum
                    FROM [ENetMIS].[dbo].[BSE_FO_AA100_view]
                    where scid like 'SENSEX%' or scid like 'BANKEX%'
                )
                SELECT *
                FROM CTE
                WHERE RowNum > {offset} AND RowNum <= {offset + page_size};
            """)
            result = db.execute(query).fetchall()
            total_rows = db.execute(text(
                rf"""Select count(*) from [ENetMIS].[dbo].[BSE_FO_AA100_view]
                            where scid like 'SENSEX%' or scid like 'BANKEX%'""")).scalar()
        json_data = {
            'data': [{k: conv_str(v) for k, v in row._mapping.items()} for row in result],
            'total_rows': total_rows,
            'page': page,
            'page_size': page_size
        }
        if not len(result):
            json_data = pd.DataFrame(columns=['data', 'total_rows', 'page', 'page_size']).to_json(orient='records')
            return Response(content=json_data, media_type='application/json')
        else:
            compressed_data = gzip.compress(json.dumps(json_data).encode('utf-8'))
            logger.info(
                f'\ntotal_rows={json_data["total_rows"]}\tpage={json_data["page"]}\tpage_size={json_data["page_size"]}\n')
            return Response(content=compressed_data, media_type='application/gzip')

    def download_source_data(self, for_date = '', for_table:str=Query(), db:Session=Depends(get_db)):
        stt = datetime.now()
        zip_path = os.path.join(zipped_dir, f'zipped_{for_table}.xlsx.gz')
        if for_table == f'sourcenotisraw':
            total_rows = db.execute(text(rf'Select count(*) from [ENetMIS].[dbo].[NSE_FO_AA100_view]')).scalar()
        elif for_table == f'sourcebseraw':
            total_rows = db.execute(text(
                rf"""Select count(*) from [ENetMIS].[dbo].[BSE_FO_AA100_view]
                where scid like 'SENSEX%' or scid like 'BANKEX%'""")).scalar()
        page_size = 5_00_000
        num_pages = total_rows // page_size + (1 if (total_rows % page_size) else 0)
        logger.info(f'Total rows in DB: {total_rows}, Splitting into {num_pages} sheets')
        buffer = io.BytesIO()
        wb = xlsxwriter.Workbook(buffer, {'in_memory': True})
        for page in range(num_pages):
            query2,query='',''
            offset = (page - 1) * page_size
            if for_table == f'sourcenotisraw':
                query = (f"""
                    WITH CTE AS (
                        SELECT *,
                               ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS RowNum
                        FROM [ENetMIS].[dbo].[NSE_FO_AA100_view]
                    )
                    SELECT *
                    FROM CTE
                    WHERE RowNum > {page * page_size} AND RowNum <= {(page + 1) * page_size};
                """)
            elif for_table == f'sourcebseraw':
                query = (f"""
                    WITH CTE AS (
                        SELECT *,
                               ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS RowNum
                        FROM [ENetMIS].[dbo].[BSE_FO_AA100_view]
                    )
                    SELECT * FROM CTE
                    WHERE RowNum > {page * page_size} AND RowNum <= {(page + 1) * page_size}
                """)
            logger.info(query)
            pbar = progressbar.ProgressBar(
                max_value=total_rows + 1,
                widgets=[
                    progressbar.Percentage(), '',
                    progressbar.Bar(marker='=', left='[', right=']'),
                    progressbar.ETA()
                ]
            )
            result = db.execute(text(query))
            ws = wb.add_worksheet(f'Sheet{page + 1}')
            for col, header in enumerate(result.keys()):
                ws.write(0, col, header)
            for rn, row in enumerate(result, start=1):
                for col, cell in enumerate(row):
                    ws.write(rn, col, cell)
                pbar.update(rn)
            pbar.finish()
        wb.close()
        logger.info('fetching data from buffer')
        buffer.seek(0)
        logger.info('Writing to xlsx file and zipping . . ')
        with gzip.open(zip_path, 'wb') as f:
            f.write(buffer.getvalue())
        ett = datetime.now()
        logger.info(f'total time taken for zip_path:{(ett - stt).total_seconds()}')
        return FileResponse(path=zip_path, media_type='application/gzip')

    def get_exposure(self, for_date:date=Query()):
        for_date = datetime.today().date().strftime('%Y-%m-%d')
        volt_df = read_file(os.path.join(volt_dir, f'FOVOLT_{yesterday.strftime("%d%m%Y")}.csv'))
        volt_df.columns = [re.sub(r'\s', '', each) for each in volt_df.columns]
        volt_df = volt_df.iloc[:, 1:3]
        volt_df.rename(columns={'UnderlyingClosePrice(A)': 'SpotClosePrice'}, inplace=True)
        sym_list = ['NIFTY','BANKNIFTY','FINNIFTY','MIDCPNIFTY','SENSEX']
        volt_df = volt_df.query("Symbol in @sym_list")

        # tablename = f'NOTIS_EOD_NET_POS_CP_NONCP_{for_date}'
        cp_df = read_data_db(for_table=n_tbl_notis_eod_net_pos_cp_noncp)
        cp_df.columns = [re.sub(r'Eod|\s', '', each) for each in cp_df.columns]

        merged_df = cp_df.merge(volt_df, how='left', left_on=['Underlying'], right_on=['Symbol'])
        merged_df.drop(columns=['Symbol'], inplace=True)
        merged_df = merged_df.query("OptionType == 'CE' or OptionType == 'PE'")
        merged_df.drop_duplicates(inplace=True)

        pivot_df = merged_df.pivot_table(
            index=['Broker', 'Underlying', 'SpotClosePrice'],
            columns=['OptionType'],
            values=['FinalNetQty'],
            aggfunc={'FinalNetQty': 'sum'},
            fill_value=0
        )
        pivot_df.columns = ['CE', 'PE']
        pivot_df.reset_index(inplace=True)
        pivot_df.SpotClosePrice = pivot_df.SpotClosePrice.astype('float64')
        pivot_df['NetQty'] = pivot_df['CE'] - pivot_df['PE']
        pivot_df['Exposure(in Crs)'] = (pivot_df['NetQty'] * pivot_df['SpotClosePrice']) / 10000000
        print(f'exposure table shape: {pivot_df.shape}')
        json_data = pivot_df.to_json(orient='records')
        if not pivot_df.empty:
            return Response(content=json_data, media_type='application/json')
    
    def get_oi(self, for_date:date=Query(), summary:bool=Query()):
        # for_date = datetime.today().date().strftime('%Y-%m-%d')
        # table_to_read = f'NOTIS_EOD_NET_POS_CP_NONCP_{for_date}'
        for_dt = pd.to_datetime(for_date).date()
        eod_df = read_data_db(for_table=f"NOTIS_EOD_NET_POS_CP_NONCP_{for_dt}")
        eod_df.columns = [re.sub(r'Eod|\s', '', each) for each in eod_df.columns]
        if not summary:
            grouped_df = eod_df.groupby(by=['Broker', 'Underlying', 'Expiry'], as_index=False).agg(
                {'PreFinalNetQty': lambda x: x.abs().sum()})
        else:
            grouped_df = eod_df.groupby(by=['Underlying'], as_index=False).agg(
                {'PreFinalNetQty':lambda x:x.abs().sum()}
            )
        json_data = grouped_df.to_json(orient='records')
        return Response(json_data, media_type='application/json')
    
    def upload_data(self, for_date=Query(), file: UploadFile = File(...), for_table:str=Query(),
                    use_carryover:bool=Query()):
        # for_date = datetime.today().date().strftime('%Y-%m-%d')
        # table_to_read = f'NOTIS_EOD_NET_POS_CP_NONCP_{for_date}'
        filename = file.filename.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
            return JSONResponse(status_code=415, content="Unsupported file format")
        data = file.file.read()
        buffer = io.BytesIO(data)
        try:
            if filename.endswith('.csv'):
                df = pd.read_csv(buffer)
            else:
                df = pd.read_excel(buffer)
            if for_table == 'SRSPL'.lower() or for_table == 'SRSPL':
                df['EodExpiry'] = pd.to_datetime(df['EodExpiry'], dayfirst=True).dt.date
                df = df[
                    ['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType', 'EodNetQuantity', 'buyQty',
                     'buyValue', 'sellQty', 'sellValue', 'PreFinalNetQty']
                ]
                write_notis_postgredb(df=df, table_name=n_tbl_srspl_trade_data, truncate_required=True)
                if use_carryover:
                    df = read_data_db(for_table=f'NOTIS_EOD_NET_POS_CP_NONCP_{yesterday}')
                    df = df.query("EodBroker in ['CP','non CP']")
                    write_notis_postgredb(df=df, table_name=f'NOTIS_EOD_NET_POS_CP_NONCP_{yesterday}',
                                          truncate_required=True)
                if datetime.today().time() > datetime.strptime('15:35:00','%H:%M:%S').time():
                    eod_df = read_data_db(for_table=n_tbl_notis_eod_net_pos_cp_noncp)
                    orig_eod_df = eod_df.query("EodBroker in ['CP','non CP']")
                    eod_df = eod_df.query("EodBroker not in ['CP','non CP']")
                    concat_df = pd.concat([eod_df, df], ignore_index=True)
                    concat_df['EodExpiry'] = pd.to_datetime(concat_df['EodExpiry'], dayfirst=True).dt.date
                    concat_df.fillna(0, inplace=True)
                    grouped_eod_df = concat_df.groupby(
                        by=['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType'],
                        as_index=False).agg(
                        {'EodNetQuantity': 'last', 'buyQty': 'sum', 'buyValue': 'sum',
                         'sellQty': 'sum', 'sellValue': 'sum', 'PreFinalNetQty': 'sum'}
                    )
                    grouped_eod_df.fillna(0, inplace=True)
                    grouped_eod_df['PreFinalNetQty'] = (
                      grouped_eod_df['EodNetQuantity'] + grouped_eod_df['buyQty'] -
                      grouped_eod_df['sellQty']
                    )
                    grouped_eod_df['ExpiredSpot_close'] = 0.0
                    grouped_eod_df['ExpiredRate'] = 0.0
                    grouped_eod_df['ExpiredAssn_value'] = 0.0
                    grouped_eod_df['ExpiredSellValue'] = 0.0
                    grouped_eod_df['ExpiredBuyValue'] = 0.0
                    grouped_eod_df['ExpiredQty'] = 0.0
                    if today in grouped_eod_df.EodExpiry.unique():
                        grouped_eod_df = analyze_expired_instruments_v2(for_date=today,grouped_final_eod=grouped_eod_df)
                    final_eod_df = pd.concat([orig_eod_df,grouped_eod_df], ignore_index=True)
                    final_eod_df['FinalNetQty'] = final_eod_df['PreFinalNetQty'] + final_eod_df['ExpiredQty']
                    final_eod_df['EodExpiry'] = pd.to_datetime(final_eod_df['EodExpiry'], dayfirst=True).dt.date
                    write_notis_postgredb(df=final_eod_df, table_name=n_tbl_notis_eod_net_pos_cp_noncp, truncate_required=True)
                    delta_df = calc_delta_v2(for_date=today,eod_df=final_eod_df)
                    write_notis_postgredb(df=delta_df, table_name=n_tbl_notis_delta_table, truncate_required=True)
            elif for_table == 'nnf' or for_table == 'nnf'.upper():
                df.columns = df.columns.str.replace(' ', '', regex=True)
                col_list = ['NNFID', 'TerminalID', 'TerminalName', 'UserID', 'SubGroup', 'MainGroup', 'NeatID']
                df = df[col_list]
                # df = df.loc[:, ~df.columns.str.startswith('Un')]
                df.dropna(how='all', inplace=True)
                df = df.drop_duplicates()
                write_notis_postgredb(df=df, table_name=n_tbl_notis_nnf_data, truncate_required=True)
        except Exception as e:
            raise HTTPException(
                status_code=400,
                detail='File format not supported.'
            )
        return JSONResponse("File Uploaded successfully")
    
    def calc_nifty_future_oi(self, for_date=Query()):
        table_name = n_tbl_notis_eod_net_pos_cp_noncp
        eod_df = read_data_db(for_table=n_tbl_notis_eod_net_pos_cp_noncp)
        eod_df.EodExpiry = pd.to_datetime(eod_df.EodExpiry, dayfirst=True).dt.date
        eod_df = eod_df.query("EodUnderlying == 'NIFTY' and EodOptionType == 'XX'")
        grouped_eod = eod_df.groupby(by=['EodBroker', 'EodUnderlying', 'EodExpiry', 'EodStrike', 'EodOptionType'], as_index=False).agg({'EodNetQuantity':'sum', 'buyQty':'sum','sellQty':'sum','PreFinalNetQty':'sum'})
        grouped_eod.fillna(0, inplace=True)
        grouped_eod['OI'] = grouped_eod['PreFinalNetQty'].abs()
        for each in grouped_eod['EodBroker'].unique():
            each_last_exp = sorted(grouped_eod.query("EodBroker == @each")['EodExpiry'])[-1]
            mask = (grouped_eod['EodBroker'] == each) & (grouped_eod['EodExpiry'] == each_last_exp)
            total_oi = grouped_eod.query("EodBroker == @each")['OI'].abs().sum()
            grouped_eod.loc[mask, 'OI Total'] = total_oi
        grouped_eod.fillna(0, inplace=True)
        bhav_pattern = rf'regularNSEBhavcopy_{yesterday.strftime("%d%m%Y")}.(xlsx|csv)' #regularNSEBhavcopy_19062025
        bhav_matched_file = [f for f in os.listdir(bhav_dir) if re.match(bhav_pattern, f)]
        bhav_df = read_file(os.path.join(bhav_dir,bhav_matched_file[0]))
        bhav_df.columns = bhav_df.columns.str.replace(' ', '')
        bhav_df = bhav_df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        bhav_df.columns = bhav_df.columns.str.capitalize()
        bhav_df = bhav_df.add_prefix('Bhav')
        bhav_df['BhavOpeninterest'] = bhav_df['BhavOpeninterest'].astype(np.int64)
        bhav_df['BhavExpiry'] = bhav_df['BhavExpiry'].apply(lambda x: pd.to_datetime(get_date_from_non_jiffy(x)).date())
        for broker in grouped_eod.EodBroker.unique():
            sum_oi = 0
            mask = None
            for index in grouped_eod.query("EodBroker == @broker")['EodUnderlying'].unique():
                fut_last_exp = sorted(grouped_eod.query("EodBroker == @broker and EodUnderlying == @index")[
                                       'EodExpiry'])[-1]
                mask = (grouped_eod['EodBroker'] == broker) & (grouped_eod['EodUnderlying'] == index) & (
                  grouped_eod['EodExpiry'] == fut_last_exp)
                for each_exp in sorted(grouped_eod.query("EodBroker == @broker and EodUnderlying == @index")[
                                        'EodExpiry'].unique()):
                    prev_oi = bhav_df.query(
                        "BhavSymbol == @index and BhavExpiry == @each_exp and BhavInstrumentname == 'FUTIDX'")[
                        'BhavOpeninterest'].unique()[0]
                    sum_oi += prev_oi
            grouped_eod.loc[mask, 'Fut OI(T-1)'] = sum_oi
        grouped_eod.replace('nan', 0, inplace=True)
        grouped_eod.fillna(0, inplace=True)
        grouped_eod['Fut OI(T-1)'] = pd.to_numeric(grouped_eod['Fut OI(T-1)'], errors='coerce')
        grouped_eod['%MktShare'] = grouped_eod.apply(
            lambda row: row['OI Total'] / row['Fut OI(T-1)'] if row['Fut OI(T-1)'] != 0 else 0, axis=1
        )
        grouped_eod['EodExpiry'] = grouped_eod['EodExpiry'].astype(str)
        json_data = grouped_eod.to_json(orient='records')
        return JSONResponse(json_data, media_type='application/json')
    
    def get_nnf_table(self):
        nnf_df = read_data_db(for_table=n_tbl_notis_nnf_data)
        nnf_df.fillna(0, inplace=True)
        json_data = nnf_df.to_json(orient='records')
        return Response(content=json_data, media_type='application/json')

service = ServiceApp()
app = service.app

if __name__ == '__main__':
    uvicorn.run('notis_app_per_minute:app', host='172.16.47.81', port=8871, workers=6)