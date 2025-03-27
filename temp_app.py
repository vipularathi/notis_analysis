import io, re
import csv
import numpy as np
import pandas as pd
import os

import time
import json
from datetime import datetime, timedelta, timezone, date
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter
import progressbar
from sqlalchemy import create_engine
from sqlalchemy.sql import text
from sqlalchemy.orm import sessionmaker, Session
import warnings
from fastapi import FastAPI, Query, status, Response, Depends
from fastapi.encoders import jsonable_encoder
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
import gzip
import uvicorn
from db_config import n_tbl_notis_trade_book, s_tbl_notis_trade_book, n_tbl_notis_raw_data, s_tbl_notis_raw_data, n_tbl_notis_nnf_data, s_tbl_notis_nnf_data, engine_str
from main import modify_file
from common import get_date_from_non_jiffy,get_date_from_jiffy, read_data_db, read_notis_file, write_notis_data, write_notis_postgredb, read_file, today, yesterday, logger, volt_dir

# today = datetime.now().date()
# today = datetime(year=2025, month=3, day=24).date()
# yesterday = datetime(year=2025, month=1, day=23).date()
pd.set_option('display.max_columns', None)
warnings.filterwarnings('ignore')

root_dir = os.path.dirname(os.path.abspath(__file__))
nnf_file_path = os.path.join(root_dir, "Final_NNF_ID.xlsx")
eod_test_dir = os.path.join(root_dir, 'eod_testing')
eod_input_dir = os.path.join(root_dir, 'eod_original')
eod_output_dir = os.path.join(root_dir, 'eod_data')
table_dir = os.path.join(root_dir, 'table_data')
bhav_path = os.path.join(root_dir, 'bhavcopy')
test_dir = os.path.join(root_dir, 'testing')
eod_net_pos_input_dir = os.path.join(root_dir, 'test_net_position_original')
eod_net_pos_output_dir = os.path.join(root_dir, 'test_net_position_code')
zipped_dir = os.path.join(root_dir, 'zipped_files')

engine = create_engine(engine_str)
sessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


def conv_str(obj):
    if isinstance(obj, datetime):
        # return obj.strftime('%Y-%m-%d %H:%M:%S')
        return obj.isoformat()
    return obj
def get_db():
    db = sessionLocal()
    try:
        yield db
    finally:
        db.close()

# def stream_data(db: Session, tablename: str):
#     def conv_str(obj):
#         if isinstance(obj,datetime):
#             # return obj.strftime('%Y-%m-%d %H:%M:%S')
#             return obj.isoformat()
#     yield '['
#     first = True
#     query = text(rf'Select * from "{tablename}"')
#     with db.connection().execution_options(stream_results=True) as conn:
#         result = conn.execute(query)
#         for row in result:
#             if not first:
#                 yield ','
#             first = False
#             # logger.info(json.dumps(dict(row), default=conv_str))
#             yield json.dumps(dict(row), default=conv_str)
#     yield ']'

class ServiceApp:
    def __init__(self):
        self.app = FastAPI(title='NOTIS_Net_Position', description='Notis_net_position', docs_url='/docs', openapi_url='/openapi.json')
        self.app.add_middleware(CORSMiddleware, allow_origins = ['*'], allow_credentials = True, allow_methods=['*'], allow_headers=['*'])
        self.add_routes()
        self.main_mod_df = pd.DataFrame()

    def add_routes(self):
        self.app.add_api_route('/netPosition/intraday', methods=['GET'], endpoint=self.get_intraday_net_position)
        self.app.add_api_route('/netPosition/eod', methods=['GET'], endpoint=self.get_intraday_net_position)
        self.app.add_api_route('/data', methods=['GET'], endpoint=self.get_data)
        # self.app.add_api_route('/data/deskwise', methods=['GET'], endpoint=self.get_deskwise_netposition)
        # self.app.add_api_route('/data/useridwise', methods=['GET'], endpoint=self.get_useridwise_netposition)
        # self.app.add_api_route('/data/nnfwise', methods=['GET'], endpoint=self.get_nnfwise_netposition)
        # self.app.add_api_route('/data/rawTradeData', methods=['GET'], endpoint=self.get_raw_trade_data)
        self.app.add_api_route('/netPosition/raw', methods=['GET'], endpoint=self.get_raw_net_position)
        # self.app.add_api_route('/data/streamModifiedTradeData', methods=['GET'], endpoint=self.test_get_notis_trade_data)
        # self.app.add_api_route('/data/pageModifiedTradeData', methods=['GET'], endpoint=self.test_page_get_notis_trade_data)
        self.app.add_api_route('/download', methods=['GET'], endpoint=self.download_data)
        self.app.add_api_route('/exposure', methods=['GET'], endpoint=self.get_exposure)

    # def get_notis_trade_data(self, for_date:date=Query()):
    #     for_dt = pd.to_datetime(for_date).date()
    #     if for_dt == today:
    #         tablename = f'NOTIS_TRADE_BOOK'
    #     else:
    #         tablename = f'NOTIS_TRADE_BOOK_{for_dt}'
    #     desk_db_df = read_data_db(for_table=tablename)
    #     json_data = desk_db_df.to_json(orient='records')
    #     if not len(desk_db_df):
    #         return Response(content=json_data, media_type='application/json')
    #     else:
    #         compressed_data = gzip.compress(json_data.encode('utf-8'))
    #         return Response(content=compressed_data, media_type='application/gzip')
    #
    # def test_get_notis_trade_data(self, for_date:date=Query(), db:Session=Depends(get_db)):
    #     for_dt = pd.to_datetime(for_date).date()
    #     tablename = f'NOTIS_TRADE_BOOK' if for_dt==today else f'NOTIS_TRADE_BOOK_{for_dt}'
    #     return StreamingResponse(stream_data(db, tablename), media_type='application/json')

    def get_data(self, for_date:date=Query(), for_table:str=Query(), page:int=Query(1), page_size:int=Query(1000),db:Session=Depends(get_db)):
        for_dt = pd.to_datetime(for_date).date()
        if for_table == 'modifiedtradebook':
            tablename = f"test_mod_{today}" if for_dt == today else f'NOTIS_TRADE_BOOK_{for_dt}'
        elif for_table == 'nnfwise':
            tablename = f'test_net_pos_nnf_{today}' if for_dt == today else f'NOTIS_NNF_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'useridwise':
            tablename = f'test_net_pos_desk_{today}' if for_dt == today else f'NOTIS_USERID_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'deskwise':
            tablename = f'test_net_pos_desk_{today}' if for_dt == today else f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'rawtradebook':
            tablename = f'test_raw_{today}' if for_dt == today else f'notis_raw_data_{for_dt}'
        elif for_table == 'eodnetposcp':
            tablename = f'test_cp_noncp_{today}' if for_dt == today else f'NOTIS_EOD_NET_POS_CP_NONCP_{for_dt.strftime("%Y-%m-%d")}'
        elif for_table == f'bsetradebook':
            tablename = f'test_bse_{today.strftime("%Y-%m-%d")}' if for_dt == today else f'BSE_TRADE_DATA_{for_dt.strftime("%Y-%m-%d")}'
        # logger.info(f'tablename={tablename}')
        # tablename = f'NOTIS_TRADE_BOOK' if for_dt==today else f'NOTIS_TRADE_BOOK_{for_dt}'
        query=text(rf'Select * from "{tablename}" limit {page_size} offset {(page -1)*page_size}')
        result = db.execute(query).fetchall()
        total_rows = db.execute(text(rf'Select count(*) from "{tablename}"')).scalar()
        # json_data = {
        #     'data':[{k: conv_str(v) for k, v in dict(row).items()} for row in result],
        #     'total_rows':total_rows,
        #     'page':page,
        #     'page_size':page_size
        # }
        json_data = {
            'data':[{k: conv_str(v) for k, v in row._mapping.items()} for row in result],
            'total_rows':total_rows,
            'page':page,
            'page_size':page_size
        }
        if not len(result):
            json_data = pd.DataFrame(columns=['data','total_rows','page','page_size']).to_json(orient='records')
            return Response(content=json_data, media_type='application/json')
        else:
            compressed_data = gzip.compress(json.dumps(json_data).encode('utf-8'))
            logger.info(f'\ntotal_rows={json_data["total_rows"]}\tpage={json_data["page"]}\tpage_size={json_data["page_size"]}\n')
            return Response(content=compressed_data, media_type='application/gzip')
            # return Response(content=json.dumps(json_data), media_type='application/json')

    def download_data(self,for_date:date=Query(),for_table:str=Query(), db:Session=Depends(get_db)):
        for_dt = pd.to_datetime(for_date).date()
        if for_table == 'modifiedtradebook':
            tablename = f'test_mod_{today}' if for_dt == today else f'NOTIS_TRADE_BOOK_{for_dt}'
        elif for_table == 'nnfwise':
            tablename = f'test_net_pos_nnf_{today}' if for_dt == today else f'NOTIS_NNF_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'useridwise':
            tablename = f'test_net_pos_nnf_{today}' if for_dt == today else f'NOTIS_USERID_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'deskwise':
            tablename = f'test_net_pos_desk_{today}' if for_dt == today else f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'rawtradebook':
            tablename = f'test_raw_{today}' if for_dt == today else f'notis_raw_data_{for_dt}'
        # netPosition eodNetPosition rawtradebooknetposi
        elif for_table == 'modifiedtradebooknetposi':
            tablename = f'test_net_pos_desk_{today}' if for_dt == today else f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
        elif for_table == 'eodNetPosition':
            tablename = f'test_net_pos_desk_{today}' if for_dt == today else f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
        elif for_table == f'rawtradebooknetposi':
            tablename = f'test_raw_{today}' if for_dt == today else f'notis_raw_data_{for_dt}'
        elif for_table == f'eodnetposcp':
            tablename = f'test_cp_noncp_{today}' if for_dt == today else f'NOTIS_EOD_NET_POS_CP_NONCP_{for_dt.strftime("%Y-%m-%d")}'
        elif for_table == f'bsetradebook':
            tablename = f'test_bse_{today.strftime("%Y-%m-%d")}' if for_dt == today else f'BSE_TRADE_DATA_{for_dt.strftime("%Y-%m-%d")}'
        # logger.info(f'tablename is {tablename}')
        if for_dt == today:
            zip_path = os.path.join(zipped_dir, f'zipped_{tablename}_{for_dt}.xlsx.gz')
        else:
            zip_path = os.path.join(zipped_dir, f'zipped_{tablename}.xlsx.gz')

        if for_table in ['modifiedtradebooknetposi','eodNetPosition','rawtradebooknetposi']:
            if for_table == 'modifiedtradebooknetposi' or for_table=='eodNetPosition':
                desk_db_df = read_data_db(for_table=tablename)
                desk_db_df.expiryDate = desk_db_df.expiryDate.astype('datetime64[ns]')
                desk_db_df.expiryDate = desk_db_df.expiryDate.dt.date
                desk_db_df.loc[desk_db_df['optionType'] == 'XX', 'strikePrice'] = 0
                desk_db_df.strikePrice = desk_db_df.strikePrice.apply(lambda x: x / 100 if x > 0 else x)
                desk_db_df.strikePrice = desk_db_df.strikePrice.astype('int64')
                grouped_desk_db_df = desk_db_df.groupby(by=['symbol', 'expiryDate', 'strikePrice', 'optionType']).agg(
                    {'buyAvgQty': 'sum', 'buyAvgPrice': 'mean', 'sellAvgQty': 'sum',
                     'sellAvgPrice': 'mean'}).reset_index()
                grouped_desk_db_df['IntradayVolume'] = grouped_desk_db_df['buyAvgQty'] - grouped_desk_db_df[
                    'sellAvgQty']
                grouped_desk_db_df.rename(columns={'buyAvgQty': 'buyQty', 'sellAvgQty': 'sellQty'})
                grouped_desk_db_df.expiryDate = grouped_desk_db_df.expiryDate.astype(str)
                if grouped_desk_db_df.empty:
                    return JSONResponse(content={"message": "No data available"}, status_code=204)
                # # json_data = grouped_desk_db_df.to_json(orient='records')
                # # if not len(grouped_desk_db_df):
                # #     return Response(content=json_data, media_type='application/json')
                # # else:
                # #     compressed_data = gzip.compress(json_data.encode('utf-8'))
                # #     return Response(content=compressed_data, media_type='application/gzip')
                # buffer = io.StringIO()
                # writer = csv.writer(buffer)
                # writer.writerow(grouped_desk_db_df.columns)
                # writer.writerow(grouped_desk_db_df.values)
                # with gzip.open(zip_path, 'wt', encoding='utf-8', newline='') as f:
                #     f.write(buffer.getvalue())
                # return FileResponse(zip_path, media_type='application/gzip')
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    grouped_desk_db_df.to_excel(writer, index=False)
                buffer.seek(0)
                with gzip.open(zip_path, 'wb') as f:
                    f.write(buffer.getvalue())
                return FileResponse(zip_path,media_type='application/gzip')
            else:
                df = read_data_db(for_table=tablename)
                if not len(df):
                    return JSONResponse(content={"message": "No data available"}, status_code=204)
                list_str_int64 = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 16, 17, 18, 19, 21, 23, 27, 28]
                list_str_none = [15, 20, 25, 30, 31, 32, 33, 34, 35, 36, 37]
                list_none_str = [38]
                for i in list_str_int64:
                    column_name = f'Column{i}'
                    df[f'Column{i}'] = df[f'Column{i}'].astype('int64')
                for i in list_str_none:
                    df[f'Column{i}'] = None
                for i in list_none_str:
                    df[f'Column{i}'] = df[f'Column{i}'].astype('str')
                logger.info('Starting file modification...')
                df.rename(columns={
                    'Column1': 'seqNo', 'Column2': 'mkt', 'Column3': 'trdNo',
                    'Column4': 'trdTm', 'Column5': 'Tkn', 'Column6': 'trdQty',
                    'Column7': 'trdPrc', 'Column8': 'bsFlg', 'Column9': 'ordNo',
                    'Column10': 'brnCd', 'Column11': 'usrId', 'Column12': 'proCli',
                    'Column13': 'cliActNo', 'Column14': 'cpCD', 'Column15': 'remarks',
                    'Column16': 'actTyp', 'Column17': 'TCd', 'Column18': 'ordTm',
                    'Column19': 'Booktype', 'Column20': 'oppTmCd', 'Column21': 'ctclid',
                    'Column22': 'status', 'Column23': 'TmCd', 'Column24': 'sym',
                    'Column25': 'ser', 'Column26': 'inst', 'Column27': 'expDt',
                    'Column28': 'strPrc', 'Column29': 'optType', 'Column30': 'sessionID',
                    'Column31': 'echoback', 'Column32': 'Fill1', 'Column33': 'Fill2',
                    'Column34': 'Fill3', 'Column35': 'Fill4', 'Column36': 'Fill5', 'Column37': 'Fill6'
                }, inplace=True)
                df['ordTm'] = df['ordTm'].apply(lambda x: get_date_from_non_jiffy(int(x)))
                df['expDt'] = df['expDt'].apply(lambda x: get_date_from_non_jiffy(int(x)))
                df.expDt = df.expDt.astype('datetime64[ns]').dt.date
                df['trdTm'] = df['trdTm'].apply(lambda x: get_date_from_jiffy(int(x)))
                df['bsFlg'] = np.where(df['bsFlg'] == 1, 'B', 'S')
                pivot_df = df.pivot_table(
                    index=['ctclid', 'sym', 'expDt', 'strPrc', 'optType'],
                    columns='bsFlg',
                    values=['trdQty', 'trdPrc'],
                    aggfunc={'trdQty': 'sum', 'trdPrc': 'mean'},
                    fill_value=0
                )
                pivot_df.columns = ['BuyPrc', 'SellPrc', 'BuyVol', 'SellVol']
                pivot_df = pivot_df.reset_index()
                pivot_df = pivot_df[
                    ['ctclid', 'sym', 'expDt', 'strPrc', 'optType', 'BuyVol', 'BuyPrc', 'SellVol', 'SellPrc']]
                pivot_df['expDt'] = pivot_df['expDt'].astype(str)
                pivot_df['IntradayVolume'] = pivot_df['BuyVol'] - pivot_df['SellVol']

                # # json_data = pivot_df.to_json(orient='records')
                # # # # return pivot_df.to_dict(orient='records')
                # # # compressed_data = gzip.compress(json_data.encode('utf-8'))
                # # # return Response(content=compressed_data, media_type='application/gzip')
                # # compressed_data = gzip.compress(json_data.encode('utf-8'))
                # # return Response(content=compressed_data, media_type='application/gzip')
                # buffer = io.StringIO()
                # writer = csv.writer(buffer)
                # writer.writerow(pivot_df.columns)
                # writer.writerow(pivot_df.values)
                # with gzip.open(zip_path,'wt',encoding='utf-8',newline='') as f:
                #     f.write(buffer.getvalue())
                # return FileResponse(zip_path, media_type='application/gzip')
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    pivot_df.to_excel(writer, index=False)
                buffer.seek(0)
                with gzip.open(zip_path, 'wb') as f:
                    f.write(buffer.getvalue())
                return FileResponse(zip_path, media_type='application/gzip')
        else:
            # if not os.path.exists(zip_path) or for_dt == today:
            #     stt = datetime.now()
            #     total_rows = db.execute(text(rf'select count(*) from "{tablename}"')).scalar()
            #     page_size = 5_00_000
            #     num_pages = total_rows // page_size + (1 if (total_rows % page_size) else 0)
            #     logger.info(f'Total rows in DB: {total_rows}, Splitting into {num_pages} sheets')
            #     buffer = io.BytesIO()
            #     wb = xlsxwriter.Workbook(buffer, {'in_memory': True})
            #     for page in range(num_pages):
            #         query = f'select * from "{tablename}" limit {page_size} offset {(page) * page_size}'
            #         logger.info(query)
            #         pbar = progressbar.ProgressBar(
            #             max_value=total_rows + 1,
            #             widgets=[
            #                 progressbar.Percentage(), '',
            #                 progressbar.Bar(marker='=',left='[',right=']'),
            #                 progressbar.ETA()
            #             ]
            #         )
            #         result = db.execute(text(query))
            #         ws = wb.add_worksheet(f'Sheet{page + 1}')
            #         for col, header in enumerate(result.keys()):
            #             ws.write(0, col, header)
            #         for rn, row in enumerate(result, start=1):
            #             for col, cell in enumerate(row):
            #                 ws.write(rn, col, cell)
            #             pbar.update(rn)
            #         pbar.finish()
            #         p = 0
            #     wb.close()
            #     logger.info('fetching data from buffer')
            #     buffer.seek(0)
            #     logger.info('Writing to xlsx file and zipping . . ')
            #     with gzip.open(zip_path, 'wb') as f:
            #         f.write(buffer.getvalue())
            #     ett = datetime.now()
            #     logger.info(f'total time taken for zip_path:{(ett - stt).total_seconds()}')
            #     return FileResponse(path=zip_path, media_type='application/gzip')
            # else:
            #     return FileResponse(path=zip_path, media_type='application/gzip')
            stt = datetime.now()
            total_rows = db.execute(text(rf'select count(*) from "{tablename}"')).scalar()
            page_size = 5_00_000
            num_pages = total_rows // page_size + (1 if (total_rows % page_size) else 0)
            logger.info(f'Total rows in DB: {total_rows}, Splitting into {num_pages} sheets')
            buffer = io.BytesIO()
            wb = xlsxwriter.Workbook(buffer, {'in_memory': True})
            for page in range(num_pages):
                query = f'select * from "{tablename}" limit {page_size} offset {(page) * page_size}'
                logger.info(query)
                pbar = progressbar.ProgressBar(
                    max_value=total_rows + 1,
                    widgets=[
                        progressbar.Percentage(), '',
                        progressbar.Bar(marker='=', left='[', right=']'),
                        progressbar.ETA()
                    ]
                )
                result = db.execute(text(query))
                ws = wb.add_worksheet(f'Sheet{page + 1}')
                for col, header in enumerate(result.keys()):
                    ws.write(0, col, header)
                for rn, row in enumerate(result, start=1):
                    for col, cell in enumerate(row):
                        ws.write(rn, col, cell)
                    pbar.update(rn)
                pbar.finish()
            wb.close()
            logger.info('fetching data from buffer')
            buffer.seek(0)
            logger.info('Writing to xlsx file and zipping . . ')
            with gzip.open(zip_path, 'wb') as f:
                f.write(buffer.getvalue())
            ett = datetime.now()
            logger.info(f'total time taken for zip_path:{(ett - stt).total_seconds()}')
            return FileResponse(path=zip_path, media_type='application/gzip')


    # def get_deskwise_netposition(self, for_date:date=Query()):
    #     for_dt = pd.to_datetime(for_date).date()
    #     if for_dt == today:
    #         tablename = f'NOTIS_DESK_WISE_NET_POSITION'
    #     else:
    #         tablename = f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
    #     desk_db_df = read_data_db(for_table=tablename)
    #     json_data = desk_db_df.to_json(orient='records')
    #     return json_data

    # def get_deskwise_netposition(self, for_date:date=Query(), page:int=Query(1), page_size:int=Query(10000), db:Session=Depends(get_db)):
    #     for_dt = pd.to_datetime(for_date).date()
    #     if for_dt == today:
    #         tablename = f'NOTIS_DESK_WISE_NET_POSITION'
    #     else:
    #         tablename = f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
    #     query=text(rf'select * from "{tablename}" limit {page_size} offset {(page-1)*page_size}')
    #     result = db.execute(query).fetchall()
    #     total_rows = db.execute(text(rf'Select count(*) from "{tablename}"')).scalar()
    #     json_data = {
    #         'data':[{k: conv_str(v) for k, v in dict(row).items()} for row in result],
    #         'total_rows':total_rows,
    #         'page':page,
    #         'page_size':page_size
    #     }
    #     if not len(result):
    #         return Response(content=json.dumps(json_data), media_type='application/json')
    #     else:
    #         compressed_data = gzip.compress(json.dumps(json_data).encode('utf-8'))
    #         logger.info(f'\ntotal_rows={json_data["total_rows"]}\tpage={json_data["page"]}\tpage_size={json_data["page_size"]}\n')
    #         return Response(content=compressed_data, media_type='application/gzip')
    #         # return Response(content=json.dumps(json_data), media_type='application/json')
    #     # desk_db_df = read_data_db(for_table=tablename)
    #     # json_data = desk_db_df.to_json(orient='records')
    #     # # compressed_data = gzip.compress(json_data.encode('utf-8'))
    #     # # return Response(content=compressed_data, media_type='application/gzip')
    #     # if not len(desk_db_df):
    #     #     return Response(content=json_data, media_type='application/json')
    #     # else:
    #     #     compressed_data = gzip.compress(json_data.encode('utf-8'))
    #     #     return Response(content=compressed_data, media_type='application/gzip')

    # def get_useridwise_netposition(self, for_date:date=Query()):
    #     for_dt = pd.to_datetime(for_date).date()
    #     if for_dt == today:
    #         tablename = f'NOTIS_USERID_WISE_NET_POSITION'
    #     else:
    #         tablename = f'NOTIS_USERID_WISE_NET_POSITION_{for_dt}'
    #     desk_db_df = read_data_db(for_table=tablename)
    #     json_data = desk_db_df.to_json(orient='records')
    #     # compressed_data = gzip.compress(json_data.encode('utf-8'))
    #     # return Response(content=compressed_data, media_type='application/gzip')
    #     if not len(desk_db_df):
    #         return Response(content=json_data, media_type='application/json')
    #     else:
    #         compressed_data = gzip.compress(json_data.encode('utf-8'))
    #         return Response(content=compressed_data, media_type='application/gzip')

    # def get_nnfwise_netposition(self, for_date:date=Query()):
    #     for_dt = pd.to_datetime(for_date).date()
    #     if for_dt == today:
    #         tablename = f'NOTIS_NNF_WISE_NET_POSITION'
    #     else:
    #         tablename = f'NOTIS_NNF_WISE_NET_POSITION_{for_dt}'
    #     desk_db_df = read_data_db(for_table=tablename)
    #     json_data = desk_db_df.to_json(orient='records')
    #     # compressed_data = gzip.compress(json_data.encode('utf-8'))
    #     # return Response(content=compressed_data, media_type='application/gzip')
    #     if not len(desk_db_df):
    #         return Response(content=json_data, media_type='application/json')
    #     else:
    #         compressed_data = gzip.compress(json_data.encode('utf-8'))
    #         return Response(content=compressed_data, media_type='application/gzip')

    def get_intraday_net_position(self, for_date:date=Query()):
        for_dt = pd.to_datetime(for_date).date()
        if for_dt == today:
            tablename = f'test_net_pos_desk_{today}'
        else:
            tablename = f'NOTIS_DESK_WISE_NET_POSITION_{for_dt}'
        desk_db_df = read_data_db(for_table=tablename)
        desk_db_df.expiryDate = desk_db_df.expiryDate.astype('datetime64[ns]')
        desk_db_df.expiryDate = desk_db_df.expiryDate.dt.date
        desk_db_df.loc[desk_db_df['optionType'] == 'XX', 'strikePrice'] = 0
        desk_db_df.strikePrice = desk_db_df.strikePrice.apply(lambda x: x / 100 if x > 0 else x)
        desk_db_df.strikePrice = desk_db_df.strikePrice.astype('int64')
        grouped_desk_db_df = desk_db_df.groupby(by=['symbol', 'expiryDate', 'strikePrice', 'optionType']).agg(
            {'buyAvgQty': 'sum', 'buyAvgPrice': 'mean', 'sellAvgQty': 'sum', 'sellAvgPrice': 'mean'}).reset_index()
        grouped_desk_db_df['IntradayVolume'] = grouped_desk_db_df['buyAvgQty'] - grouped_desk_db_df['sellAvgQty']
        grouped_desk_db_df.rename(columns={'buyAvgQty': 'buyQty', 'sellAvgQty': 'sellQty'})
        grouped_desk_db_df.expiryDate = grouped_desk_db_df.expiryDate.astype(str)
        json_data = grouped_desk_db_df.to_json(orient='records')
        # compressed_data = gzip.compress(json_data.encode('utf-8'))
        # return Response(content=compressed_data, media_type='application/gzip')
        if not len(grouped_desk_db_df):
            return Response(content=json_data, media_type='application/json')
        else:
            compressed_data = gzip.compress(json_data.encode('utf-8'))
            return Response(content=compressed_data, media_type='application/gzip')

    # def get_raw_trade_data(self, for_date:date=Query()):
    #     for_dt = pd.to_datetime(for_date).date()
    #     if for_dt == today:
    #         tablename = f'notis_raw_data'
    #     else:
    #         tablename = f'notis_raw_data_{for_dt}'
    #     desk_db_df = read_data_db(for_table=tablename)
    #     json_data = desk_db_df.to_json(orient='records')
    #     # compressed_data = gzip.compress(json_data.encode('utf-8'))
    #     # return Response(content=compressed_data, media_type='application/gzip')
    #     if not len(desk_db_df):
    #         return Response(content=json_data, media_type='application/json')
    #     else:
    #         compressed_data = gzip.compress(json_data.encode('utf-8'))
    #         return Response(content=compressed_data, media_type='application/gzip')

    def get_raw_net_position(self, for_date:date=Query()):
        for_dt = pd.to_datetime(for_date).date()
        if for_dt == today:
            tablename = f'test_raw_{today}'
        else:
            tablename = f'notis_raw_data_{for_dt}'
        df = read_data_db(for_table=tablename)
        if not len(df):
            json_data = df.to_json(orient='records')
            return Response(content=json_data, media_type='application/json')
        else:
            list_str_int64 = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 16, 17, 18, 19, 21, 23, 27, 28]
            list_str_none = [15, 20, 25, 30, 31, 32, 33, 34, 35, 36, 37]
            list_none_str = [38]
            for i in list_str_int64:
                column_name = f'Column{i}'
                df[f'Column{i}'] = df[f'Column{i}'].astype('int64')
            for i in list_str_none:
                df[f'Column{i}'] = None
            for i in list_none_str:
                df[f'Column{i}'] = df[f'Column{i}'].astype('str')
            logger.info('Starting file modification...')
            df.rename(columns={
                'Column1': 'seqNo', 'Column2': 'mkt', 'Column3': 'trdNo',
                'Column4': 'trdTm', 'Column5': 'Tkn', 'Column6': 'trdQty',
                'Column7': 'trdPrc', 'Column8': 'bsFlg', 'Column9': 'ordNo',
                'Column10': 'brnCd', 'Column11': 'usrId', 'Column12': 'proCli',
                'Column13': 'cliActNo', 'Column14': 'cpCD', 'Column15': 'remarks',
                'Column16': 'actTyp', 'Column17': 'TCd', 'Column18': 'ordTm',
                'Column19': 'Booktype', 'Column20': 'oppTmCd', 'Column21': 'ctclid',
                'Column22': 'status', 'Column23': 'TmCd', 'Column24': 'sym',
                'Column25': 'ser', 'Column26': 'inst', 'Column27': 'expDt',
                'Column28': 'strPrc', 'Column29': 'optType', 'Column30': 'sessionID',
                'Column31': 'echoback', 'Column32': 'Fill1', 'Column33': 'Fill2',
                'Column34': 'Fill3', 'Column35': 'Fill4', 'Column36': 'Fill5', 'Column37': 'Fill6'
            }, inplace=True)
            df['ordTm'] = df['ordTm'].apply(lambda x: get_date_from_non_jiffy(int(x)))
            df['expDt'] = df['expDt'].apply(lambda x: get_date_from_non_jiffy(int(x)))
            df.expDt = df.expDt.astype('datetime64[ns]').dt.date
            df['trdTm'] = df['trdTm'].apply(lambda x: get_date_from_jiffy(int(x)))
            df['bsFlg'] = np.where(df['bsFlg'] == 1, 'B', 'S')
            pivot_df = df.pivot_table(
                index=['ctclid', 'sym', 'expDt', 'strPrc', 'optType'],
                columns='bsFlg',
                values=['trdQty', 'trdPrc'],
                aggfunc={'trdQty': 'sum', 'trdPrc': 'mean'},
                fill_value=0
            )
            pivot_df.columns = ['BuyPrc', 'SellPrc', 'BuyVol', 'SellVol']
            pivot_df = pivot_df.reset_index()
            pivot_df = pivot_df[['ctclid', 'sym', 'expDt', 'strPrc', 'optType', 'BuyVol', 'BuyPrc', 'SellVol', 'SellPrc']]
            pivot_df['expDt'] = pivot_df['expDt'].astype(str)
            pivot_df['IntradayVolume'] = pivot_df['BuyVol'] - pivot_df['SellVol']

            json_data = pivot_df.to_json(orient='records')
            # # return pivot_df.to_dict(orient='records')
            # compressed_data = gzip.compress(json_data.encode('utf-8'))
            # return Response(content=compressed_data, media_type='application/gzip')
            compressed_data = gzip.compress(json_data.encode('utf-8'))
            return Response(content=compressed_data, media_type='application/gzip')

    def get_exposure(self, for_date:date=Query()):
        volt_df = read_file(os.path.join(volt_dir, f'FOVOLT_{yesterday.strftime("%d%m%Y")}.csv'))
        volt_df.columns = [re.sub(r'\s', '', each) for each in volt_df.columns]
        volt_df = volt_df.iloc[:, 1:3]
        volt_df.rename(columns={'UnderlyingClosePrice(A)': 'SpotClosePrice'}, inplace=True)
        sym_list = ['NIFTY', 'BANKNIFTY', 'FINNIFTY', 'MIDCPNIFTY']
        volt_df = volt_df.query("Symbol in @sym_list")

        tablename = f'test_cp_noncp_{today}'
        cp_df = read_data_db(for_table=tablename)
        cp_df.columns = [re.sub(r'Eod|\s', '', each) for each in cp_df.columns]

        merged_df = cp_df.merge(volt_df, how='left', left_on=['Underlying'], right_on=['Symbol'])
        merged_df.drop(columns=['Symbol'], inplace=True)
        merged_df = merged_df.query("OptionType == 'CE' or OptionType == 'PE'")
        merged_df.drop_duplicates(inplace=True)

        pivot_df = merged_df.pivot_table(
            index=['Broker', 'Underlying', 'SpotClosePrice'],
            columns=['OptionType'],
            values=['FinalNetQty'],
            aggfunc={'FinalNetQty': 'sum'},
            fill_value=0
        )
        pivot_df.columns = ['CE', 'PE']
        pivot_df.reset_index(inplace=True)
        pivot_df.SpotClosePrice = pivot_df.SpotClosePrice.astype('float64')
        pivot_df['NetQty'] = pivot_df['CE'] - pivot_df['PE']
        pivot_df['Exposure(in Crs)'] = (pivot_df['NetQty'] * pivot_df['SpotClosePrice']) / 10000000
        print(f'exposure table shape: {pivot_df.shape}')
        # list_to_int = ['SpotClosePrice','CE','PE','NetQty','Exposure(in Crs)']
        # for each in list_to_int:
        #     pivot_df[each] = pivot_df[each].astype(str)
        json_data = pivot_df.to_json(orient='records')
        if not pivot_df.empty:
            return Response(content=json_data, media_type='application/json')

service = ServiceApp()
app = service.app

if __name__ == '__main__':
    uvicorn.run('temp_app:app', host='172.16.47.81', port=8871, workers=6)

